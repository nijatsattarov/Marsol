from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import re
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any, Union
import uuid
import io
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import shutil

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Email notification service (Resend)
from email_service import notify as _email_notify  # noqa: E402
import email_service  # noqa: E402
import asyncio  # noqa: E402

# Cloudinary upload service
from cloudinary_service import upload_file as _cl_upload, delete_asset as _cl_delete  # noqa: E402
from invitation_card import render_invitation_png  # noqa: E402

# Firebase Cloud Messaging (push notifications)
from push_service import init_firebase as _init_fcm, push_to_users as _push_to_users  # noqa: E402
_init_fcm()


def _safe_push(recipient_names, title: str, body: str, link: Optional[str] = None, data: Optional[dict] = None):
    """Fire-and-forget push to users. Never raises — DB notification is the
    primary record; push is best-effort and survives FCM outages.
    """
    async def _runner():
        try:
            await _push_to_users(db, recipient_names, title, body, data=data, link=link)
        except Exception as exc:
            logging.error("push notification failed: %s", exc)
    try:
        asyncio.create_task(_runner())
    except RuntimeError:
        # Not in an event loop (rare path). Skip silently.
        pass

# Default invitation message templates per event type. Each template supports
# placeholders {guest_name}, {event_name}, {event_date}, {event_time},
# {event_location}. Newlines split into separate lines on the rendered card.
DEFAULT_INVITATION_TEMPLATES = {
    "Breakfast": (
        "Hörmətli {guest_name},\n"
        "sizi {event_date} tarixində baş tutacaq\n"
        '"{event_name}" işgüzar səhər yeməyinə dəvət edirik.'
    ),
    "B2B görüş": (
        "Hörmətli {guest_name},\n"
        "sizi region partnyorları ilə baş tutacaq\n"
        '"{event_name}" B2B görüşünə dəvət edirik.'
    ),
    "Sosial fəaliyyət": (
        "Hörmətli {guest_name},\n"
        "sizi keçiriləcək\n"
        '"{event_name}" sosial fəaliyyətinə dəvət edirik.'
    ),
    "Mafia": (
        "Hörmətli {guest_name},\n"
        "sizi keçiriləcək\n"
        '"{event_name}" Mafia oyununa dəvət edirik.'
    ),
    "Təlim": (
        "Hörmətli {guest_name},\n"
        "sizi keçiriləcək\n"
        '"{event_name}" təliminə dəvət edirik.'
    ),
    "Ofis ziyarəti": (
        "Hörmətli {guest_name},\n"
        "sizi keçiriləcək\n"
        '"{event_name}" ofis ziyarətinə dəvət edirik.'
    ),
    "default": (
        "Hörmətli {guest_name},\n"
        "sizi keçiriləcək\n"
        '"{event_name}" tədbirinə dəvət edirik.'
    ),
}


async def _get_invitation_template(event_type: str) -> str:
    """Look up template from db.invitation_templates falling back to defaults."""
    doc = None
    if event_type:
        doc = await db.invitation_templates.find_one({"event_type": event_type}, {"_id": 0})
    if doc and doc.get("body"):
        return doc["body"]
    return DEFAULT_INVITATION_TEMPLATES.get(event_type) or DEFAULT_INVITATION_TEMPLATES["default"]

# SMS service (LSIM Quick SMS)
import sms_service  # noqa: E402

# Mailchimp Marketing service
import mailchimp_service  # noqa: E402

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL')
if not mongo_url:
    raise RuntimeError("MONGO_URL environment variable is not set")
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'marsol_db')]

# JWT Settings
SECRET_KEY = os.environ.get('SECRET_KEY', 'marsol-secret-key-2024')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 90  # 90 days — long-lived session for PWAs

security = HTTPBearer()

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Static files for uploads
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

# ==================== MODELS ====================

class UserLogin(BaseModel):
    email: str
    password: str

class UserCreate(BaseModel):
    email: str
    password: str
    name: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: dict

# Company (Şirkət) Models
class CompanyCreate(BaseModel):
    brand_name: str
    legal_name: Optional[str] = ""
    sector: str
    company_size: str  # Böyük, Orta, Kiçik, Mikro
    registration_date: Optional[str] = ""
    address: Optional[str] = ""
    bank_details: Optional[str] = ""
    # Owner info
    owner_name: str
    owner_phone: str
    owner_email: Optional[str] = ""
    owner_social_links: Optional[str] = ""
    # Co-founders
    co_founders: Optional[List[str]] = []
    # Representative
    representative_name: Optional[str] = ""
    representative_phone: Optional[str] = ""
    representative_email: Optional[str] = ""
    # Company contacts
    company_phone: Optional[str] = ""
    company_website: Optional[str] = ""
    company_social_links: Optional[str] = ""
    # Children info (owner's)
    children_count: Optional[int] = 0
    children_info: Optional[List[Dict]] = []
    # Reference
    reference_source: Optional[str] = ""  # Media, Partnyor, Digər
    reference_person: Optional[str] = ""
    reference_company: Optional[str] = ""
    # Marsol
    marsol_representative: str
    # Project/Package
    joined_project: str  # Üzvlük, Sərgi, Təlim, ICMA, Sosial layihə
    package: str  # Premium, Business, Business Plus
    # Contract
    contract_start_date: Optional[str] = ""
    contract_end_date: Optional[str] = ""
    contract_file: Optional[str] = ""
    # Payment
    total_amount: Optional[float] = 0
    paid_amount: Optional[float] = 0
    debt_amount: Optional[float] = 0
    last_payment_date: Optional[str] = ""
    payment_due_date: Optional[str] = ""
    # Status
    status: Optional[str] = "Aktiv"

class CompanyUpdate(BaseModel):
    brand_name: Optional[str] = None
    legal_name: Optional[str] = None
    sector: Optional[str] = None
    company_size: Optional[str] = None
    registration_date: Optional[str] = None
    address: Optional[str] = None
    bank_details: Optional[str] = None
    owner_name: Optional[str] = None
    owner_phone: Optional[str] = None
    owner_email: Optional[str] = None
    owner_social_links: Optional[str] = None
    co_founders: Optional[List[str]] = None
    representative_name: Optional[str] = None
    representative_phone: Optional[str] = None
    representative_email: Optional[str] = None
    company_phone: Optional[str] = None
    company_website: Optional[str] = None
    company_social_links: Optional[str] = None
    children_count: Optional[int] = None
    children_info: Optional[List[Dict]] = None
    reference_source: Optional[str] = None
    reference_person: Optional[str] = None
    reference_company: Optional[str] = None
    marsol_representative: Optional[str] = None
    joined_project: Optional[str] = None
    package: Optional[str] = None
    contract_start_date: Optional[str] = None
    contract_end_date: Optional[str] = None
    contract_file: Optional[str] = None
    total_amount: Optional[float] = None
    paid_amount: Optional[float] = None
    debt_amount: Optional[float] = None
    last_payment_date: Optional[str] = None
    payment_due_date: Optional[str] = None
    status: Optional[str] = None

# Employee Model
class EmployeeCreate(BaseModel):
    photo: Optional[str] = ""
    full_name: str
    father_name: Optional[str] = ""
    birth_date: Optional[str] = ""
    gender: str
    id_card_number: Optional[str] = ""
    fin_code: Optional[str] = ""
    education_level: Optional[str] = ""
    education_institution: Optional[str] = ""
    marital_status: Optional[str] = ""
    children_count: Optional[int] = 0
    registration_address: Optional[str] = ""
    actual_address: Optional[str] = ""
    company_phone: Optional[str] = ""
    personal_phone: str
    email: str
    emergency_contact_name: Optional[str] = ""
    emergency_contact_relation: Optional[str] = ""
    emergency_contact_phone: Optional[str] = ""
    # Contract
    department: str
    position: str
    contract_start_date: Optional[str] = ""
    work_start_date: Optional[str] = ""
    contract_end_date: Optional[str] = ""
    probation_end_date: Optional[str] = ""
    main_vacation_days: Optional[int] = 21
    additional_vacation_days: Optional[int] = 0
    gross_salary: Optional[float] = 0
    net_salary: Optional[float] = 0
    work_schedule: Optional[str] = ""
    status: Optional[str] = "Aktiv"

# Finance Models
class IncomeCreate(BaseModel):
    company_id: str
    company_name: str
    owner_name: str
    marsol_representative: str
    project: str
    package: str
    amount: float
    paid_amount: Optional[float] = 0
    debt_amount: Optional[float] = 0
    currency: Optional[str] = "AZN"
    contract_start_date: Optional[str] = ""
    contract_end_date: Optional[str] = ""
    payment_method: Optional[str] = ""
    marsol_company: Optional[str] = ""

class ExpenseCreate(BaseModel):
    expense_name: str
    category: str
    sub_category: Optional[str] = ""
    amount: float
    currency: Optional[str] = "AZN"
    date: str
    project: Optional[str] = ""
    department: Optional[str] = ""
    responsible_person: Optional[str] = ""
    payment_type: Optional[str] = ""
    payment_method: Optional[str] = ""
    status: Optional[str] = "Ödənilib"
    marsol_company: Optional[str] = ""

# Task Model
class TaskCreate(BaseModel):
    task_name: str
    department: Optional[str] = ""
    assignee: Optional[Union[str, List[str]]] = ""
    responsible_person: Optional[Union[str, List[str]]] = ""
    priority: str  # Yüksək, Orta, Aşağı
    difficulty: Optional[str] = ""  # Çətin / Orta / Asan
    estimated_duration: Optional[str] = ""  # "2 saat", "3 gün", etc.
    start_date: Optional[str] = ""
    end_date: Optional[str] = ""
    related_object_type: Optional[str] = ""
    related_object_id: Optional[str] = ""
    related_object: Optional[str] = ""
    phase: Optional[str] = ""
    status: Optional[str] = "Gözləyir"
    notes: Optional[str] = ""
    subtasks: Optional[List[dict]] = []  # [{title, done}, ...]

# Meeting Model  
class MeetingCreate(BaseModel):
    employee: str
    meeting_setter: str
    date: str
    time: str
    company: Optional[str] = ""
    contact_person: Optional[str] = ""
    project: Optional[str] = ""
    meeting_type: str
    meeting_mode: Optional[str] = "Offline"
    department: Optional[str] = ""
    location: Optional[str] = ""
    result: Optional[str] = ""
    next_meeting: Optional[str] = ""
    notes: Optional[str] = ""
    reminders: Optional[list] = []

# ==================== AUTH FUNCTIONS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ==================== RBAC HELPER ====================

MODULES = [
    "dashboard", "companies", "hr", "sales", "members", "obligations",
    "finance", "organization", "meetings", "assembly", "tasks",
    "marketing", "projects", "reports", "messages", "files", "notes",
    "settings", "notifications", "sms"
]

async def get_user_permissions(user: dict) -> dict:
    """Get merged permissions for a user based on their role"""
    if user.get("role") == "admin":
        return {m: "write" for m in MODULES}
    role_name = user.get("role", "")
    if role_name:
        role = await db.roles.find_one({"name": role_name}, {"_id": 0})
        if role:
            return role.get("permissions", {})
    return {m: "read" for m in MODULES}

def check_permission(module: str, level: str = "read"):
    """Dependency: check if current user has required permission for a module"""
    async def checker(current_user: dict = Depends(get_current_user)):
        if current_user.get("role") == "admin":
            return current_user
        perms = await get_user_permissions(current_user)
        user_level = perms.get(module, "none")
        if level == "read" and user_level in ("read", "write"):
            return current_user
        if level == "write" and user_level == "write":
            return current_user
        raise HTTPException(status_code=403, detail="Bu əməliyyat üçün icazəniz yoxdur")
    return checker

# ==================== SCOPE (RECORD-LEVEL VISIBILITY) HELPER ====================
# Each module defines which stored fields identify "ownership" of a record.
# If a role has scopes[module] == "own", only records where one of these
# fields equals the user's name are visible / editable. Default is "all".
SCOPE_FIELDS = {
    "members": ["curator", "created_by"],
    "companies": ["curator", "created_by"],
    "tasks": ["assignee", "responsible_person", "created_by"],
    "meetings": ["employee", "meeting_setter", "created_by", "participant_names"],
    "sales": ["curator", "created_by"],
    "projects": ["created_by"],
    "assembly": ["created_by", "curator"],
    "files": ["uploaded_by", "owner"],
    "notes": ["created_by", "shared_with_users"],
    "organization": ["created_by"],
}

# Modules that participate in Müəssisə (multi-tenant) isolation. Each entry maps
# to the corresponding MongoDB collection field used as the tenant marker.
TENANT_MODULES = {
    "members", "companies", "tasks", "meetings", "sales",
    "projects", "assembly", "files", "notes", "organization",
}


def _user_tenant(user: dict) -> Optional[str]:
    """Return the active müəssisə name for `user`, or None if no filter should be
    applied. Admin users are always cross-tenant (return None)."""
    if not user or user.get("role") == "admin":
        return None
    val = (user.get("marsol_company") or "").strip()
    return val or None


def _merge_filter(query: dict, extra: dict) -> dict:
    """Safely merge `extra` filter clauses into `query` without clobbering keys."""
    if not extra:
        return query
    if not query:
        return dict(extra)
    out = dict(query)
    for k, v in extra.items():
        if k not in out:
            out[k] = v
        else:
            # key collision → wrap into $and
            out = {"$and": [out, {k: v}]}
            break
    return out


def _as_name_list(value) -> list:
    """Normalise a value that may be a string or a list/tuple into a clean list of trimmed names."""
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(v).strip() for v in value if v and str(v).strip()]
    s = str(value).strip()
    return [s] if s else []

async def get_user_scopes(user: dict) -> dict:
    if user.get("role") == "admin":
        return {}
    role = await db.roles.find_one({"name": user.get("role", "")}, {"_id": 0})
    if not role:
        return {}
    return role.get("scopes", {}) or {}

async def apply_scope(query: dict, user: dict, module: str) -> dict:
    """Merge scope-based ownership filter into query. Returns new query dict.

    Scope values:
        - "all"        → no filter (everyone visible)
        - "own"        → only records the user is involved in (assignee/curator/...)
        - "department" → records owned by users in the same department as `user`

    Additionally, when the module participates in multi-tenant (Müəssisə)
    isolation and the user has a `marsol_company` assigned, records are
    filtered to that tenant. Admin users bypass both scope and tenant filters.
    """
    if user.get("role") == "admin":
        return query
    scopes = await get_user_scopes(user)
    # Default scope per module is "own" for non-admin users when the role
    # hasn't explicitly configured it. This prevents legacy roles (no scopes
    # field) from accidentally exposing data system-wide. Admins are
    # short-circuited above.
    scope = scopes.get(module, "own")
    fields = SCOPE_FIELDS.get(module, [])
    user_name = user.get("name", "")
    tenant = _user_tenant(user)
    tenant_active = bool(tenant and module in TENANT_MODULES)

    # Build "personal involvement" clauses — records where the user is
    # explicitly named (assignee/curator/created_by/…). These records remain
    # visible to the user EVEN IF the record sits on a different müəssisə.
    own_clauses = [{f: user_name} for f in fields] if (fields and user_name) else []

    if scope == "own":
        # Pure ownership — tenant filter is implicit (user is named ⇒ relevant).
        if not own_clauses:
            return query
        if "$or" in query or "$and" in query:
            return {"$and": [query, {"$or": own_clauses}]}
        new_query = dict(query)
        new_query["$or"] = own_clauses
        return new_query

    if scope == "department":
        my_dept = (user.get("department") or "").strip()
        dept_clauses = []
        if my_dept and fields:
            dept_users = await db.users.find({"department": my_dept}, {"_id": 0, "name": 1}).to_list(500)
            names = [u.get("name") for u in dept_users if u.get("name")]
            if user_name and user_name not in names:
                names.append(user_name)
            if names:
                dept_clauses = [{f: {"$in": names}} for f in fields]
        # Compose: (department records restricted by tenant)  OR  (own records cross-tenant)
        outer = []
        if dept_clauses:
            inner = {"$or": dept_clauses}
            if tenant_active:
                inner = {"$and": [{"marsol_company": tenant}, inner]}
            outer.append(inner)
        # Add own_clauses individually so $or stays flat (cheap to index)
        outer.extend(own_clauses)
        if not outer:
            return query
        clause = {"$or": outer} if len(outer) > 1 else outer[0]
        if "$or" in query or "$and" in query:
            return {"$and": [query, clause]}
        new_query = dict(query)
        new_query.update(clause if "$or" not in clause else {"$or": clause["$or"]})
        if "$and" in clause:
            new_query["$and"] = clause["$and"]
        return new_query

    # scope == "all"
    # See everyone in my müəssisə + my personal cross-tenant records.
    if not tenant_active:
        return query
    outer = [{"marsol_company": tenant}] + own_clauses
    clause = {"$or": outer} if len(outer) > 1 else {"marsol_company": tenant}
    if "$or" in query or "$and" in query:
        return {"$and": [query, clause]}
    new_query = dict(query)
    new_query.update(clause)
    return new_query


async def assert_scope_ownership(user: dict, module: str, record: Optional[dict]):
    """Raise 403 if user has 'own'/'department' scope and record is not in their bucket."""
    if user.get("role") == "admin" or not record:
        return

    fields = SCOPE_FIELDS.get(module, [])
    user_name = user.get("name", "")

    def _field_matches(val) -> bool:
        if isinstance(val, (list, tuple, set)):
            return user_name in val
        return val == user_name

    # User is explicitly named in the record → ALWAYS allowed (their own item,
    # even if it sits on a different müəssisə — cross-tenant assignment case)
    is_personal = any(_field_matches(record.get(f)) for f in fields) if fields else False
    if is_personal:
        return

    # Tenant isolation — applied only for non-personal records
    tenant = _user_tenant(user)
    if tenant and module in TENANT_MODULES:
        rec_tenant = (record.get("marsol_company") or "").strip()
        if rec_tenant and rec_tenant != tenant:
            raise HTTPException(status_code=403, detail="Bu qeyd başqa müəssisəyə aiddir")

    scopes = await get_user_scopes(user)
    scope = scopes.get(module, "own")
    if scope == "all":
        return
    if scope == "own":
        if any(_field_matches(record.get(f)) for f in fields):
            return
        raise HTTPException(status_code=403, detail="Bu qeyd sizə aid deyil")
    if scope == "department":
        my_dept = (user.get("department") or "").strip()
        if not my_dept:
            if any(_field_matches(record.get(f)) for f in fields):
                return
            raise HTTPException(status_code=403, detail="Bu qeyd sizə aid deyil")
        dept_users = await db.users.find({"department": my_dept}, {"_id": 0, "name": 1}).to_list(500)
        names = {u.get("name") for u in dept_users if u.get("name")}
        names.add(user_name)
        for f in fields:
            v = record.get(f)
            if isinstance(v, (list, tuple, set)):
                if any(n in names for n in v):
                    return
            elif v in names:
                return
        raise HTTPException(status_code=403, detail="Bu qeyd şöbənizə aid deyil")


async def _user_email_by_name(name: str) -> Optional[str]:
    """Look up a system user's email by display name."""
    if not name:
        return None
    user = await db.users.find_one({"name": name}, {"_id": 0, "email": 1})
    return user.get("email") if user else None


async def _email_notify_safe(**kwargs) -> bool:
    """Wrapper that logs but never raises so notifications never break the API.
    Returns True if at least one recipient received the email."""
    try:
        from email_service import notify as _en
        return bool(await _en(**kwargs))
    except Exception as e:
        logging.error("Email notify failed: %s", e)
        return False

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "email": user_data.email,
        "name": user_data.name,
        "password": hash_password(user_data.password),
        "role": "admin",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    access_token = create_access_token({"sub": user_id})
    return Token(
        access_token=access_token,
        token_type="bearer",
        user={"id": user_id, "email": user_data.email, "name": user_data.name, "role": "admin"}
    )

@api_router.post("/auth/login", response_model=Token)
async def login(user_data: UserLogin):
    user = await db.users.find_one({"email": user_data.email})
    if not user or not verify_password(user_data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Yanlış email və ya şifrə")
    # Block inactive / suspended users from logging in
    user_status = (user.get("status") or "Aktiv").strip()
    if user_status and user_status.lower() not in ("aktiv", "active"):
        raise HTTPException(status_code=403, detail="Sizin hesab deaktiv edilmişdir")
    
    access_token = create_access_token({"sub": user["id"]})
    user_dict = {"id": user["id"], "email": user["email"], "name": user["name"], "role": user["role"]}
    perms = await get_user_permissions(user)
    user_dict["permissions"] = perms
    # Record system session (for Davamiyyət — Sistem fəaliyyəti)
    try:
        now = datetime.now(timezone.utc)
        # Auto-close any previous still-open session(s) for this user. The
        # logout_at is pinned to the LAST heartbeat (last_active_at) — not
        # to "now" — so an abandoned tab from days ago doesn't suddenly count
        # the intervening idle time.
        async for prev in db.user_sessions.find({"user_id": user["id"], "logout_at": None}):
            close_ts = prev.get("last_active_at") or prev.get("login_at") or now.isoformat()
            await db.user_sessions.update_one(
                {"_id": prev["_id"]},
                {"$set": {"logout_at": close_ts, "auto_closed": True}},
            )
        await db.user_sessions.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "user_email": user["email"],
            "user_name": user["name"],
            "login_at": now.isoformat(),
            "last_active_at": now.isoformat(),
            "logout_at": None,
        })
    except Exception as e:
        logging.error("Session record failed: %s", e)
    return Token(
        access_token=access_token,
        token_type="bearer",
        user=user_dict
    )

@api_router.post("/auth/logout")
async def logout(current_user: dict = Depends(get_current_user)):
    """Mark the user's most recent open session as closed."""
    now = datetime.now(timezone.utc).isoformat()
    await db.user_sessions.update_one(
        {"user_id": current_user["id"], "logout_at": None},
        {"$set": {"logout_at": now, "last_active_at": now}},
        upsert=False,
    )
    return {"ok": True}

@api_router.post("/auth/heartbeat")
async def heartbeat(current_user: dict = Depends(get_current_user), request: Request = None):
    """Bumps last_active_at on the latest open session so we can compute
    accurate active duration even when the user closes the tab without logout.

    Also performs a SLIDING-WINDOW token refresh: when the caller's JWT has
    less than half of its lifetime remaining, we mint a fresh long-lived
    token. The frontend swaps it into localStorage so PWAs effectively stay
    logged in indefinitely as long as the user opens the app at least once
    every 45 days (half of the 90-day window).
    """
    now = datetime.now(timezone.utc).isoformat()
    await db.user_sessions.update_one(
        {"user_id": current_user["id"], "logout_at": None},
        {"$set": {"last_active_at": now}},
        upsert=False,
    )
    new_token: Optional[str] = None
    try:
        auth_header = request.headers.get("Authorization", "") if request else ""
        if auth_header.lower().startswith("bearer "):
            tok = auth_header.split(None, 1)[1]
            decoded = jwt.decode(tok, SECRET_KEY, algorithms=[ALGORITHM])
            exp_ts = decoded.get("exp")
            if exp_ts:
                # Refresh once the token's remaining lifetime drops below 50%
                # of the full window.
                from datetime import datetime as _dt, timezone as _tz
                remaining = exp_ts - _dt.now(_tz.utc).timestamp()
                full = ACCESS_TOKEN_EXPIRE_MINUTES * 60
                if remaining < full * 0.5:
                    new_token = create_access_token({"sub": current_user["id"], "email": current_user["email"]})
    except Exception:
        # Heartbeat must never 5xx — token refresh is best effort.
        pass
    resp: Dict[str, Any] = {"ok": True, "ts": now}
    if new_token:
        resp["access_token"] = new_token
    return resp

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return current_user

# ==================== PUSH NOTIFICATIONS (FCM) ====================

@api_router.post("/push/subscribe")
async def push_subscribe(payload: dict, current_user: dict = Depends(get_current_user)):
    """Persist an FCM device token for the current user (multi-device aware).

    Body: { token, platform? } — token from firebase/messaging getToken().
    Idempotent: upsert keyed on token (one device = one row).
    """
    token = (payload or {}).get("token", "").strip()
    if not token:
        raise HTTPException(status_code=400, detail="token missing")
    now = datetime.now(timezone.utc).isoformat()
    await db.push_tokens.update_one(
        {"token": token},
        {
            "$set": {
                "user_id": current_user["id"],
                "user_email": current_user.get("email", ""),
                "user_name": current_user.get("name", ""),
                "platform": (payload.get("platform") or "web")[:32],
                "last_used_at": now,
            },
            "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": now},
        },
        upsert=True,
    )
    return {"ok": True}


@api_router.post("/push/unsubscribe")
async def push_unsubscribe(payload: dict, current_user: dict = Depends(get_current_user)):
    token = (payload or {}).get("token", "").strip()
    if token:
        await db.push_tokens.delete_one({"token": token, "user_id": current_user["id"]})
    else:
        # Remove all tokens for this user
        await db.push_tokens.delete_many({"user_id": current_user["id"]})
    return {"ok": True}


@api_router.post("/push/test")
async def push_test(payload: dict = None, current_user: dict = Depends(get_current_user)):
    """Send a sample push to the calling user's own devices — Settings UI uses this.

    On failure the response includes the FCM-reported reason so the UI can
    actually display *why* a push didn't go through (instead of "uğursuz: 1").
    """
    title = (payload or {}).get("title") or "Marsol MMS — test bildirişi"
    body = (payload or {}).get("body") or f"Salam, {current_user.get('name', '')}! Push bildirişləri aktivdir."
    me = current_user.get("name", "")
    rows = await db.push_tokens.find({"user_name": me}, {"_id": 0, "token": 1}).to_list(50)
    tokens = list({r["token"] for r in rows if r.get("token")})
    if not tokens:
        return {"ok": True, "result": {"success": 0, "failure": 0, "invalid": [], "no_tokens": True}}
    from push_service import send_push as _send_push
    result = await _send_push(db, tokens, title, body, link="/dashboard")
    # After auto-pruning invalid tokens, tell the UI how many devices are left so
    # the user knows whether they need to re-subscribe.
    remaining = await db.push_tokens.count_documents({"user_name": me})
    result["remaining_devices"] = remaining
    return {"ok": True, "result": result}


@api_router.get("/push/status")
async def push_status(current_user: dict = Depends(get_current_user)):
    """Returns device count + last_used per device for the current user."""
    rows = await db.push_tokens.find(
        {"user_id": current_user["id"]}, {"_id": 0, "platform": 1, "last_used_at": 1, "created_at": 1}
    ).to_list(50)
    return {"devices": rows, "count": len(rows)}

# ==================== DASHBOARD ====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(check_permission("dashboard", "read"))):
    # Get real counts from database
    companies_count = await db.companies.count_documents({})
    employees_count = await db.employees.count_documents({})
    # Dashboard 'tasks' widget: admin sees TOTAL across all users; non-admin
    # sees only their own workload (tasks they are assignee / responsible /
    # created for).
    is_admin_dash = (current_user.get("role") or "").lower() == "admin"
    me_name = current_user.get("name", "")
    if is_admin_dash:
        me_tasks_query: Dict[str, Any] = {}
    else:
        me_clauses = [
            {"assignee": me_name},
            {"responsible_person": me_name},
            {"created_by": me_name},
        ] if me_name else [{}]
        me_tasks_query = {"$or": me_clauses}
    tasks_count = await db.tasks.count_documents(me_tasks_query)
    # Meetings count — admin sees total; non-admin sees only meetings scoped to them
    # (assignee/setter/created_by/participant via apply_scope).
    today_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if is_admin_dash:
        base_meetings_q: Dict[str, Any] = {}
    else:
        base_meetings_q = await apply_scope({}, current_user, "meetings")
    meetings_count = await db.meetings.count_documents(base_meetings_q)
    upcoming_q = {"$and": [base_meetings_q, {"date": {"$gt": today_iso}}]} if base_meetings_q else {"date": {"$gt": today_iso}}
    today_q = {"$and": [base_meetings_q, {"date": today_iso}]} if base_meetings_q else {"date": today_iso}
    past_q = {"$and": [base_meetings_q, {"date": {"$lt": today_iso}}]} if base_meetings_q else {"date": {"$lt": today_iso}}
    meetings_upcoming = await db.meetings.count_documents(upcoming_q)
    meetings_today = await db.meetings.count_documents(today_q)
    meetings_past = await db.meetings.count_documents(past_q)
    
    # Events stats
    events_count = await db.events.count_documents({})
    invitations_count = await db.invitations.count_documents({})
    events_type_pipeline = [
        {"$group": {"_id": "$event_type", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    events_by_type = await db.events.aggregate(events_type_pipeline).to_list(20)
    
    # Invitations stats
    inv_attended = await db.invitations.count_documents({"participation_status": "Qatılır"})
    inv_declined = await db.invitations.count_documents({"participation_status": "Qatılmır"})
    inv_no_answer = await db.invitations.count_documents({"call_status": "Cavab vermədi"})
    inv_pending = await db.invitations.count_documents({"call_status": "Gözləyir"})
    
    # Invitations by event type
    inv_type_pipeline = [
        {"$group": {"_id": "$event_type", "total": {"$sum": 1},
                     "attended": {"$sum": {"$cond": [{"$eq": ["$participation_status", "Qatılır"]}, 1, 0]}},
                     "declined": {"$sum": {"$cond": [{"$eq": ["$participation_status", "Qatılmır"]}, 1, 0]}}}},
        {"$sort": {"total": -1}}
    ]
    inv_by_type = await db.invitations.aggregate(inv_type_pipeline).to_list(20)
    
    # Get companies by package — dynamically based on configured packages
    pkg_palette = ["#3D4F6F", "#9ACD32", "#F59E0B", "#8B5CF6", "#EF4444", "#06B6D4", "#EC4899", "#10B981", "#F97316"]
    package_docs = await db.packages.find({}, {"_id": 0, "name": 1}).to_list(100)
    package_names = [p["name"] for p in package_docs if p.get("name")]
    if not package_names:
        package_names = ["Premium", "Business", "Business Plus"]
    package_breakdown = []
    for idx, pname in enumerate(package_names):
        cnt = await db.companies.count_documents({"package": pname})
        package_breakdown.append({"name": f"{pname} paket", "count": cnt, "color": pkg_palette[idx % len(pkg_palette)]})
    other_count = await db.companies.count_documents({"$or": [
        {"package": {"$nin": package_names + [""]}},
        {"package": {"$exists": False}},
        {"package": ""},
    ]})
    if other_count > 0:
        package_breakdown.append({"name": "Digər / paket yoxdur", "count": other_count, "color": "#CBD5E1"})

    # Get companies by sector
    sectors_pipeline = [
        {"$group": {"_id": "$sector", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    sectors_data = await db.companies.aggregate(sectors_pipeline).to_list(20)
    
    # Calculate financials from companies
    company_finance_pipeline = [
        {"$group": {"_id": None, "total": {"$sum": "$total_amount"}, "paid": {"$sum": "$paid_amount"}, "debt": {"$sum": "$debt_amount"}}}
    ]
    company_finance = await db.companies.aggregate(company_finance_pipeline).to_list(1)
    
    expense_pipeline = [
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    expense_data = await db.expenses.aggregate(expense_pipeline).to_list(1)
    
    total_income = company_finance[0]["total"] if company_finance else 0
    total_paid = company_finance[0]["paid"] if company_finance else 0
    total_debt = company_finance[0]["debt"] if company_finance else 0
    total_expenses = expense_data[0]["total"] if expense_data else 0
    
    # Build sector breakdown with distinct colors (high-contrast palette)
    colors = ["#3D4F6F", "#9ACD32", "#F59E0B", "#8B5CF6", "#EF4444", "#06B6D4", "#EC4899", "#10B981", "#F97316", "#6366F1", "#14B8A6", "#84CC16"]
    sector_breakdown = [
        {"name": s["_id"] or "Digər", "count": s["count"], "color": colors[i % len(colors)]}
        for i, s in enumerate(sectors_data)
    ]
    
    return {
        "companies": {
            "total": companies_count,
            "breakdown": package_breakdown,
        },
        "employees": {
            "total": employees_count
        },
        "tasks": {
            "total": tasks_count,
            "pending": await db.tasks.count_documents({"$and": [me_tasks_query, {"status": "Gözləyir"}]}) if me_tasks_query else await db.tasks.count_documents({"status": "Gözləyir"}),
            "in_progress": await db.tasks.count_documents({"$and": [me_tasks_query, {"status": "İcrada"}]}) if me_tasks_query else await db.tasks.count_documents({"status": "İcrada"}),
            "completed": await db.tasks.count_documents({"$and": [me_tasks_query, {"status": "Tamamlandı"}]}) if me_tasks_query else await db.tasks.count_documents({"status": "Tamamlandı"}),
            "cancelled": await db.tasks.count_documents({"$and": [me_tasks_query, {"status": "Ləğv edildi"}]}) if me_tasks_query else await db.tasks.count_documents({"status": "Ləğv edildi"})
        },
        "meetings": {
            "total": meetings_count,
            "upcoming": meetings_upcoming,
            "today": meetings_today,
            "past": meetings_past,
        },
        "events": {
            "total": events_count,
            "by_type": [{"name": e["_id"] or "Digər", "count": e["count"]} for e in events_by_type],
        },
        "invitations": {
            "total": invitations_count,
            "attended": inv_attended,
            "declined": inv_declined,
            "no_answer": inv_no_answer,
            "pending": inv_pending,
            "by_type": [{"name": i["_id"] or "Digər", "total": i["total"], "attended": i["attended"], "declined": i["declined"]} for i in inv_by_type],
        },
        "sectors": {
            "total": len(sectors_data),
            "breakdown": sector_breakdown
        },
        "financials": {
            "income": total_income,
            "paid": total_paid,
            "debt": total_debt,
            "expenses": total_expenses,
            "profit": total_income - total_expenses,
            "currency": "AZN"
        },
        "payments": {
            "total": total_income,
            "paid": total_paid,
            "remaining": total_debt,
            "currency": "AZN"
        }
    }

# ==================== COMPANIES (ŞİRKƏT MƏLUMATLARI) ====================

@api_router.get("/companies")
async def get_companies(
    sector: Optional[str] = None,
    package: Optional[str] = None,
    company_size: Optional[str] = None,
    marsol_representative: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if sector and sector != "all":
        query["sector"] = sector
    if package and package != "all":
        query["package"] = package
    if company_size and company_size != "all":
        query["company_size"] = company_size
    if marsol_representative and marsol_representative != "all":
        query["marsol_representative"] = marsol_representative
    if status and status != "all":
        query["status"] = status
    query = await apply_scope(query, current_user, "companies")
    companies = await db.companies.find(query, {"_id": 0}).sort("created_at", 1).to_list(1000)
    # Compute finance_id (FN001, FN002...) and contract_days
    today = datetime.now(timezone.utc).date()
    for idx, c in enumerate(companies):
        c["finance_id"] = f"FN{str(idx + 1).zfill(3)}"
        cs = c.get("contract_start_date", "")
        if cs:
            try:
                d = datetime.strptime(cs, "%Y-%m-%d").date()
                c["contract_days"] = (today - d).days
            except (ValueError, TypeError):
                c["contract_days"] = None
        else:
            c["contract_days"] = None
    return companies

# =====================================================
# COMPANIES MODULE
# =====================================================
async def _next_company_display_id() -> str:
    """Return the next sequential C-prefixed company display id (e.g. C0042)."""
    cursor = db.companies.find({"display_id": {"$regex": r"^C\d+$"}}, {"_id": 0, "display_id": 1})
    max_n = 0
    async for d in cursor:
        try:
            n = int((d.get("display_id") or "C0").lstrip("C"))
            if n > max_n:
                max_n = n
        except (ValueError, TypeError):
            continue
    return f"C{max_n + 1:04d}"


async def _backfill_company_display_ids():
    """One-time backfill — assigns C0001, C0002, ... in created_at order to
    every company that doesn't already have a display_id. Safe to call on
    every startup (no-op when all companies already have one)."""
    have = await db.companies.count_documents({"display_id": {"$exists": True, "$ne": None, "$ne": ""}})
    total = await db.companies.count_documents({})
    if have >= total:
        return
    # Find max existing C-id so we don't reuse numbers.
    max_n = 0
    async for d in db.companies.find({"display_id": {"$regex": r"^C\d+$"}}, {"_id": 0, "display_id": 1}):
        try:
            n = int((d.get("display_id") or "C0").lstrip("C"))
            max_n = max(max_n, n)
        except (ValueError, TypeError):
            pass
    # Assign in created_at ascending order.
    cursor = db.companies.find(
        {"$or": [{"display_id": {"$exists": False}}, {"display_id": None}, {"display_id": ""}]},
        {"_id": 0, "id": 1, "created_at": 1},
    ).sort("created_at", 1)
    async for d in cursor:
        max_n += 1
        await db.companies.update_one({"id": d["id"]}, {"$set": {"display_id": f"C{max_n:04d}"}})


@api_router.post("/companies/backfill-ids")
async def backfill_ids(current_user: dict = Depends(check_permission("companies", "write"))):
    """Manually trigger backfill — useful after a fresh deploy."""
    before = await db.companies.count_documents({"display_id": {"$exists": True}})
    await _backfill_company_display_ids()
    after = await db.companies.count_documents({"display_id": {"$exists": True}})
    return {"assigned": after - before, "total": after}


@api_router.post("/companies/renumber-ids")
async def renumber_company_ids(payload: dict = None, current_user: dict = Depends(check_permission("companies", "write"))):
    """Renumber ALL companies' display_id sequentially starting at C0001.

    Order is controlled via `payload.order_by`:
      - "created_at" (default, ascending): preserves the registration timeline
      - "brand_name": A-Z by name
      - "current_id": keeps the current relative order of display_id (and assigns
        sequential ones to those without one)
    """
    payload = payload or {}
    order_by = payload.get("order_by", "created_at")
    if order_by not in ("created_at", "brand_name", "current_id"):
        order_by = "created_at"

    docs = await db.companies.find({}, {"_id": 0, "id": 1, "brand_name": 1, "created_at": 1, "display_id": 1}).to_list(10000)
    if order_by == "brand_name":
        docs.sort(key=lambda d: (d.get("brand_name") or "").lower())
    elif order_by == "current_id":
        def _id_sort_key(d):
            did = d.get("display_id") or ""
            try:
                return (0, int(did.lstrip("C")))
            except ValueError:
                return (1, 9999999)
        docs.sort(key=_id_sort_key)
    else:  # created_at
        docs.sort(key=lambda d: d.get("created_at") or "")

    for idx, d in enumerate(docs, start=1):
        new_id = f"C{idx:04d}"
        if d.get("display_id") != new_id:
            await db.companies.update_one({"id": d["id"]}, {"$set": {"display_id": new_id}})
    return {"renumbered": len(docs), "order_by": order_by, "first": "C0001", "last": f"C{len(docs):04d}"}


@api_router.post("/companies")
async def create_company(company_data: dict, current_user: dict = Depends(check_permission("companies", "write"))):
    company_id = str(uuid.uuid4())
    display_id = await _next_company_display_id()
    company_doc = {
        "id": company_id,
        "display_id": display_id,
        **company_data,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    # Tenant isolation: auto-set marsol_company from creating user if not provided
    if not (company_doc.get("marsol_company") or "").strip():
        company_doc["marsol_company"] = (current_user.get("marsol_company") or "").strip()
    # Calculate debt
    total = company_doc.get("total_amount", 0) or 0
    paid = company_doc.get("paid_amount", 0) or 0
    company_doc["debt_amount"] = total - paid
    
    await db.companies.insert_one(company_doc)
    company_doc.pop("_id", None)
    return company_doc

@api_router.post("/companies/import-excel")
async def import_companies_excel(
    file: UploadFile = File(...),
    current_user: dict = Depends(check_permission("companies", "write")),
):
    """Bulk-import companies from a 2-column Excel file.

    Expected columns (header row, case-insensitive — first match wins):
      - Şirkət adı | Brand name | Company | Ad
      - Paket     | Package

    Behavior: existing company matched by brand_name (case-insensitive trim) →
    `package` field is updated. New rows → company created with sensible
    defaults so the row passes validation; user can edit other fields later.
    Returns counts: {created, updated, skipped, total, errors[]}.
    """
    from openpyxl import load_workbook
    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel oxuna bilmədi: {str(e)[:120]}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"created": 0, "updated": 0, "skipped": 0, "total": 0, "errors": []}

    # Detect column indexes from header row
    header = [str(c or "").strip().lower() for c in rows[0]]
    name_aliases = {"şirkət adı", "sirket adi", "brand name", "company", "company name", "ad", "name", "şirkət"}
    pkg_aliases = {"paket", "package", "paket adı"}
    paid_aliases = {"ödənilib", "ödənilmiş", "ödəniş", "odenilib", "paid", "paid amount", "ödənmə", "ödəniş məbləği"}
    name_idx = next((i for i, h in enumerate(header) if h in name_aliases), None)
    pkg_idx = next((i for i, h in enumerate(header) if h in pkg_aliases), None)
    paid_idx = next((i for i, h in enumerate(header) if h in paid_aliases), None)
    if name_idx is None:
        raise HTTPException(status_code=400, detail="'Şirkət adı' sütunu tapılmadı (header row)")
    if pkg_idx is None:
        raise HTTPException(status_code=400, detail="'Paket' sütunu tapılmadı (header row)")

    created = updated = skipped = 0
    errors: List[str] = []
    now_iso = datetime.now(timezone.utc).isoformat()
    actor = current_user.get("name", "")

    # Pre-load existing companies by lowercased brand_name for O(1) match
    existing = {}
    async for doc in db.companies.find({}, {"_id": 0, "id": 1, "brand_name": 1}):
        bn = (doc.get("brand_name") or "").strip().lower()
        if bn:
            existing[bn] = doc["id"]

    # Pre-load package -> price so imported contract & total_amount reflect the
    # agreed price automatically (same fallback defaults the API returns when
    # no packages are configured in db).
    pkg_docs = await db.packages.find({}, {"_id": 0, "name": 1, "price": 1}).to_list(100)
    if not pkg_docs:
        pkg_docs = [
            {"name": "Premium", "price": 5000},
            {"name": "Business", "price": 3000},
            {"name": "Business Plus", "price": 4000},
            {"name": "Sponsor", "price": 8000},
        ]
    pkg_price = {(p.get("name") or "").strip().lower(): float(p.get("price") or 0) for p in pkg_docs}

    for line, row in enumerate(rows[1:], start=2):
        if not row or row[name_idx] is None:
            continue
        brand_name = str(row[name_idx]).strip()
        package = str(row[pkg_idx] or "").strip() if pkg_idx < len(row) else ""
        # Optional paid_amount column — accepts numbers or numeric strings.
        paid_val = 0.0
        if paid_idx is not None and paid_idx < len(row) and row[paid_idx] is not None:
            try:
                paid_val = float(str(row[paid_idx]).replace(",", ".").strip() or 0)
            except (ValueError, TypeError):
                paid_val = 0.0
        if not brand_name:
            continue
        try:
            key = brand_name.lower()
            price = pkg_price.get(package.lower(), 0)
            debt = max(0.0, price - paid_val)
            if key in existing:
                # Update package + price on the matched company AND propagate
                # to the first contract entry if one exists.
                existing_doc = await db.companies.find_one({"id": existing[key]}, {"_id": 0, "contracts": 1, "paid_amount": 1})
                existing_contracts = (existing_doc or {}).get("contracts") or []
                if existing_contracts:
                    existing_contracts[0] = {
                        **existing_contracts[0],
                        "package": package,
                        "total_amount": price,
                        "paid_amount": paid_val,
                        "debt_amount": debt,
                    }
                else:
                    existing_contracts = [{
                        "project": "Üzvlük", "package": package, "start_date": "", "end_date": "",
                        "join_date": "", "total_amount": price, "paid_amount": paid_val, "debt_amount": debt, "contract_file": "",
                    }]
                await db.companies.update_one(
                    {"id": existing[key]},
                    {"$set": {
                        "package": package,
                        "contracts": existing_contracts,
                        "total_amount": price,
                        "paid_amount": paid_val,
                        "debt_amount": debt,
                        "updated_at": now_iso,
                        "updated_by": actor,
                    }},
                )
                updated += 1
            else:
                # Create with defaults — only brand_name + package have user values
                new_display = await _next_company_display_id()
                doc = {
                    "id": str(uuid.uuid4()),
                    "display_id": new_display,
                    "brand_name": brand_name,
                    "legal_name": "",
                    "sector": "",
                    "company_size": "",
                    "owner_name": "",
                    "owner_phone": "",
                    "owner_email": "",
                    "marsol_representative": "",
                    "joined_project": "Üzvlük",
                    "package": package,
                    # Populate contracts array so the edit modal's package <Select>
                    # pre-selects the imported value out-of-the-box.
                    "contracts": [{
                        "project": "Üzvlük",
                        "package": package,
                        "start_date": "",
                        "end_date": "",
                        "join_date": "",
                        "total_amount": price,
                        "paid_amount": paid_val,
                        "debt_amount": debt,
                        "contract_file": "",
                    }],
                    "total_amount": price,
                    "paid_amount": paid_val,
                    "debt_amount": debt,
                    "status": "Aktiv",
                    "created_at": now_iso,
                    "created_by": actor,
                    "imported_via": "excel",
                }
                await db.companies.insert_one(doc.copy())
                existing[key] = doc["id"]
                created += 1
        except Exception as e:
            skipped += 1
            errors.append(f"Sətir {line} ({brand_name}): {str(e)[:120]}")

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total": created + updated + skipped,
        "errors": errors[:50],
    }



@api_router.get("/companies/{company_id}")
async def get_company(company_id: str, current_user: dict = Depends(get_current_user)):
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Şirkət tapılmadı")
    return company

@api_router.put("/companies/{company_id}")
async def update_company(company_id: str, company_data: dict, current_user: dict = Depends(check_permission("companies", "write"))):
    update_data = {k: v for k, v in company_data.items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Yenilənəcək məlumat yoxdur")
    
    # Recalculate debt if amounts changed
    if "total_amount" in update_data or "paid_amount" in update_data:
        company = await db.companies.find_one({"id": company_id}, {"_id": 0})
        total = update_data.get("total_amount", company.get("total_amount", 0)) or 0
        paid = update_data.get("paid_amount", company.get("paid_amount", 0)) or 0
        update_data["debt_amount"] = total - paid
    
    result = await db.companies.update_one({"id": company_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Şirkət tapılmadı")
    
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    return company

# Update finance note and payment info for a company
@api_router.put("/companies/{company_id}/finance")
async def update_company_finance(company_id: str, data: dict, current_user: dict = Depends(check_permission("finance", "write"))):
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Şirkət tapılmadı")
    
    update_data = {}
    
    # Handle new payment (additive)
    new_payment = data.get("new_payment_amount")
    if new_payment and float(new_payment) > 0:
        payment_amount = float(new_payment)
        old_paid = float(company.get("paid_amount", 0) or 0)
        total = float(company.get("total_amount", 0) or company.get("payment_amount", 0) or 0)
        new_paid = old_paid + payment_amount
        update_data["paid_amount"] = new_paid
        update_data["debt_amount"] = total - new_paid
        update_data["last_payment_date"] = data.get("payment_date", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
        
        # Save payment history
        payment_record = {
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "amount": payment_amount,
            "date": data.get("payment_date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            "note": data.get("payment_note", ""),
            "payment_method": data.get("payment_method", ""),
            "recorded_by": current_user.get("name", ""),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.payment_history.insert_one(payment_record)
    
    # Handle direct field updates (total_amount, finance_note, contract dates, finance tracking fields)
    for key in ["finance_note", "total_amount", "payment_amount", "last_payment_date",
                "contract_start_date", "contract_end_date", "contract_status",
                "finance_contract_number", "payment_due_date", "voen",
                "e_invoice_date", "e_invoice_number", "follow_up", "marsol_company"]:
        if key in data and key != "new_payment_amount":
            update_data[key] = data[key]
    
    # Recalculate debt if total changed
    if "total_amount" in update_data or "payment_amount" in update_data:
        total = float(update_data.get("total_amount", update_data.get("payment_amount", company.get("total_amount", 0) or company.get("payment_amount", 0))) or 0)
        paid = float(update_data.get("paid_amount", company.get("paid_amount", 0)) or 0)
        update_data["debt_amount"] = total - paid
    
    if not update_data:
        raise HTTPException(status_code=400, detail="Yenilənəcək məlumat yoxdur")
    
    await db.companies.update_one({"id": company_id}, {"$set": update_data})
    updated = await db.companies.find_one({"id": company_id}, {"_id": 0})
    return updated

@api_router.get("/companies/{company_id}/payments")
async def get_payment_history(company_id: str, current_user: dict = Depends(get_current_user)):
    payments = await db.payment_history.find({"company_id": company_id}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return payments

@api_router.delete("/companies/{company_id}")
async def delete_company(company_id: str, current_user: dict = Depends(check_permission("companies", "write"))):
    result = await db.companies.delete_one({"id": company_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Şirkət tapılmadı")
    return {"message": "Şirkət silindi"}


@api_router.post("/companies/bulk-delete")
async def bulk_delete_companies(payload: dict, current_user: dict = Depends(check_permission("companies", "write"))):
    """Delete multiple companies at once. Body: {ids: [...]}."""
    ids = payload.get("ids") or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="ids siyahısı boş ola bilməz")
    result = await db.companies.delete_many({"id": {"$in": ids}})
    return {"deleted": result.deleted_count, "requested": len(ids)}

# ==================== EMPLOYEES (İNSAN RESURSLARI) ====================

@api_router.get("/employees")
async def get_employees(
    department: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if department and department != "all":
        query["department"] = department
    if status and status != "all":
        query["status"] = status
    
    employees = await db.employees.find(query, {"_id": 0}).to_list(1000)
    return employees

@api_router.post("/employees")
async def create_employee(employee_data: dict, current_user: dict = Depends(check_permission("hr", "write"))):
    employee_id = str(uuid.uuid4())
    # Auto-generate employee code (E001, E002...)
    count = await db.employees.count_documents({})
    employee_code = f"E{str(count + 1).zfill(3)}"
    employee_doc = {
        "id": employee_id,
        "employee_code": employee_code,
        **employee_data,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.employees.insert_one(employee_doc)
    employee_doc.pop("_id", None)
    return employee_doc

@api_router.get("/employees/{employee_id}")
async def get_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Əməkdaş tapılmadı")
    return employee

@api_router.put("/employees/{employee_id}")
async def update_employee(employee_id: str, employee_data: dict, current_user: dict = Depends(check_permission("hr", "write"))):
    update_data = {k: v for k, v in employee_data.items() if v is not None}
    result = await db.employees.update_one({"id": employee_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Əməkdaş tapılmadı")
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    return employee

@api_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str, current_user: dict = Depends(check_permission("hr", "write"))):
    result = await db.employees.delete_one({"id": employee_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Əməkdaş tapılmadı")
    return {"message": "Əməkdaş silindi"}

# ==================== HR — KPI (TAPŞIRIQ HESABATI) ====================

def _date_in_range(iso_str: str, start: Optional[str], end: Optional[str]) -> bool:
    """Compare an ISO date string against optional YYYY-MM-DD bounds (inclusive)."""
    if not iso_str:
        return False
    d = iso_str[:10]
    if start and d < start:
        return False
    if end and d > end:
        return False
    return True


@api_router.get("/hr/kpi")
async def hr_kpi(
    period: Optional[str] = "month",  # day | week | month | year | custom
    start: Optional[str] = None,       # YYYY-MM-DD (used when period=custom)
    end: Optional[str] = None,
    department: Optional[str] = None,
    marsol_company: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(check_permission("hr", "read")),
):
    """Aggregated KPI report — per user / department / müəssisə task counts.

    Returns three breakdowns sharing the same filter window:
      - users:   [{name, department, marsol_company, total, completed, in_progress, pending, overdue, completion_rate}]
      - departments: same shape grouped by department
      - companies:   grouped by müəssisə
      - totals: overall counts for the filtered window
    """
    from datetime import datetime as _dt, timedelta as _td

    today = _dt.now(timezone.utc).date()
    if period == "day":
        s_date = today.isoformat()
        e_date = today.isoformat()
    elif period == "week":
        s_date = (today - _td(days=today.weekday())).isoformat()
        e_date = today.isoformat()
    elif period == "month":
        s_date = today.replace(day=1).isoformat()
        e_date = today.isoformat()
    elif period == "year":
        s_date = today.replace(month=1, day=1).isoformat()
        e_date = today.isoformat()
    elif period == "all":
        s_date = None
        e_date = None
    else:  # custom
        s_date = start
        e_date = end

    # Load users keyed by name (so we can join task assignee/responsible)
    users_db = await db.users.find({}, {"_id": 0, "name": 1, "department": 1, "marsol_company": 1, "status": 1}).to_list(1000)
    user_by_name = {u["name"]: u for u in users_db if u.get("name")}

    # Load tasks (admin = unscoped, others = apply_scope so KPI respects RBAC)
    q = await apply_scope({}, current_user, "tasks")
    tasks_active = await db.tasks.find(q, {"_id": 0}).to_list(20000)
    # Archived tasks: still count toward stats so completed tasks moved into
    # the archive are visible in KPI summaries.
    tasks_archived = await db.tasks_archive.find(q, {"_id": 0}).to_list(20000)
    tasks_all = list(tasks_active) + list(tasks_archived)

    # Filter by date window (created_at OR start_date — pick created_at)
    def in_window(t):
        if not s_date and not e_date:
            return True
        return _date_in_range(t.get("created_at") or "", s_date, e_date)

    tasks_window = [t for t in tasks_all if in_window(t)]

    if department and department != "all":
        tasks_window = [t for t in tasks_window if (t.get("department") or "") == department]
    if marsol_company and marsol_company != "all":
        tasks_window = [t for t in tasks_window if (t.get("marsol_company") or "") == marsol_company]

    def _stat_row():
        return {
            "total": 0, "completed": 0, "in_progress": 0, "pending": 0, "overdue": 0,
            # Per-role splits (only meaningful for users_bucket)
            "as_executor": 0, "as_executor_completed": 0,
            "as_responsible": 0, "as_responsible_completed": 0,
        }

    today_iso = today.isoformat()

    def _status_buckets(t: dict):
        st = (t.get("status") or "").strip()
        end_d = (t.get("end_date") or "")[:10]
        overdue = bool(end_d and end_d < today_iso and st != "Tamamlandı")
        if st == "Tamamlandı":
            return {"completed": 1, "in_progress": 0, "pending": 0, "overdue": int(overdue)}
        if st == "İcradadır" or st == "İcrada":
            return {"completed": 0, "in_progress": 1, "pending": 0, "overdue": int(overdue)}
        return {"completed": 0, "in_progress": 0, "pending": 1, "overdue": int(overdue)}

    def _add(row, buckets):
        row["total"] += 1
        row["completed"] += buckets["completed"]
        row["in_progress"] += buckets["in_progress"]
        row["pending"] += buckets["pending"]
        row["overdue"] += buckets["overdue"]

    # ----- USERS BUCKET (normalized name → single row, separate role columns) -----
    def _norm_name(s: str) -> str:
        return (s or "").strip().lower()

    # Build canonical-name map (lower-cased → original display name) to ensure
    # we show ONE row per user, with the prettiest form of their name.
    name_canon: dict = {}
    name_meta: dict = {}
    for u in users_db:
        nm = (u.get("name") or "").strip()
        if not nm:
            continue
        key = _norm_name(nm)
        name_canon.setdefault(key, nm)
        name_meta.setdefault(key, {
            "department": u.get("department") or "—",
            "marsol_company": u.get("marsol_company") or "—",
        })

    users_bucket: dict = {}

    def _ensure_user_row(name: str, dept_fallback: str = "—", mc_fallback: str = "—"):
        key = _norm_name(name)
        if not key:
            return None, None
        canon = name_canon.get(key, name.strip())
        if key not in users_bucket:
            meta = name_meta.get(key, {"department": dept_fallback, "marsol_company": mc_fallback})
            users_bucket[key] = {**_stat_row(), "name": canon, "department": meta["department"], "marsol_company": meta["marsol_company"]}
        return key, users_bucket[key]

    dept_bucket: dict = {}
    company_bucket: dict = {}

    for t in tasks_window:
        buckets = _status_buckets(t)
        # Collect executors (assignee) and responsible separately
        executors = {_norm_name(n) for n in _as_name_list(t.get("assignee")) if (n or "").strip()}
        responsibles = {_norm_name(n) for n in _as_name_list(t.get("responsible_person")) if (n or "").strip()}

        # Per-user breakdown — UNION (so user appears ONCE)
        persons = executors | responsibles
        for k in persons:
            display = name_canon.get(k) or k
            _, row = _ensure_user_row(display, t.get("department") or "—", t.get("marsol_company") or "—")
            if not row:
                continue
            _add(row, buckets)
            if k in executors:
                row["as_executor"] += 1
                row["as_executor_completed"] += buckets["completed"]
            if k in responsibles:
                row["as_responsible"] += 1
                row["as_responsible_completed"] += buckets["completed"]

        # Per-department & per-company breakdowns (task-level)
        dep = (t.get("department") or "—").strip() or "—"
        if dep not in dept_bucket:
            dept_bucket[dep] = {**_stat_row(), "name": dep}
        _add(dept_bucket[dep], buckets)

        mc = (t.get("marsol_company") or "—").strip() or "—"
        if mc not in company_bucket:
            company_bucket[mc] = {**_stat_row(), "name": mc}
        _add(company_bucket[mc], buckets)

    def _finalize(bucket):
        rows = list(bucket.values())
        for r in rows:
            r["completion_rate"] = round(r["completed"] * 100 / r["total"], 1) if r["total"] else 0
        rows.sort(key=lambda x: (-x["total"], x["name"]))
        return rows

    users_rows = _finalize(users_bucket)
    if search:
        s = search.strip().lower()
        users_rows = [r for r in users_rows if s in r["name"].lower() or s in r["department"].lower() or s in r["marsol_company"].lower()]

    totals = _stat_row()
    for t in tasks_window:
        b = _status_buckets(t)
        _add(totals, b)
    totals["completion_rate"] = round(totals["completed"] * 100 / totals["total"], 1) if totals["total"] else 0

    return {
        "period": period,
        "start": s_date,
        "end": e_date,
        "totals": totals,
        "users": users_rows,
        "departments": _finalize(dept_bucket),
        "companies": _finalize(company_bucket),
    }


@api_router.post("/tasks/dispatch-reminders")
async def dispatch_task_reminders(current_user: dict = Depends(get_current_user)):
    """Idempotent — creates in-app + push notifications for tasks whose deadline
    is approaching (today/tomorrow) or already overdue. Skips tasks where a
    reminder has already been dispatched today (tracked via `last_reminder_at`).
    """
    from datetime import datetime as _dt
    today = _dt.now(timezone.utc).date()
    today_iso = today.isoformat()
    tomorrow_iso = (today + timedelta(days=1)).isoformat()

    cur = db.tasks.find({"status": {"$ne": "Tamamlandı"}, "end_date": {"$nin": [None, ""]}}, {"_id": 0})
    dispatched = 0
    async for t in cur:
        last = (t.get("last_reminder_at") or "")[:10]
        if last == today_iso:
            continue
        end_d = (t.get("end_date") or "")[:10]
        if not end_d:
            continue
        if end_d > tomorrow_iso:
            continue  # not yet within reminder window
        # Reminder window: due today, due tomorrow, or already overdue
        if end_d < today_iso:
            label = f"Gecikmiş ({(today - _dt.strptime(end_d, '%Y-%m-%d').date()).days} gün)"
            push_title = "Tapşırıq gecikib"
        elif end_d == today_iso:
            label = "Bu gün bitir"
            push_title = "Tapşırıq bu gün bitir"
        else:
            label = "Sabah bitir"
            push_title = "Tapşırıq sabah bitir"

        body = f"{t.get('task_name','')}{label}"
        recipients = set()
        for nm in _as_name_list(t.get("assignee")):
            if nm:
                recipients.add(nm)
        for nm in _as_name_list(t.get("responsible_person")):
            if nm:
                recipients.add(nm)
        for r_name in recipients:
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()),
                "type": "task_reminder",
                "title": push_title,
                "body": body,
                "task_id": t["id"],
                "task_code": t.get("task_code", ""),
                "recipient_name": r_name,
                "is_read": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
        if recipients:
            _safe_push(list(recipients), push_title, body, link=f"/tasks?id={t['id']}", data={"type": "task_reminder", "task_id": t["id"]})
        await db.tasks.update_one({"id": t["id"]}, {"$set": {"last_reminder_at": datetime.now(timezone.utc).isoformat()}})
        dispatched += 1
    return {"dispatched": dispatched}


# ==================== ATTENDANCE (DAVAMİYYƏT) ====================

ATTENDANCE_STATUSES = ["İşdə", "Gəlməyib", "Məzuniyyət", "Xəstəlik", "İcazəli", "Uzaq"]

@api_router.get("/attendance")
async def get_attendance(
    date: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    employee_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if date:
        query["date"] = date
    elif start and end:
        query["date"] = {"$gte": start, "$lte": end}
    if employee_id:
        query["employee_id"] = employee_id
    records = await db.attendance.find(query, {"_id": 0}).sort("date", -1).to_list(5000)
    return records

@api_router.post("/attendance")
async def upsert_attendance(data: dict, current_user: dict = Depends(check_permission("hr", "write"))):
    employee_id = data.get("employee_id", "")
    date = data.get("date", "")
    if not employee_id or not date:
        raise HTTPException(status_code=400, detail="employee_id və date tələb olunur")
    emp = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    update = {
        "employee_id": employee_id,
        "employee_name": (emp.get("full_name") or f"{emp.get('first_name','')} {emp.get('last_name','')}".strip()) if emp else data.get("employee_name", ""),
        "date": date,
        "status": data.get("status", "İşdə"),
        "check_in": data.get("check_in", ""),
        "check_out": data.get("check_out", ""),
        "notes": data.get("notes", ""),
        "updated_by": current_user.get("name", ""),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    existing = await db.attendance.find_one({"employee_id": employee_id, "date": date})
    if existing:
        await db.attendance.update_one({"employee_id": employee_id, "date": date}, {"$set": update})
        update["id"] = existing["id"]
    else:
        update["id"] = str(uuid.uuid4())
        update["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.attendance.insert_one(dict(update))
    update.pop("_id", None)
    return update

@api_router.post("/attendance/bulk")
async def bulk_attendance(data: dict, current_user: dict = Depends(check_permission("hr", "write"))):
    date = data.get("date", "")
    records = data.get("records", [])
    if not date or not records:
        raise HTTPException(status_code=400, detail="date və records tələb olunur")
    count = 0
    for r in records:
        emp = await db.employees.find_one({"id": r.get("employee_id", "")}, {"_id": 0})
        doc = {
            "employee_id": r.get("employee_id", ""),
            "employee_name": (emp.get("full_name") or f"{emp.get('first_name','')} {emp.get('last_name','')}".strip()) if emp else "",
            "date": date,
            "status": r.get("status", "İşdə"),
            "check_in": r.get("check_in", ""),
            "check_out": r.get("check_out", ""),
            "notes": r.get("notes", ""),
            "updated_by": current_user.get("name", ""),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        existing = await db.attendance.find_one({"employee_id": doc["employee_id"], "date": date})
        if existing:
            await db.attendance.update_one({"employee_id": doc["employee_id"], "date": date}, {"$set": doc})
        else:
            doc["id"] = str(uuid.uuid4())
            doc["created_at"] = datetime.now(timezone.utc).isoformat()
            await db.attendance.insert_one(dict(doc))
        count += 1
    return {"message": f"{count} qeyd saxlanıldı"}

@api_router.get("/attendance/stats")
async def attendance_stats(month: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")
    start = f"{month}-01"
    end = f"{month}-31"
    records = await db.attendance.find({"date": {"$gte": start, "$lte": end}}, {"_id": 0}).to_list(10000)
    employees = await db.employees.find({"status": "Aktiv"}, {"_id": 0}).to_list(1000)
    per_employee = {}
    for e in employees:
        eid = e["id"]
        per_employee[eid] = {
            "employee_id": eid,
            "employee_name": e.get("full_name") or f"{e.get('first_name','')} {e.get('last_name','')}".strip(),
            "department": e.get("department", ""),
            "position": e.get("position", ""),
            **{s: 0 for s in ATTENDANCE_STATUSES}
        }
    for r in records:
        eid = r.get("employee_id", "")
        st = r.get("status", "")
        if eid in per_employee and st in per_employee[eid]:
            per_employee[eid][st] += 1
    totals = {s: sum(p[s] for p in per_employee.values()) for s in ATTENDANCE_STATUSES}
    return {"month": month, "per_employee": list(per_employee.values()), "totals": totals}

@api_router.delete("/attendance/{record_id}")
async def delete_attendance(record_id: str, current_user: dict = Depends(check_permission("hr", "write"))):
    await db.attendance.delete_one({"id": record_id})
    return {"message": "Qeyd silindi"}

@api_router.get("/attendance/system-sessions")
async def attendance_system_sessions(
    date: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_id: Optional[str] = None,
    current_user: dict = Depends(check_permission("hr", "read")),
):
    """Returns user login/logout sessions with active duration in seconds.

    - `date` filter (YYYY-MM-DD) keeps sessions whose login_at is on that day
    - `start_date` / `end_date` (YYYY-MM-DD) bound the login_at range inclusively
    - Open sessions report active_seconds based on the LAST heartbeat
      (`last_active_at`), NOT wall-clock now. A session that hasn't heartbeated
      in over STALE_MINUTES is treated as effectively closed (browser tab
      killed) so we don't keep accruing idle time.
    """
    STALE_MINUTES = 5  # heartbeat gap that flags a tab as effectively closed
    query: Dict[str, Any] = {}
    if user_id:
        query["user_id"] = user_id
    if date:
        # date filter on login_at prefix (ISO strings start with YYYY-MM-DD)
        query["login_at"] = {"$regex": f"^{date}"}
    elif start_date or end_date:
        # Range filter (string compare works because ISO timestamps are lexicographically sortable)
        rng: Dict[str, Any] = {}
        if start_date:
            rng["$gte"] = f"{start_date}T00:00:00"
        if end_date:
            rng["$lte"] = f"{end_date}T23:59:59.999999+00:00"
        query["login_at"] = rng
    sessions = await db.user_sessions.find(query, {"_id": 0}).sort("login_at", -1).to_list(2000)

    def _parse(ts: Optional[str]) -> Optional[datetime]:
        if not ts:
            return None
        try:
            return datetime.fromisoformat(ts.replace("Z", "+00:00"))
        except Exception:
            return None

    out = []
    now_dt = datetime.now(timezone.utc)
    stale_threshold = timedelta(minutes=STALE_MINUTES)
    for s in sessions:
        login_dt = _parse(s.get("login_at"))
        logout_dt = _parse(s.get("logout_at"))
        last_dt = _parse(s.get("last_active_at"))
        is_open = s.get("logout_at") is None
        is_stale = False
        # Determine end timestamp:
        #   - closed session → logout_at (fallback last_active_at)
        #   - open session   → last_active_at (when heartbeat stalled the user
        #     has effectively disconnected; using `now` would falsely count
        #     every minute the page sits closed)
        if is_open:
            end = last_dt or login_dt
            # Mark as stale if no heartbeat in N minutes
            if last_dt and (now_dt - last_dt) > stale_threshold:
                is_stale = True
        else:
            end = logout_dt or last_dt or login_dt
        active_seconds = 0
        if login_dt and end:
            active_seconds = max(0, int((end - login_dt).total_seconds()))
        out.append({
            **s,
            "active_seconds": active_seconds,
            "is_open": is_open and not is_stale,
            "is_stale": is_stale,
        })
    # Aggregate per-user totals (sum of all session durations within range)
    totals: Dict[str, Dict[str, Any]] = {}
    for row in out:
        uid = row.get("user_id") or ""
        if uid not in totals:
            totals[uid] = {
                "user_id": uid,
                "user_email": row.get("user_email", ""),
                "user_name": row.get("user_name", ""),
                "total_seconds": 0,
                "sessions": 0,
                "has_open": False,
            }
        totals[uid]["total_seconds"] += int(row.get("active_seconds") or 0)
        totals[uid]["sessions"] += 1
        if row.get("is_open"):
            totals[uid]["has_open"] = True
    return {"sessions": out, "totals": list(totals.values())}

# ==================== LEAVE REQUESTS (MƏZUNİYYƏT SORĞULARI) ====================

@api_router.get("/leave-requests")
async def get_leave_requests(
    employee_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if status and status != "all":
        query["status"] = status
    items = await db.leave_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items

@api_router.post("/leave-requests")
async def create_leave_request(data: dict, current_user: dict = Depends(check_permission("hr", "write"))):
    emp = await db.employees.find_one({"id": data.get("employee_id", "")}, {"_id": 0})
    doc = {
        "id": str(uuid.uuid4()),
        "employee_id": data.get("employee_id", ""),
        "employee_name": (emp.get("full_name") or f"{emp.get('first_name','')} {emp.get('last_name','')}".strip()) if emp else "",
        "type": data.get("type", "Məzuniyyət"),
        "start_date": data.get("start_date", ""),
        "end_date": data.get("end_date", ""),
        "reason": data.get("reason", ""),
        "status": "Gözləyir",
        "created_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.leave_requests.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/leave-requests/{req_id}")
async def update_leave_request(req_id: str, data: dict, current_user: dict = Depends(check_permission("hr", "write"))):
    update = {k: v for k, v in data.items() if k not in ("id",)}
    update["updated_by"] = current_user.get("name", "")
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.leave_requests.update_one({"id": req_id}, {"$set": update})
    # If approved, auto-fill attendance records for the period
    if update.get("status") == "Təsdiqlənib":
        req = await db.leave_requests.find_one({"id": req_id}, {"_id": 0})
        if req and req.get("start_date") and req.get("end_date"):
            try:
                from datetime import date as _date, timedelta
                s = datetime.strptime(req["start_date"], "%Y-%m-%d").date()
                e = datetime.strptime(req["end_date"], "%Y-%m-%d").date()
                att_status = "Xəstəlik" if req.get("type") == "Xəstəlik" else "Məzuniyyət"
                cur = s
                while cur <= e:
                    ds = cur.strftime("%Y-%m-%d")
                    existing = await db.attendance.find_one({"employee_id": req["employee_id"], "date": ds})
                    att_doc = {
                        "employee_id": req["employee_id"],
                        "employee_name": req.get("employee_name", ""),
                        "date": ds,
                        "status": att_status,
                        "notes": f"Sorğu: {req.get('reason', '')}",
                        "updated_by": current_user.get("name", ""),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                    if existing:
                        await db.attendance.update_one({"employee_id": req["employee_id"], "date": ds}, {"$set": att_doc})
                    else:
                        att_doc["id"] = str(uuid.uuid4())
                        att_doc["created_at"] = datetime.now(timezone.utc).isoformat()
                        await db.attendance.insert_one(dict(att_doc))
                    cur += timedelta(days=1)
            except (ValueError, TypeError):
                pass
    doc = await db.leave_requests.find_one({"id": req_id}, {"_id": 0})
    return doc

@api_router.delete("/leave-requests/{req_id}")
async def delete_leave_request(req_id: str, current_user: dict = Depends(check_permission("hr", "write"))):
    await db.leave_requests.delete_one({"id": req_id})
    return {"message": "Sorğu silindi"}

# ==================== FINANCE (MALİYYƏ) ====================

@api_router.get("/finance/incomes")
async def get_incomes(current_user: dict = Depends(get_current_user)):
    incomes = await db.incomes.find({}, {"_id": 0}).to_list(1000)
    return incomes

@api_router.post("/finance/incomes")
async def create_income(income_data: IncomeCreate, current_user: dict = Depends(check_permission("finance", "write"))):
    income_id = str(uuid.uuid4())
    income_doc = {
        "id": income_id,
        **income_data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    income_doc["debt_amount"] = income_doc["amount"] - income_doc["paid_amount"]
    await db.incomes.insert_one(income_doc)
    income_doc.pop("_id", None)
    return income_doc

@api_router.put("/finance/incomes/{income_id}")
async def update_income(income_id: str, income_data: dict, current_user: dict = Depends(check_permission("finance", "write"))):
    update_data = {k: v for k, v in income_data.items() if v is not None}
    if "amount" in update_data or "paid_amount" in update_data:
        current = await db.incomes.find_one({"id": income_id}, {"_id": 0})
        if current:
            amount = update_data.get("amount", current.get("amount", 0))
            paid = update_data.get("paid_amount", current.get("paid_amount", 0))
            update_data["debt_amount"] = amount - paid
    result = await db.incomes.update_one({"id": income_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Gəlir tapılmadı")
    income = await db.incomes.find_one({"id": income_id}, {"_id": 0})
    return income

@api_router.delete("/finance/incomes/{income_id}")
async def delete_income(income_id: str, current_user: dict = Depends(check_permission("finance", "write"))):
    result = await db.incomes.delete_one({"id": income_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Gəlir tapılmadı")
    return {"message": "Gəlir silindi"}


# ---------- Inventar (Inventory) ----------
def _safe_nonneg(v, default=0.0):
    """Coerce to non-negative float (mənfi dəyər icazəsizdir)."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return float(default)
    return max(f, 0.0)


def _months_between(start_iso: str, end_dt: datetime) -> int:
    """Return whole months between an ISO date string (YYYY-MM-DD or full ISO) and end_dt."""
    if not start_iso:
        return 0
    s = str(start_iso).strip()
    try:
        if "T" in s:
            start = datetime.fromisoformat(s.replace("Z", "+00:00")).replace(tzinfo=None)
        else:
            # Accept YYYY-MM-DD or DD/MM/YYYY
            if "/" in s:
                parts = s.split("/")
                if len(parts) == 3:
                    start = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
                else:
                    return 0
            else:
                start = datetime.fromisoformat(s)
    except Exception:
        return 0
    delta_months = (end_dt.year - start.year) * 12 + (end_dt.month - start.month)
    if end_dt.day < start.day:
        delta_months -= 1
    return max(delta_months, 0)


def _years_completed(start_iso: str, end_dt: datetime) -> int:
    """Return number of FULL years completed since start date."""
    return _months_between(start_iso, end_dt) // 12


def _compute_inventory_valuation(item: dict) -> dict:
    """Compute structured financial valuation for a single inventory item.

    Returns a dict with: total_initial_value, annual_depreciation, monthly_depreciation,
    months_used, accumulated_depreciation, book_value, market_value, operational_status,
    market_above_book, suggestion.
    """
    qty = max(int(item.get("quantity") or 1), 1)
    unit_value = _safe_nonneg(item.get("unit_value"))
    # purchase_price defaults to qty*unit_value for backward compat
    purchase_price = _safe_nonneg(item.get("purchase_price"), default=qty * unit_value)
    delivery_cost = _safe_nonneg(item.get("delivery_cost"))
    customs_cost = _safe_nonneg(item.get("customs_cost"))
    installation_cost = _safe_nonneg(item.get("installation_cost"))
    other_costs = _safe_nonneg(item.get("other_costs"))

    total_initial_value = round(
        purchase_price + delivery_cost + customs_cost + installation_cost + other_costs, 2
    )

    # Depreciation rate (declining-balance per FULL year). When provided, this takes
    # precedence over the legacy straight-line useful_life_years calculation.
    depreciation_rate = _safe_nonneg(item.get("depreciable_asset_rate"))  # percent (0..100)
    useful_life_years = _safe_nonneg(item.get("useful_life_years"))
    annual_depr = 0.0
    monthly_depr = 0.0

    now_dt = datetime.now(timezone.utc).replace(tzinfo=None)
    months_used = _months_between(item.get("purchase_date") or "", now_dt)
    years_completed = months_used // 12

    if depreciation_rate > 0 and total_initial_value > 0:
        # Declining balance: each completed year deducts rate% from the remaining book.
        book = total_initial_value
        yearly_breakdown: List[Dict[str, Any]] = []
        for y in range(years_completed):
            depr = round(book * (depreciation_rate / 100.0), 2)
            if depr <= 0 or book <= 0:
                break
            new_book = round(max(book - depr, 0.0), 2)
            yearly_breakdown.append({
                "year": y + 1,
                "opening_balance": round(book, 2),
                "depreciation": depr,
                "closing_balance": new_book,
            })
            book = new_book
            if book <= 0:
                break
        accumulated_depr = round(total_initial_value - book, 2)
        book_value = round(max(book, 0.0), 2)
        # For reporting: first-year depreciation as the "annual" representative
        annual_depr = yearly_breakdown[0]["depreciation"] if yearly_breakdown else round(
            total_initial_value * (depreciation_rate / 100.0), 2
        )
        monthly_depr = round(annual_depr / 12.0, 2)
        method = "declining_balance"
    else:
        # Legacy straight-line fallback
        if useful_life_years > 0 and total_initial_value > 0:
            annual_depr = round(total_initial_value / useful_life_years, 2)
            monthly_depr = round(annual_depr / 12.0, 2)
        accumulated_depr = round(min(monthly_depr * months_used, total_initial_value), 2)
        book_value = round(max(total_initial_value - accumulated_depr, 0.0), 2)
        yearly_breakdown = []
        method = "straight_line" if useful_life_years > 0 else "none"

    market_value = _safe_nonneg(item.get("market_value"))
    is_operational = bool(item.get("is_operational", True))

    # Determine operational status
    fully_depreciated = total_initial_value > 0 and book_value <= 0
    if fully_depreciated and is_operational:
        operational_status = "Tam amortizasiya olunub, amma istifadədədir"
    elif fully_depreciated and not is_operational:
        operational_status = "Silinməyə namizəd"
    elif not is_operational:
        operational_status = "İstifadəyə yararsız"
    else:
        operational_status = "İstifadədədir"

    market_above_book = market_value > book_value and market_value > 0

    suggestion = None
    if fully_depreciated and not is_operational and market_value <= 0:
        suggestion = "Utilizasiya / silinmə tövsiyə olunur"
    elif market_above_book:
        suggestion = f"Bazar dəyəri ({market_value:.2f}) mühasibat qalıq dəyərindən ({book_value:.2f}) yüksəkdir — yenidən qiymətləndirmə nəzərə alına bilər"

    return {
        "purchase_price": purchase_price,
        "delivery_cost": delivery_cost,
        "customs_cost": customs_cost,
        "installation_cost": installation_cost,
        "other_costs": other_costs,
        "total_initial_value": total_initial_value,
        "useful_life_years": useful_life_years,
        "depreciation_rate": depreciation_rate,
        "depreciation_method": method,
        "annual_depreciation": annual_depr,
        "monthly_depreciation": monthly_depr,
        "months_used": months_used,
        "years_completed": years_completed,
        "yearly_breakdown": yearly_breakdown,
        "accumulated_depreciation": accumulated_depr,
        "book_value": book_value,
        "market_value": market_value,
        "is_operational": is_operational,
        "operational_status": operational_status,
        "market_above_book": market_above_book,
        "suggestion": suggestion,
    }


def _enrich_inventory(item: dict) -> dict:
    """Attach a `valuation` block to an inventory document for client consumption."""
    item["valuation"] = _compute_inventory_valuation(item)
    return item


@api_router.get("/finance/inventory")
async def list_inventory(current_user: dict = Depends(check_permission("finance", "read"))):
    items = await db.inventory.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    return [_enrich_inventory(it) for it in items]


@api_router.post("/finance/inventory")
async def create_inventory_item(data: dict, current_user: dict = Depends(check_permission("finance", "write"))):
    # Auto-generate sequential display ID: I001, I002, ...
    count = await db.inventory.count_documents({})
    display_id = data.get("display_id") or f"I{(count + 1):03d}"
    now_iso = datetime.now(timezone.utc).isoformat()
    category_name = (data.get("category") or "").strip()
    # Resolve auto-generated inventory_code based on category prefix when not provided
    inventory_code = (data.get("inventory_code") or "").strip()
    if not inventory_code and category_name:
        inventory_code = await _generate_inventory_code(category_name)
    # Snapshot depreciation rate from the depreciable_assets registry when name provided but rate not
    depreciable_asset = (data.get("depreciable_asset") or "").strip()
    depreciable_asset_rate = _safe_nonneg(data.get("depreciable_asset_rate"))
    if depreciable_asset and not depreciable_asset_rate:
        reg = await db.depreciable_assets.find_one({"name": depreciable_asset}, {"_id": 0})
        if reg and reg.get("rate"):
            depreciable_asset_rate = _safe_nonneg(reg.get("rate"))
    item = {
        "id": str(uuid.uuid4()),
        "display_id": display_id,
        "marsol_company": (data.get("marsol_company") or "").strip(),
        "depreciable_asset": depreciable_asset,
        "depreciable_asset_rate": depreciable_asset_rate,
        "department": (data.get("department") or "").strip(),
        "asset_name": (data.get("asset_name") or "").strip(),
        "category": category_name,
        "inventory_code": inventory_code,
        "quantity": max(int(data.get("quantity") or 1), 1),
        "condition": (data.get("condition") or "").strip(),
        "responsible_person": (data.get("responsible_person") or "").strip(),
        "location": (data.get("location") or "").strip(),
        "purchase_date": (data.get("purchase_date") or "").strip(),
        "last_check_date": (data.get("last_check_date") or "").strip(),
        "status": (data.get("status") or "Aktiv").strip(),
        "note": (data.get("note") or "").strip(),
        "unit_value": _safe_nonneg(data.get("unit_value")),
        # NEW financial fields
        "purchase_price": _safe_nonneg(data.get("purchase_price")),
        "delivery_cost": _safe_nonneg(data.get("delivery_cost")),
        "customs_cost": _safe_nonneg(data.get("customs_cost")),
        "installation_cost": _safe_nonneg(data.get("installation_cost")),
        "other_costs": _safe_nonneg(data.get("other_costs")),
        "useful_life_years": _safe_nonneg(data.get("useful_life_years")),
        "market_value": _safe_nonneg(data.get("market_value")),
        "is_operational": bool(data.get("is_operational", True)),
        "created_at": now_iso,
        "created_by": current_user.get("name", ""),
        "updated_at": now_iso,
    }
    if not item["asset_name"]:
        raise HTTPException(status_code=400, detail="Əmlakın adı boş ola bilməz")
    await db.inventory.insert_one(item)
    item.pop("_id", None)
    return _enrich_inventory(item)


@api_router.put("/finance/inventory/{item_id}")
async def update_inventory_item(item_id: str, data: dict, current_user: dict = Depends(check_permission("finance", "write"))):
    existing = await db.inventory.find_one({"id": item_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="İnventar tapılmadı")
    update: Dict[str, Any] = {}
    for k in ("marsol_company", "depreciable_asset", "department", "asset_name", "category",
              "inventory_code", "condition", "responsible_person", "location", "purchase_date",
              "last_check_date", "status", "note"):
        if k in data:
            update[k] = (str(data.get(k) or "")).strip()
    if "quantity" in data:
        update["quantity"] = max(int(data.get("quantity") or 1), 1)
    # Non-negative numeric fields
    for k in ("unit_value", "purchase_price", "delivery_cost", "customs_cost",
              "installation_cost", "other_costs", "useful_life_years", "market_value",
              "depreciable_asset_rate"):
        if k in data:
            update[k] = _safe_nonneg(data.get(k))
    if "is_operational" in data:
        update["is_operational"] = bool(data.get("is_operational"))
    # Auto-fill rate from registry if depreciable_asset changed and explicit rate not provided
    if "depreciable_asset" in data and "depreciable_asset_rate" not in data:
        reg = await db.depreciable_assets.find_one({"name": update.get("depreciable_asset", "")}, {"_id": 0})
        if reg and reg.get("rate"):
            update["depreciable_asset_rate"] = _safe_nonneg(reg.get("rate"))
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.inventory.update_one({"id": item_id}, {"$set": update})
    updated = await db.inventory.find_one({"id": item_id}, {"_id": 0})
    return _enrich_inventory(updated)


async def _generate_inventory_code(category_name: str) -> str:
    """Generate `<PREFIX>-NNN` code based on category. Prefix comes from inventory_categories
    registry (`code_prefix` field) or falls back to first 3 uppercase letters of the name."""
    cat = await db.inventory_categories.find_one({"name": category_name}, {"_id": 0})
    prefix = ""
    if cat and cat.get("code_prefix"):
        prefix = str(cat["code_prefix"]).strip().upper()
    if not prefix:
        # Strip non-alphabetic characters, uppercase, take first 3 chars
        cleaned = "".join(ch for ch in category_name if ch.isalpha()).upper()
        prefix = cleaned[:3] if cleaned else "INV"
    # Find max existing code with this prefix
    pattern = f"^{prefix}-(\\d+)$"
    cursor = db.inventory.find({"inventory_code": {"$regex": pattern}}, {"_id": 0, "inventory_code": 1})
    max_n = 0
    async for doc in cursor:
        try:
            n = int(doc["inventory_code"].split("-")[-1])
            if n > max_n:
                max_n = n
        except (ValueError, IndexError):
            continue
    return f"{prefix}-{(max_n + 1):03d}"


@api_router.delete("/finance/inventory/{item_id}")
async def delete_inventory_item(item_id: str, current_user: dict = Depends(check_permission("finance", "write"))):
    result = await db.inventory.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="İnventar tapılmadı")
    return {"message": "İnventar silindi"}


@api_router.get("/finance/inventory/{item_id}/valuation")
async def get_inventory_valuation(item_id: str, current_user: dict = Depends(check_permission("finance", "read"))):
    """Return only the structured valuation object for a single inventory item."""
    item = await db.inventory.find_one({"id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="İnventar tapılmadı")
    return _compute_inventory_valuation(item)


@api_router.get("/finance/inventory/value-report")
async def inventory_value_report(current_user: dict = Depends(check_permission("finance", "read"))):
    """Aggregate inventory financials: initial value, book value, market value, accumulated
    depreciation — grouped by department, category and operational status."""
    items = await db.inventory.find({}, {"_id": 0}).to_list(5000)
    by_dept: Dict[str, Dict[str, Any]] = {}
    by_cat: Dict[str, Dict[str, Any]] = {}
    by_op_status: Dict[str, int] = {}
    grand_initial = 0.0
    grand_book = 0.0
    grand_market = 0.0
    grand_accum_depr = 0.0
    grand_qty = 0
    fully_depreciated_count = 0
    writeoff_candidates: List[Dict[str, Any]] = []
    revaluation_candidates: List[Dict[str, Any]] = []
    for it in items:
        v = _compute_inventory_valuation(it)
        qty = int(it.get("quantity") or 0)
        grand_initial += v["total_initial_value"]
        grand_book += v["book_value"]
        grand_market += v["market_value"]
        grand_accum_depr += v["accumulated_depreciation"]
        grand_qty += qty
        if v["total_initial_value"] > 0 and v["book_value"] <= 0:
            fully_depreciated_count += 1
        if v["operational_status"] == "Silinməyə namizəd":
            writeoff_candidates.append({
                "id": it.get("id"), "display_id": it.get("display_id"),
                "asset_name": it.get("asset_name"), "department": it.get("department"),
                "book_value": v["book_value"], "market_value": v["market_value"],
            })
        if v["market_above_book"]:
            revaluation_candidates.append({
                "id": it.get("id"), "display_id": it.get("display_id"),
                "asset_name": it.get("asset_name"),
                "book_value": v["book_value"], "market_value": v["market_value"],
                "delta": round(v["market_value"] - v["book_value"], 2),
            })
        d = it.get("department") or "(boş)"
        if d not in by_dept:
            by_dept[d] = {"department": d, "items": 0, "quantity": 0,
                          "initial_value": 0.0, "book_value": 0.0, "market_value": 0.0,
                          "accumulated_depreciation": 0.0}
        by_dept[d]["items"] += 1
        by_dept[d]["quantity"] += qty
        by_dept[d]["initial_value"] += v["total_initial_value"]
        by_dept[d]["book_value"] += v["book_value"]
        by_dept[d]["market_value"] += v["market_value"]
        by_dept[d]["accumulated_depreciation"] += v["accumulated_depreciation"]
        c = it.get("category") or "(boş)"
        if c not in by_cat:
            by_cat[c] = {"category": c, "items": 0, "quantity": 0,
                         "initial_value": 0.0, "book_value": 0.0, "market_value": 0.0,
                         "accumulated_depreciation": 0.0}
        by_cat[c]["items"] += 1
        by_cat[c]["quantity"] += qty
        by_cat[c]["initial_value"] += v["total_initial_value"]
        by_cat[c]["book_value"] += v["book_value"]
        by_cat[c]["market_value"] += v["market_value"]
        by_cat[c]["accumulated_depreciation"] += v["accumulated_depreciation"]
        op = v["operational_status"]
        by_op_status[op] = by_op_status.get(op, 0) + 1

    def _round_group(rows):
        for r in rows:
            for k in ("initial_value", "book_value", "market_value", "accumulated_depreciation"):
                r[k] = round(r[k], 2)
        return rows

    return {
        "totals": {
            "items": len(items),
            "quantity": grand_qty,
            "initial_value": round(grand_initial, 2),
            "book_value": round(grand_book, 2),
            "market_value": round(grand_market, 2),
            "accumulated_depreciation": round(grand_accum_depr, 2),
            "fully_depreciated_count": fully_depreciated_count,
        },
        "by_department": _round_group(sorted(by_dept.values(), key=lambda x: x["book_value"], reverse=True)),
        "by_category": _round_group(sorted(by_cat.values(), key=lambda x: x["book_value"], reverse=True)),
        "by_operational_status": [{"status": k, "count": v} for k, v in by_op_status.items()],
        "writeoff_candidates": writeoff_candidates,
        "revaluation_candidates": revaluation_candidates,
    }


# ---------- Settings: Depreciable Assets registry ----------
DEFAULT_DEPRECIABLE_ASSETS = [
    {"name": "Binalar və tikililər", "rate": 5},
    {"name": "Maşın və avadanlıqlar", "rate": 20},
    {"name": "Nəqliyyat vasitələri", "rate": 25},
    {"name": "İT və ofis avadanlığı", "rate": 25},
    {"name": "Mebel", "rate": 20},
    {"name": "Digər əsas vəsait", "rate": 10},
]


@api_router.get("/settings/depreciable-assets")
async def list_depreciable_assets(current_user: dict = Depends(get_current_user)):
    items = await db.depreciable_assets.find({}, {"_id": 0}).sort("name", 1).to_list(200)
    if not items:
        seeded = []
        for d in DEFAULT_DEPRECIABLE_ASSETS:
            doc = {"id": str(uuid.uuid4()), "name": d["name"], "rate": float(d["rate"])}
            await db.depreciable_assets.insert_one(doc)
            doc.pop("_id", None)
            seeded.append(doc)
        return seeded
    return items


@api_router.post("/settings/depreciable-assets")
async def create_depreciable_asset(data: dict, current_user: dict = Depends(get_current_user)):
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Ad boş ola bilməz")
    rate = _safe_nonneg(data.get("rate"))
    if rate > 100:
        rate = 100.0
    if await db.depreciable_assets.find_one({"name": name}):
        raise HTTPException(status_code=400, detail="Bu ad artıq mövcuddur")
    doc = {"id": str(uuid.uuid4()), "name": name, "rate": rate}
    await db.depreciable_assets.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/settings/depreciable-assets/{item_id}")
async def update_depreciable_asset(item_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    update: Dict[str, Any] = {}
    if "name" in data:
        n = (data.get("name") or "").strip()
        if not n:
            raise HTTPException(status_code=400, detail="Ad boş ola bilməz")
        update["name"] = n
    if "rate" in data:
        update["rate"] = min(_safe_nonneg(data.get("rate")), 100.0)
    if not update:
        raise HTTPException(status_code=400, detail="Yenilənəcək məlumat yoxdur")
    result = await db.depreciable_assets.update_one({"id": item_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tapılmadı")
    return await db.depreciable_assets.find_one({"id": item_id}, {"_id": 0})


@api_router.delete("/settings/depreciable-assets/{item_id}")
async def delete_depreciable_asset(item_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.depreciable_assets.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tapılmadı")
    return {"message": "Silindi"}


# ---------- Settings: Inventory Categories registry ----------
DEFAULT_INVENTORY_CATEGORIES = [
    {"name": "Kompüter texnikası", "code_prefix": "KOM"},
    {"name": "Mebel", "code_prefix": "MEB"},
    {"name": "Nəqliyyat", "code_prefix": "NQL"},
    {"name": "Ofis avadanlığı", "code_prefix": "OFS"},
    {"name": "Texniki avadanlıq", "code_prefix": "TXN"},
    {"name": "Digər", "code_prefix": "DGR"},
]


@api_router.get("/settings/inventory-categories")
async def list_inventory_categories(current_user: dict = Depends(get_current_user)):
    items = await db.inventory_categories.find({}, {"_id": 0}).sort("name", 1).to_list(200)
    if not items:
        seeded = []
        for d in DEFAULT_INVENTORY_CATEGORIES:
            doc = {"id": str(uuid.uuid4()), "name": d["name"], "code_prefix": d["code_prefix"]}
            await db.inventory_categories.insert_one(doc)
            doc.pop("_id", None)
            seeded.append(doc)
        return seeded
    return items


@api_router.post("/settings/inventory-categories")
async def create_inventory_category(data: dict, current_user: dict = Depends(get_current_user)):
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Ad boş ola bilməz")
    prefix = (data.get("code_prefix") or "").strip().upper()[:6]
    if not prefix:
        cleaned = "".join(ch for ch in name if ch.isalpha()).upper()
        prefix = cleaned[:3] if cleaned else "INV"
    if await db.inventory_categories.find_one({"name": name}):
        raise HTTPException(status_code=400, detail="Bu ad artıq mövcuddur")
    doc = {"id": str(uuid.uuid4()), "name": name, "code_prefix": prefix}
    await db.inventory_categories.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/settings/inventory-categories/{item_id}")
async def update_inventory_category(item_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    update: Dict[str, Any] = {}
    if "name" in data:
        n = (data.get("name") or "").strip()
        if not n:
            raise HTTPException(status_code=400, detail="Ad boş ola bilməz")
        update["name"] = n
    if "code_prefix" in data:
        update["code_prefix"] = (data.get("code_prefix") or "").strip().upper()[:6]
    if not update:
        raise HTTPException(status_code=400, detail="Yenilənəcək məlumat yoxdur")
    result = await db.inventory_categories.update_one({"id": item_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tapılmadı")
    return await db.inventory_categories.find_one({"id": item_id}, {"_id": 0})


@api_router.delete("/settings/inventory-categories/{item_id}")
async def delete_inventory_category(item_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.inventory_categories.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tapılmadı")
    return {"message": "Silindi"}



@api_router.get("/finance/expenses")
async def get_expenses(current_user: dict = Depends(get_current_user)):
    expenses = await db.expenses.find({}, {"_id": 0}).to_list(1000)
    return expenses


# ---------- Settings: Social Media Platforms registry ----------
DEFAULT_SOCIAL_PLATFORMS = [
    {"name": "Facebook", "icon": "facebook"},
    {"name": "Instagram", "icon": "instagram"},
    {"name": "LinkedIn", "icon": "linkedin"},
    {"name": "Twitter / X", "icon": "twitter"},
    {"name": "TikTok", "icon": "tiktok"},
    {"name": "YouTube", "icon": "youtube"},
    {"name": "WhatsApp", "icon": "message-circle"},
    {"name": "Telegram", "icon": "send"},
    {"name": "Website", "icon": "globe"},
]


@api_router.get("/settings/social-platforms")
async def list_social_platforms(current_user: dict = Depends(get_current_user)):
    items = await db.social_platforms.find({}, {"_id": 0}).sort("name", 1).to_list(200)
    if not items:
        seeded = []
        for d in DEFAULT_SOCIAL_PLATFORMS:
            doc = {"id": str(uuid.uuid4()), "name": d["name"], "icon": d["icon"]}
            await db.social_platforms.insert_one(doc)
            doc.pop("_id", None)
            seeded.append(doc)
        return seeded
    return items


@api_router.post("/settings/social-platforms")
async def create_social_platform(data: dict, current_user: dict = Depends(get_current_user)):
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Ad boş ola bilməz")
    icon = (data.get("icon") or "globe").strip().lower()[:32]
    if await db.social_platforms.find_one({"name": name}):
        raise HTTPException(status_code=400, detail="Bu ad artıq mövcuddur")
    doc = {"id": str(uuid.uuid4()), "name": name, "icon": icon}
    await db.social_platforms.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/settings/social-platforms/{item_id}")
async def update_social_platform(item_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    update: Dict[str, Any] = {}
    if "name" in data:
        n = (data.get("name") or "").strip()
        if not n:
            raise HTTPException(status_code=400, detail="Ad boş ola bilməz")
        update["name"] = n
    if "icon" in data:
        update["icon"] = (data.get("icon") or "globe").strip().lower()[:32]
    if not update:
        raise HTTPException(status_code=400, detail="Yenilənəcək məlumat yoxdur")
    result = await db.social_platforms.update_one({"id": item_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tapılmadı")
    return await db.social_platforms.find_one({"id": item_id}, {"_id": 0})


@api_router.delete("/settings/social-platforms/{item_id}")
async def delete_social_platform(item_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.social_platforms.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tapılmadı")
    return {"message": "Silindi"}

@api_router.post("/finance/expenses")
async def create_expense(expense_data: ExpenseCreate, current_user: dict = Depends(check_permission("finance", "write"))):
    expense_id = str(uuid.uuid4())
    expense_doc = {
        "id": expense_id,
        **expense_data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.expenses.insert_one(expense_doc)
    expense_doc.pop("_id", None)
    return expense_doc

@api_router.put("/finance/expenses/{expense_id}")
async def update_expense(expense_id: str, expense_data: dict, current_user: dict = Depends(check_permission("finance", "write"))):
    update_data = {k: v for k, v in expense_data.items() if v is not None}
    result = await db.expenses.update_one({"id": expense_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Xərc tapılmadı")
    expense = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    return expense

@api_router.delete("/finance/expenses/{expense_id}")
async def delete_expense(expense_id: str, current_user: dict = Depends(check_permission("finance", "write"))):
    result = await db.expenses.delete_one({"id": expense_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Xərc tapılmadı")
    return {"message": "Xərc silindi"}

# ==================== BARTER ƏMƏLİYYATLARI ====================

BARTER_STATUSES = ["Təklif", "Müzakirədə", "Aktiv", "Tamamlandı", "Ləğv edilib"]

@api_router.get("/barters")
async def get_barters(
    status: Optional[str] = None,
    partner_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status and status != "all":
        query["status"] = status
    if partner_id:
        query["partner_id"] = partner_id
    items = await db.barters.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items

@api_router.post("/barters")
async def create_barter(data: dict, current_user: dict = Depends(check_permission("finance", "write"))):
    count = await db.barters.count_documents({})
    doc = {
        "id": str(uuid.uuid4()),
        "barter_code": f"B-{str(count + 1).zfill(3)}",
        "partner_id": data.get("partner_id", ""),
        "partner_name": data.get("partner_name", ""),
        "partner_contact": data.get("partner_contact", ""),
        "partner_phone": data.get("partner_phone", ""),
        "our_service": data.get("our_service", ""),
        "their_service": data.get("their_service", ""),
        "our_value": float(data.get("our_value", 0) or 0),
        "their_value": float(data.get("their_value", 0) or 0),
        "status": data.get("status", "Təklif"),
        "start_date": data.get("start_date", ""),
        "end_date": data.get("end_date", ""),
        "notes": data.get("notes", ""),
        "responsible": data.get("responsible", current_user.get("name", "")),
        "created_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.barters.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/barters/{barter_id}")
async def update_barter(barter_id: str, data: dict, current_user: dict = Depends(check_permission("finance", "write"))):
    update = {k: v for k, v in data.items() if k not in ("id", "barter_code", "created_at", "created_by")}
    if "our_value" in update:
        update["our_value"] = float(update["our_value"] or 0)
    if "their_value" in update:
        update["their_value"] = float(update["their_value"] or 0)
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.barters.update_one({"id": barter_id}, {"$set": update})
    doc = await db.barters.find_one({"id": barter_id}, {"_id": 0})
    return doc

@api_router.delete("/barters/{barter_id}")
async def delete_barter(barter_id: str, current_user: dict = Depends(check_permission("finance", "write"))):
    await db.barters.delete_one({"id": barter_id})
    return {"message": "Barter silindi"}

@api_router.get("/barters/stats")
async def barter_stats(current_user: dict = Depends(get_current_user)):
    items = await db.barters.find({}, {"_id": 0}).to_list(5000)
    by_status = {s: 0 for s in BARTER_STATUSES}
    total_our = 0
    total_their = 0
    active_count = 0
    for b in items:
        st = b.get("status", "Təklif")
        if st in by_status:
            by_status[st] += 1
        if st in ("Aktiv", "Tamamlandı"):
            total_our += b.get("our_value", 0) or 0
            total_their += b.get("their_value", 0) or 0
        if st == "Aktiv":
            active_count += 1
    return {
        "total": len(items),
        "active": active_count,
        "by_status": by_status,
        "total_our_value": round(total_our, 2),
        "total_their_value": round(total_their, 2),
        "net_balance": round(total_their - total_our, 2)
    }

@api_router.get("/finance/summary")
async def get_finance_summary(current_user: dict = Depends(get_current_user)):
    income_pipeline = [
        {"$group": {"_id": None, "total": {"$sum": "$amount"}, "paid": {"$sum": "$paid_amount"}, "debt": {"$sum": "$debt_amount"}}}
    ]
    expense_pipeline = [
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    
    income_data = await db.incomes.aggregate(income_pipeline).to_list(1)
    expense_data = await db.expenses.aggregate(expense_pipeline).to_list(1)
    
    total_income = income_data[0]["total"] if income_data else 0
    total_paid = income_data[0]["paid"] if income_data else 0
    total_debt = income_data[0]["debt"] if income_data else 0
    total_expenses = expense_data[0]["total"] if expense_data else 0
    
    return {
        "total_income": total_income,
        "paid_income": total_paid,
        "debt": total_debt,
        "total_expenses": total_expenses,
        "net_profit": total_income - total_expenses,
        "current_profit": total_paid - total_expenses,
        "currency": "AZN"
    }

# ==================== TASKS (TAPŞIRIQLAR) ====================

@api_router.get("/tasks")
async def get_tasks(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    assignee: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status and status != "all":
        query["status"] = status
    if priority and priority != "all":
        query["priority"] = priority
    if assignee and assignee != "all":
        query["assignee"] = {"$regex": assignee, "$options": "i"}
    query = await apply_scope(query, current_user, "tasks")
    tasks = await db.tasks.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return tasks

@api_router.post("/tasks")
async def create_task(task_data: TaskCreate, current_user: dict = Depends(check_permission("tasks", "write"))):
    # İcraçı şöbəsi məcburi sahədir
    if not (task_data.department or "").strip():
        raise HTTPException(status_code=400, detail="İcraçı şöbə seçilməlidir")
    # Idempotency guard: if the SAME user submitted the SAME task name within the
    # last 10 seconds, return the existing record instead of inserting a duplicate.
    # Protects against double-clicks, retried POSTs (e.g. flaky 3G), and React
    # StrictMode dev double-renders.
    creator_name = current_user.get("name", "")
    now_dt = datetime.now(timezone.utc)
    cutoff = (now_dt - timedelta(seconds=10)).isoformat()
    dupe = await db.tasks.find_one(
        {
            "created_by": creator_name,
            "task_name": task_data.task_name,
            "created_at": {"$gte": cutoff},
        },
        {"_id": 0},
    )
    if dupe:
        logging.info("Duplicate task suppressed for %s: %s", creator_name, task_data.task_name)
        return dupe

    count = await db.tasks.count_documents({})
    task_code = f"T-{str(count + 1).zfill(3)}"
    task_doc = {
        "id": str(uuid.uuid4()),
        "task_code": task_code,
        **task_data.model_dump(),
        "created_by": creator_name,
        "creator_department": current_user.get("department", "") or "",
        "marsol_company": (current_user.get("marsol_company") or "").strip(),
        "created_at": now_dt.isoformat()
    }
    await db.tasks.insert_one(task_doc)
    task_doc.pop("_id", None)

    # In-app notification to assignee(s) + responsible_person(s) (excluding the creator)
    recipients = set()
    actor_name = current_user.get("name", "")
    for nm in _as_name_list(task_doc.get("assignee")):
        if nm and nm != actor_name:
            recipients.add(nm)
    for nm in _as_name_list(task_doc.get("responsible_person")):
        if nm and nm != actor_name:
            recipients.add(nm)
    notif_body = (
        f"Yeni tapşırıq sizə təyin olundu: {task_doc['task_name']} "
        f"(Prioritet: {task_doc.get('priority', '')}, "
        f"Bitmə: {task_doc.get('end_date', '—')})"
    )
    for r_name in recipients:
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "type": "task_assigned",
            "title": "Yeni tapşırıq",
            "body": notif_body,
            "task_id": task_doc["id"],
            "task_code": task_code,
            "recipient_name": r_name,
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    if recipients:
        _safe_push(list(recipients), "Yeni tapşırıq", notif_body, link=f"/tasks?id={task_doc['id']}", data={"type": "task_assigned", "task_id": task_doc["id"]})

    # Email notification
    recipient_emails = []
    for r_name in recipients:
        em = await _user_email_by_name(r_name)
        if em:
            recipient_emails.append(em)
    if recipient_emails:
        body_html = f"""<p>Sizə yeni tapşırıq təyin olunub:</p>
        <table cellpadding='6' cellspacing='0' style='width:100%;border:1px solid #e2e8f0;border-radius:8px;font-size:13px'>
          <tr><td style='color:#64748b'>Tapşırıq</td><td style='font-weight:600'>{task_doc.get('task_name', '')}</td></tr>
          <tr><td style='color:#64748b'>Kod</td><td>{task_code}</td></tr>
          <tr><td style='color:#64748b'>Prioritet</td><td>{task_doc.get('priority', '')}</td></tr>
          <tr><td style='color:#64748b'>Başlanğıc</td><td>{task_doc.get('start_date') or '—'}</td></tr>
          <tr><td style='color:#64748b'>Bitmə</td><td>{task_doc.get('end_date') or '—'}</td></tr>
          <tr><td style='color:#64748b'>Yaradıcı</td><td>{task_doc.get('created_by', '')}</td></tr>
        </table>"""
        await _email_notify_safe(
            title=f"Yeni tapşırıq: {task_doc.get('task_name', '')}",
            body_html=body_html,
            extra_recipients=recipient_emails,
        )

    return task_doc

@api_router.put("/tasks/{task_id}")
async def update_task(task_id: str, task_data: dict, current_user: dict = Depends(check_permission("tasks", "write"))):
    existing = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Tapşırıq tapılmadı")
    await assert_scope_ownership(current_user, "tasks", existing)
    update_data = {k: v for k, v in task_data.items() if v is not None}

    # ---------------------------------------------------------------------
    # RBAC: Yalnız tapşırığı YARADAN və ya admin TAM redaktə edə bilər.
    # Assignee/Responsible (təyin olunan şəxs) yalnız "status", "subtasks"
    # (checklist toggle) və "result" (nəticə mətni) sahələrini yeniləyə bilər.
    # ---------------------------------------------------------------------
    me = (current_user.get("name") or "").strip()
    is_admin = (current_user.get("role") or "").lower() == "admin"
    creator = (existing.get("created_by") or "").strip()
    is_creator = creator and creator == me
    ASSIGNEE_ALLOWED_FIELDS = {"status", "subtasks", "result"}
    if not is_admin and not is_creator:
        forbidden = [k for k in update_data.keys() if k not in ASSIGNEE_ALLOWED_FIELDS]
        if forbidden:
            raise HTTPException(
                status_code=403,
                detail="Yalnız tapşırığı yaradan və ya admin redaktə edə bilər"
            )

    # Stamp completed_at the first time status flips to "Tamamlandı" so that
    # the auto-archive job has a reliable date marker.
    new_status = (update_data.get("status") or "").strip()
    old_status = (existing.get("status") or "").strip()
    if new_status == "Tamamlandı" and old_status != "Tamamlandı" and not existing.get("completed_at"):
        update_data["completed_at"] = datetime.now(timezone.utc).isoformat()

    result = await db.tasks.update_one({"id": task_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tapşırıq tapılmadı")
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})

    # When the creator (or anyone else) modifies a meaningful field, notify the
    # assignee + responsible_person (excluding whoever made the change).
    actor = (current_user.get("name") or "").strip()
    tracked_fields = ("task_name", "priority", "start_date", "end_date", "status", "notes", "phase", "assignee", "responsible_person", "subtasks")
    changes = []
    for k in tracked_fields:
        if k in update_data and existing.get(k) != update_data[k]:
            label = {
                "task_name": "Ad", "priority": "Prioritet", "start_date": "Başlama tarixi",
                "end_date": "Bitmə tarixi", "status": "Status", "notes": "Qeyd",
                "phase": "Mərhələ", "assignee": "İcraçı", "responsible_person": "Məsul",
                "subtasks": "Alt mərhələlər",
            }.get(k, k)
            changes.append(label)
    if changes:
        recipients = set()
        for nm in _as_name_list(task.get("assignee")):
            if nm and nm != actor:
                recipients.add(nm)
        for nm in _as_name_list(task.get("responsible_person")):
            if nm and nm != actor:
                recipients.add(nm)
        body = f"\"{task.get('task_name','')}\" tapşırığı yeniləndi: {', '.join(changes)}. Dəyişikliyi edən: {actor or '—'}"
        for r_name in recipients:
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()),
                "type": "task_updated",
                "title": "Tapşırıq yeniləndi",
                "body": body,
                "task_id": task["id"],
                "task_code": task.get("task_code", ""),
                "recipient_name": r_name,
                "is_read": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
        if recipients:
            _safe_push(list(recipients), "Tapşırıq yeniləndi", body, link=f"/tasks?id={task['id']}", data={"type": "task_updated", "task_id": task["id"]})
    return task

@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, current_user: dict = Depends(check_permission("tasks", "write"))):
    existing = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Tapşırıq tapılmadı")
    await assert_scope_ownership(current_user, "tasks", existing)
    # Only the creator OR admin may delete a task. Assignees can update status but
    # may NOT delete tasks created for them by others.
    is_admin = (current_user.get("role") or "").lower() == "admin"
    creator = (existing.get("created_by") or "").strip()
    me = (current_user.get("name") or "").strip()
    if not is_admin and creator and creator != me:
        raise HTTPException(status_code=403, detail="Yalnız tapşırığı yaradan və ya admin silə bilər")
    # Archive the task before deleting
    archived = {**existing}
    archived.pop("_id", None)
    archived["archived_at"] = datetime.now(timezone.utc).isoformat()
    archived["archived_by"] = current_user.get("name", "")
    archived["archive_id"] = str(uuid.uuid4())
    await db.tasks_archive.insert_one(archived)
    result = await db.tasks.delete_one({"id": task_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tapşırıq tapılmadı")
    return {"message": "Tapşırıq arxivləndi"}


@api_router.post("/tasks/bulk-delete")
async def bulk_delete_tasks(payload: dict, current_user: dict = Depends(check_permission("tasks", "write"))):
    """Bulk-delete (archive) selected tasks. Admin can delete any; non-admins
    can only delete tasks they created."""
    ids = payload.get("ids", []) if isinstance(payload, dict) else []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="ids massivi boşdur")
    is_admin = (current_user.get("role") or "").lower() == "admin"
    me = (current_user.get("name") or "").strip()
    docs = await db.tasks.find({"id": {"$in": ids}}, {"_id": 0}).to_list(len(ids))
    deletable = [d for d in docs if is_admin or (d.get("created_by") or "") == me]
    deletable_ids = [d["id"] for d in deletable]
    if not deletable_ids:
        raise HTTPException(status_code=403, detail="Heç bir tapşırığı silmək icazəniz yoxdur")
    now_iso = datetime.now(timezone.utc).isoformat()
    archive_docs = []
    for d in deletable:
        archive_docs.append({**d, "archive_id": str(uuid.uuid4()), "archived_at": now_iso, "archived_by": me})
    if archive_docs:
        await db.tasks_archive.insert_many(archive_docs)
    res = await db.tasks.delete_many({"id": {"$in": deletable_ids}})
    return {"deleted": res.deleted_count, "skipped": len(ids) - len(deletable_ids)}


@api_router.get("/tasks/{task_id}/comments")
async def list_task_comments(task_id: str, current_user: dict = Depends(check_permission("tasks", "read"))):
    """List comments on a single task, oldest first."""
    items = await db.task_comments.find({"task_id": task_id}, {"_id": 0}).sort("created_at", 1).to_list(500)
    return items


@api_router.post("/tasks/{task_id}/comments")
async def add_task_comment(task_id: str, payload: dict, current_user: dict = Depends(check_permission("tasks", "write"))):
    """Append a comment to a task and notify the other stakeholders."""
    text = (payload.get("text") or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Şərh boş ola bilməz")
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Tapşırıq tapılmadı")
    comment = {
        "id": str(uuid.uuid4()),
        "task_id": task_id,
        "author_name": current_user.get("name", ""),
        "author_id": current_user.get("id", ""),
        "text": text,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.task_comments.insert_one(comment)
    comment.pop("_id", None)

    # Notify other stakeholders (creator + assignee(s) + responsible_person(s))
    actor = current_user.get("name", "")
    recipients = set()
    for nm in _as_name_list(task.get("assignee")):
        if nm and nm != actor:
            recipients.add(nm)
    for nm in _as_name_list(task.get("responsible_person")):
        if nm and nm != actor:
            recipients.add(nm)
    cb = (task.get("created_by") or "").strip()
    if cb and cb != actor:
        recipients.add(cb)
    for r_name in recipients:
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "type": "task_comment",
            "title": f"Yeni şərh — {task.get('task_name', '')}",
            "body": f"{actor}: {text[:160]}",
            "task_id": task_id,
            "recipient_name": r_name,
            "is_read": False,
            "created_at": comment["created_at"],
        })
    if recipients:
        _safe_push(list(recipients), f"Yeni şərh — {task.get('task_name','')}", f"{actor}: {text[:160]}", link=f"/tasks?id={task_id}", data={"type": "task_comment", "task_id": task_id})
    return comment


@api_router.delete("/tasks/{task_id}/comments/{comment_id}")
async def delete_task_comment(task_id: str, comment_id: str, current_user: dict = Depends(check_permission("tasks", "write"))):
    """Delete a comment (author or admin only)."""
    c = await db.task_comments.find_one({"id": comment_id, "task_id": task_id}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Şərh tapılmadı")
    is_admin = (current_user.get("role") or "").lower() == "admin"
    if c.get("author_id") != current_user.get("id") and not is_admin:
        raise HTTPException(status_code=403, detail="Yalnız müəllif və ya admin silə bilər")
    await db.task_comments.delete_one({"id": comment_id})
    return {"deleted": True}


@api_router.get("/tasks/archive")
async def list_archived_tasks(current_user: dict = Depends(check_permission("tasks", "read"))):
    """List archived (soft-deleted) tasks. Admin sees all; others see only those
    they created or were assigned to."""
    is_admin = (current_user.get("role") or "").lower() == "admin"
    me = (current_user.get("name") or "").strip()
    q: Dict[str, Any] = {}
    if not is_admin and me:
        q = {"$or": [{"created_by": me}, {"assignee": me}]}
    items = await db.tasks_archive.find(q, {"_id": 0}).sort("archived_at", -1).to_list(500)
    return items


@api_router.post("/tasks/archive/{archive_id}/restore")
async def restore_archived_task(archive_id: str, current_user: dict = Depends(check_permission("tasks", "write"))):
    """Restore an archived task back to the active tasks collection. Admin
    or original creator only."""
    archived = await db.tasks_archive.find_one({"archive_id": archive_id}, {"_id": 0})
    if not archived:
        raise HTTPException(status_code=404, detail="Arxiv qeydi tapılmadı")
    is_admin = (current_user.get("role") or "").lower() == "admin"
    creator = (archived.get("created_by") or "").strip()
    me = (current_user.get("name") or "").strip()
    if not is_admin and creator and creator != me:
        raise HTTPException(status_code=403, detail="Yalnız tapşırığı yaradan və ya admin bərpa edə bilər")
    restored = {k: v for k, v in archived.items() if k not in ("archived_at", "archived_by", "archive_id")}
    await db.tasks.insert_one(restored)
    await db.tasks_archive.delete_one({"archive_id": archive_id})
    restored.pop("_id", None)
    return restored


@api_router.delete("/tasks/archive/{archive_id}")
async def delete_archived_task(archive_id: str, current_user: dict = Depends(check_permission("tasks", "write"))):
    """Permanently delete a single archived task. Admin only."""
    if (current_user.get("role") or "").lower() != "admin":
        raise HTTPException(status_code=403, detail="Yalnız admin arxivdən silə bilər")
    archived = await db.tasks_archive.find_one({"archive_id": archive_id}, {"_id": 0})
    if not archived:
        raise HTTPException(status_code=404, detail="Arxiv qeydi tapılmadı")
    await db.tasks_archive.delete_one({"archive_id": archive_id})
    return {"message": "Arxiv qeydi silindi", "archive_id": archive_id}


@api_router.post("/tasks/archive/bulk-delete")
async def bulk_delete_archived_tasks(data: dict, current_user: dict = Depends(check_permission("tasks", "write"))):
    """Permanently delete multiple archived tasks at once. Admin only."""
    if (current_user.get("role") or "").lower() != "admin":
        raise HTTPException(status_code=403, detail="Yalnız admin arxivdən silə bilər")
    archive_ids = data.get("archive_ids") or []
    if not isinstance(archive_ids, list) or not archive_ids:
        raise HTTPException(status_code=400, detail="archive_ids siyahısı boş ola bilməz")
    result = await db.tasks_archive.delete_many({"archive_id": {"$in": archive_ids}})
    return {"deleted": result.deleted_count, "skipped": 0, "skipped_ids": []}

# ==================== MEETINGS (GÖRÜŞLƏR) ====================

@api_router.get("/meetings")
async def get_meetings(
    meeting_type: Optional[str] = None,
    department: Optional[str] = None,
    employee: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if meeting_type and meeting_type != "all":
        query["meeting_type"] = meeting_type
    if department and department != "all":
        query["department"] = department
    if employee and employee != "all":
        query["employee"] = employee
    if date_from:
        query.setdefault("date", {})["$gte"] = date_from
    if date_to:
        query.setdefault("date", {})["$lte"] = date_to
    query = await apply_scope(query, current_user, "meetings")
    
    meetings = await db.meetings.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return meetings

@api_router.post("/meetings")
async def create_meeting(data: dict, current_user: dict = Depends(check_permission("meetings", "write"))):
    meeting_id = str(uuid.uuid4())
    meeting_doc = {
        "id": meeting_id,
        "employee": data.get("employee", ""),
        "meeting_setter": data.get("meeting_setter", ""),
        "date": data.get("date", ""),
        "time": data.get("time", ""),
        "company": data.get("company", ""),
        "contact_person": data.get("contact_person", ""),
        "project": data.get("project", ""),
        "meeting_type": data.get("meeting_type", ""),
        "meeting_mode": data.get("meeting_mode", "Offline"),
        "department": data.get("department", ""),
        "location": data.get("location", ""),
        "result": data.get("result", ""),
        "next_meeting": data.get("next_meeting", ""),
        "notes": data.get("notes", ""),
        "reminders": data.get("reminders", []),
        "created_by": current_user.get("name", ""),
        "marsol_company": (current_user.get("marsol_company") or "").strip(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.meetings.insert_one(meeting_doc)
    meeting_doc.pop("_id", None)
    # Auto-track service usage for this company's package
    await _auto_track_meeting_service(meeting_doc)
    # Create notification for each reminder
    for rem in meeting_doc.get("reminders", []):
        notif_doc = {
            "id": str(uuid.uuid4()),
            "title": f"Görüş xatırlatması: {meeting_doc['meeting_type']}",
            "message": f"{meeting_doc['date']} {meeting_doc['time']} - {meeting_doc['company'] or meeting_doc['employee']}. {rem.get('note', '')}",
            "type": "reminder",
            "meeting_id": meeting_id,
            "reminder_date": rem.get("date", ""),
            "reminder_time": rem.get("time", ""),
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notif_doc)
    
    # Email notify the meeting employee + setter (and admin)
    emp_email = await _user_email_by_name(meeting_doc.get("employee", ""))
    setter_email = await _user_email_by_name(meeting_doc.get("meeting_setter", ""))
    body = f"""<p>Yeni görüş təyin edildi:</p>
    <table cellpadding='6' cellspacing='0' style='width:100%;border:1px solid #e2e8f0;border-radius:8px;font-size:13px'>
      <tr><td style='color:#64748b'>Tarix</td><td style='font-weight:600'>{meeting_doc.get('date', '')} {meeting_doc.get('time', '')}</td></tr>
      <tr><td style='color:#64748b'>Şirkət</td><td>{meeting_doc.get('company') or '—'}</td></tr>
      <tr><td style='color:#64748b'>Növ</td><td>{meeting_doc.get('meeting_type', '')}</td></tr>
      <tr><td style='color:#64748b'>Məkan</td><td>{meeting_doc.get('location') or meeting_doc.get('meeting_mode', '')}</td></tr>
      <tr><td style='color:#64748b'>Əməkdaş</td><td>{meeting_doc.get('employee', '')}</td></tr>
    </table>"""
    await _email_notify_safe(
        title=f"Görüş təyin edildi: {meeting_doc.get('date', '')}",
        body_html=body,
        extra_recipients=[emp_email, setter_email],
    )
    return meeting_doc

@api_router.put("/meetings/{meeting_id}")
async def update_meeting(meeting_id: str, data: dict, current_user: dict = Depends(check_permission("meetings", "write"))):
    existing = await db.meetings.find_one({"id": meeting_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Görüş tapılmadı")
    await assert_scope_ownership(current_user, "meetings", existing)
    update_data = {k: v for k, v in data.items() if k != "id"}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.meetings.update_one({"id": meeting_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Görüş tapılmadı")
    # Update reminders as notifications
    await db.notifications.delete_many({"meeting_id": meeting_id, "type": "reminder"})
    meeting = await db.meetings.find_one({"id": meeting_id}, {"_id": 0})
    for rem in update_data.get("reminders", []):
        notif_doc = {
            "id": str(uuid.uuid4()),
            "title": f"Görüş xatırlatması: {meeting.get('meeting_type', '')}",
            "message": f"{meeting.get('date', '')} {meeting.get('time', '')} - {meeting.get('company', '') or meeting.get('employee', '')}. {rem.get('note', '')}",
            "type": "reminder",
            "meeting_id": meeting_id,
            "reminder_date": rem.get("date", ""),
            "reminder_time": rem.get("time", ""),
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.notifications.insert_one(notif_doc)
    return meeting

@api_router.delete("/meetings/{meeting_id}")
async def delete_meeting(meeting_id: str, current_user: dict = Depends(check_permission("meetings", "write"))):
    existing = await db.meetings.find_one({"id": meeting_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Görüş tapılmadı")
    await assert_scope_ownership(current_user, "meetings", existing)
    result = await db.meetings.delete_one({"id": meeting_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Görüş tapılmadı")
    await db.notifications.delete_many({"meeting_id": meeting_id, "type": "reminder"})
    return {"message": "Görüş silindi"}


# ==================== MEETING REQUEST FLOW (internal user-to-user) ====================

@api_router.post("/meeting-requests")
async def create_meeting_request(data: dict, current_user: dict = Depends(check_permission("meetings", "write"))):
    """Send a meeting request to one or more internal users.

    Body: { date, time, meeting_type, meeting_mode, location, notes, recipient_ids: [user_id,...] }
    Creates a `meeting_requests` document with status='pending' and a notification for each recipient.
    """
    recipient_ids = data.get("recipient_ids", [])
    if not recipient_ids or not isinstance(recipient_ids, list):
        raise HTTPException(status_code=400, detail="Ən az 1 iştirakçı seçilməlidir")
    recipients = await db.users.find({"id": {"$in": recipient_ids}}, {"_id": 0, "password": 0}).to_list(100)
    if not recipients:
        raise HTTPException(status_code=400, detail="İştirakçılar tapılmadı")
    req_id = str(uuid.uuid4())
    sender_id = current_user.get("id") or current_user.get("user_id")
    sender_name = current_user.get("name", "")
    now_iso = datetime.now(timezone.utc).isoformat()
    request_doc = {
        "id": req_id,
        "sender_id": sender_id,
        "sender_name": sender_name,
        "recipients": [
            {"id": r["id"], "name": r.get("name", ""), "email": r.get("email", ""), "status": "pending", "responded_at": None}
            for r in recipients
        ],
        "date": (data.get("date") or "").strip(),
        "time": (data.get("time") or "").strip(),
        "meeting_type": (data.get("meeting_type") or "").strip(),
        "meeting_mode": (data.get("meeting_mode") or "Offline").strip(),
        "location": (data.get("location") or "").strip(),
        "notes": (data.get("notes") or "").strip(),
        "status": "pending",  # pending | accepted | rejected | cancelled
        "created_at": now_iso,
    }
    await db.meeting_requests.insert_one(request_doc)
    request_doc.pop("_id", None)
    # Notify each recipient
    recipient_names_push = []
    for r in recipients:
        notif = {
            "id": str(uuid.uuid4()),
            "title": f"Görüş təklifi: {sender_name}",
            "message": f"{sender_name} sizinlə görüş təklif etdi — {request_doc['date']} {request_doc['time']} · {request_doc['meeting_type'] or 'Görüş'}",
            "type": "meeting_request",
            "meeting_request_id": req_id,
            "recipient_id": r["id"],
            "is_read": False,
            "created_at": now_iso,
        }
        await db.notifications.insert_one(notif)
        if r.get("name"):
            recipient_names_push.append(r["name"])
    if recipient_names_push:
        _safe_push(
            recipient_names_push,
            f"Görüş təklifi: {sender_name}",
            f"{request_doc['date']} {request_doc['time']} · {request_doc['meeting_type'] or 'Görüş'}",
            link="/meetings",
            data={"type": "meeting_request", "request_id": req_id},
        )
    return request_doc


@api_router.get("/meeting-requests")
async def list_meeting_requests(current_user: dict = Depends(check_permission("meetings", "read"))):
    """Return requests sent BY or addressed TO the current user."""
    uid = current_user.get("id") or current_user.get("user_id")
    cursor = db.meeting_requests.find(
        {"$or": [{"sender_id": uid}, {"recipients.id": uid}]},
        {"_id": 0},
    ).sort("created_at", -1)
    items = await cursor.to_list(500)
    return items


@api_router.post("/meeting-requests/{req_id}/respond")
async def respond_to_meeting_request(req_id: str, data: dict, current_user: dict = Depends(check_permission("meetings", "write"))):
    """Recipient accepts or rejects a meeting request.

    Body: { action: 'accept' | 'reject' }
    When ALL recipients accept → status='accepted' and meeting docs are inserted for sender + all recipients.
    Any rejection → status='rejected', no meeting created.
    """
    action = (data.get("action") or "").lower()
    if action not in ("accept", "reject"):
        raise HTTPException(status_code=400, detail="action 'accept' və ya 'reject' olmalıdır")
    req = await db.meeting_requests.find_one({"id": req_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Görüş təklifi tapılmadı")
    uid = current_user.get("id") or current_user.get("user_id")
    # Find this recipient
    found = False
    for r in req["recipients"]:
        if r["id"] == uid:
            r["status"] = "accepted" if action == "accept" else "rejected"
            r["responded_at"] = datetime.now(timezone.utc).isoformat()
            found = True
            break
    if not found:
        raise HTTPException(status_code=403, detail="Bu təklif sizin üçün deyil")
    # Update overall status
    statuses = [r["status"] for r in req["recipients"]]
    if "rejected" in statuses:
        req["status"] = "rejected"
    elif all(s == "accepted" for s in statuses):
        req["status"] = "accepted"
    else:
        req["status"] = "pending"
    await db.meeting_requests.update_one({"id": req_id}, {"$set": {
        "recipients": req["recipients"],
        "status": req["status"],
    }})
    # Mark all related notifications as read for this user
    await db.notifications.update_many(
        {"meeting_request_id": req_id, "recipient_id": uid},
        {"$set": {"is_read": True}},
    )
    # If now fully accepted → insert ONE shared meeting visible to all participants
    if req["status"] == "accepted":
        sender = await db.users.find_one({"id": req["sender_id"]}, {"_id": 0, "password": 0})
        participants = [
            {"id": req["sender_id"], "name": req["sender_name"], "email": sender.get("email") if sender else ""},
            *[{"id": r["id"], "name": r["name"], "email": r.get("email", "")} for r in req["recipients"]],
        ]
        names = ", ".join({p["name"] for p in participants if p["name"]})
        now_iso = datetime.now(timezone.utc).isoformat()
        # Single shared meeting doc; `employee` lists everyone (comma-separated)
        # and `participant_ids`/`participant_names` enable scope filtering.
        meeting_doc = {
            "id": str(uuid.uuid4()),
            "employee": names,  # display: comma-separated participant list
            "meeting_setter": req["sender_name"],
            "date": req.get("date", ""),
            "time": req.get("time", ""),
            "company": "",
            "contact_person": "",
            "project": "",
            "meeting_type": req.get("meeting_type", ""),
            "meeting_mode": req.get("meeting_mode", "Offline"),
            "department": "",
            "location": req.get("location", ""),
            "result": "",
            "next_meeting": "",
            "notes": (req.get("notes", "") + (f"\nİştirakçılar: {names}" if names else "")).strip(),
            "reminders": [],
            "created_by": req["sender_name"],
            "created_at": now_iso,
            "meeting_request_id": req_id,
            "is_internal": True,
            "participant_ids": [p["id"] for p in participants if p.get("id")],
            "participant_names": [p["name"] for p in participants if p.get("name")],
        }
        await db.meetings.insert_one(meeting_doc)
        # Notify sender that all accepted
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "title": "Görüş təklifi qəbul edildi",
            "message": f"{names} sizinlə görüşməyi qəbul etdi — {req['date']} {req['time']}",
            "type": "meeting_request_accepted",
            "meeting_request_id": req_id,
            "recipient_id": req["sender_id"],
            "is_read": False,
            "created_at": now_iso,
        })
        if req.get("sender_name"):
            _safe_push([req["sender_name"]], "Görüş təklifi qəbul edildi", f"{names} — {req['date']} {req['time']}", link="/meetings", data={"type": "meeting_request_accepted"})
    elif req["status"] == "rejected":
        rejecter = current_user.get("name", "")
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "title": "Görüş təklifi rədd edildi",
            "message": f"{rejecter} görüş təklifinizi rədd etdi — {req['date']} {req['time']}",
            "type": "meeting_request_rejected",
            "meeting_request_id": req_id,
            "recipient_id": req["sender_id"],
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        if req.get("sender_name"):
            _safe_push([req["sender_name"]], "Görüş təklifi rədd edildi", f"{rejecter} təklifinizi rədd etdi", link="/meetings", data={"type": "meeting_request_rejected"})
    return req



# ==================== PROJECT EVENTS (LAYİHƏLƏR/TƏDBİRLƏR) ====================

@api_router.get("/project-events")
async def get_project_events(current_user: dict = Depends(get_current_user)):
    query = await apply_scope({}, current_user, "projects")
    events = await db.project_events.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    for e in events:
        e["guest_count"] = await db.event_invitations.count_documents({"event_id": e["id"]})
        e["attended_count"] = await db.event_invitations.count_documents({"event_id": e["id"], "status": "İştirak etdi"})
    return events

@api_router.post("/project-events")
async def create_project_event(data: dict, current_user: dict = Depends(check_permission("projects", "write"))):
    doc = {
        "id": str(uuid.uuid4()),
        "name": data.get("name", ""),
        "type": data.get("type", ""),
        "date": data.get("date", ""),
        "end_date": data.get("end_date", ""),
        "location": data.get("location", ""),
        "description": data.get("description", ""),
        "status": data.get("status", "Planlaşdırılır"),
        "price_per_sqm": data.get("price_per_sqm"),
        "total_price": data.get("total_price"),
        "created_by": current_user.get("name", ""),
        "marsol_company": (current_user.get("marsol_company") or "").strip(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.project_events.insert_one(doc)
    doc.pop("_id", None)
    doc["guest_count"] = 0
    doc["attended_count"] = 0
    return doc

@api_router.put("/project-events/{event_id}")
async def update_project_event(event_id: str, data: dict, current_user: dict = Depends(check_permission("projects", "write"))):
    existing = await db.project_events.find_one({"id": event_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Layihə tapılmadı")
    await assert_scope_ownership(current_user, "projects", existing)
    update = {k: v for k, v in data.items() if k not in ("id",)}
    await db.project_events.update_one({"id": event_id}, {"$set": update})
    doc = await db.project_events.find_one({"id": event_id}, {"_id": 0})
    return doc

@api_router.delete("/project-events/{event_id}")
async def delete_project_event(event_id: str, current_user: dict = Depends(check_permission("projects", "write"))):
    existing = await db.project_events.find_one({"id": event_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Layihə tapılmadı")
    await assert_scope_ownership(current_user, "projects", existing)
    await db.project_events.delete_one({"id": event_id})
    await db.event_invitations.delete_many({"event_id": event_id})
    return {"message": "Layihə silindi"}

@api_router.get("/project-events/{event_id}/sales")
async def get_project_sales(event_id: str, current_user: dict = Depends(get_current_user)):
    """Return all sales-leads linked to this project (status Satıldı or Üzv oldu)."""
    event = await db.project_events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Layihə tapılmadı")
    query = {
        "project_id": event_id,
        "status": {"$in": ["Satıldı", "Üzv oldu"]}
    }
    query = await apply_scope(query, current_user, "sales")
    sales = await db.sales_leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    # Enrich each sale with sector info from companies if linked by name
    for s in sales:
        comp = await db.companies.find_one({"brand_name": s.get("company_name", "")}, {"_id": 0, "sector": 1, "sub_sector": 1})
        s["sector"] = comp.get("sector", "") if comp else ""
        s["sub_sector"] = comp.get("sub_sector", "") if comp else ""
        # Compute debt
        total = s.get("total_amount") or 0
        paid = s.get("paid_amount") or 0
        s["debt_amount"] = max(float(total) - float(paid), 0)
    return {"event": event, "sales": sales}

@api_router.post("/sales-leads/{lead_id}/payment")
async def add_lead_payment(lead_id: str, data: dict, current_user: dict = Depends(check_permission("sales", "write"))):
    """Append a payment to a lead and update paid_amount & metadata (contract, e-qaimə, follow-up)."""
    lead = await db.sales_leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead tapılmadı")
    await assert_scope_ownership(current_user, "sales", lead)
    
    updates = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    # Metadata updates (all optional)
    for k in ("contract_number", "e_invoice_date", "e_invoice_number", "voen", "payment_due_date", "follow_up", "notes", "marsol_company"):
        if k in data:
            updates[k] = data[k]
    
    # Add new payment if amount > 0
    amount = data.get("new_payment_amount")
    if amount is not None and str(amount).strip() and float(amount) > 0:
        payment = {
            "id": str(uuid.uuid4()),
            "amount": float(amount),
            "date": data.get("payment_date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            "note": data.get("payment_note", ""),
            "payment_method": data.get("payment_method", ""),
            "added_by": current_user.get("name", ""),
            "added_at": datetime.now(timezone.utc).isoformat()
        }
        history = lead.get("payment_history", []) or []
        history.append(payment)
        updates["payment_history"] = history
        updates["paid_amount"] = float(lead.get("paid_amount") or 0) + float(amount)
    
    await db.sales_leads.update_one({"id": lead_id}, {"$set": updates})
    updated = await db.sales_leads.find_one({"id": lead_id}, {"_id": 0})
    return updated

@api_router.get("/sales-leads/{lead_id}/payments")
async def get_lead_payments(lead_id: str, current_user: dict = Depends(get_current_user)):
    lead = await db.sales_leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead tapılmadı")
    return lead.get("payment_history", []) or []

# ==================== EVENT INVITATIONS (QONAQLAR / DƏVƏTLƏR) ====================

@api_router.get("/event-invitations")
async def get_event_invitations(
    event_id: Optional[str] = None,
    status: Optional[str] = None,
    invited_by: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if event_id and event_id != "all":
        query["event_id"] = event_id
    if status and status != "all":
        query["status"] = status
    if invited_by and invited_by != "all":
        query["invited_by"] = invited_by
    invitations = await db.event_invitations.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return invitations

@api_router.post("/event-invitations")
async def create_event_invitation(data: dict, current_user: dict = Depends(check_permission("sales", "write"))):
    doc = {
        "id": str(uuid.uuid4()),
        "event_id": data.get("event_id", ""),
        "event_name": data.get("event_name", ""),
        "guest_name": data.get("guest_name", ""),
        "guest_company": data.get("guest_company", ""),
        "guest_position": data.get("guest_position", ""),
        "guest_phone": data.get("guest_phone", ""),
        "guest_email": data.get("guest_email", ""),
        "status": "Dəvət edilib",
        "decline_reason": "",
        "notes": data.get("notes", ""),
        "invited_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.event_invitations.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/event-invitations/{inv_id}")
async def update_event_invitation(inv_id: str, data: dict, current_user: dict = Depends(check_permission("sales", "write"))):
    update = {k: v for k, v in data.items() if k not in ("id",)}
    await db.event_invitations.update_one({"id": inv_id}, {"$set": update})
    doc = await db.event_invitations.find_one({"id": inv_id}, {"_id": 0})
    return doc

@api_router.delete("/event-invitations/{inv_id}")
async def delete_event_invitation(inv_id: str, current_user: dict = Depends(check_permission("sales", "write"))):
    await db.event_invitations.delete_one({"id": inv_id})
    return {"message": "Dəvət silindi"}

@api_router.post("/event-invitations/{inv_id}/convert-to-lead")
async def convert_event_invitation_to_lead(inv_id: str, current_user: dict = Depends(check_permission("sales", "write"))):
    inv = await db.event_invitations.find_one({"id": inv_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Dəvət tapılmadı")
    count = await db.sales_leads.count_documents({})
    lead_code = f"SB-{str(count + 1).zfill(3)}"
    lead = {
        "id": str(uuid.uuid4()), "lead_code": lead_code,
        "company_name": inv.get("guest_company", ""),
        "contact_name": inv.get("guest_name", ""),
        "position": inv.get("guest_position", ""),
        "phone": inv.get("guest_phone", ""),
        "email": inv.get("guest_email", ""),
        "source": f"Dəvət - {inv.get('event_name', '')}",
        "sale_type": "Üzvlük", "status": "Yeni",
        "notes": inv.get("notes", ""),
        "curator": current_user.get("name", ""),
        "created_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.sales_leads.insert_one(lead)
    lead.pop("_id", None)
    await db.event_invitations.update_one({"id": inv_id}, {"$set": {"converted_to_lead": True, "lead_id": lead["id"]}})
    return lead


def _fmt_dd_mm_yyyy(iso: str) -> str:
    """Convert ISO YYYY-MM-DD to DD/MM/YYYY for display."""
    if not iso or "-" not in iso:
        return iso or ""
    try:
        parts = iso.split("T")[0].split("-")
        if len(parts) == 3:
            return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except Exception:
        pass
    return iso


@api_router.post("/event-invitations/{inv_id}/generate-card")
async def generate_invitation_card(inv_id: str, current_user: dict = Depends(check_permission("sales", "write"))):
    """Render a personalised invitation PNG and upload to Cloudinary.

    Looks up the invitation + linked event, renders the branded card with
    dynamic guest name / event title / date+time / venue, uploads the PNG
    to Cloudinary (folder marsol/invitations) and persists the URL back on
    the invitation document. Returns { url, public_id, whatsapp_link }.
    """
    inv = await db.event_invitations.find_one({"id": inv_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Dəvət tapılmadı")
    # Look up the linked event. Organization "Fəaliyyətlər" uses db.events;
    # Sales "Projects" uses db.project_events. We probe both.
    event = None
    if inv.get("event_id"):
        event = await db.events.find_one({"id": inv["event_id"]}, {"_id": 0})
        if not event:
            event = await db.project_events.find_one({"id": inv["event_id"]}, {"_id": 0})
    event = event or {}

    event_name = event.get("name") or inv.get("event_name") or "Tədbir"
    event_date_iso = event.get("date", "")
    event_date = _fmt_dd_mm_yyyy(event_date_iso)
    event_time = event.get("time", "")
    event_location = event.get("venue") or event.get("location") or ""

    # Look up message template for the event_type
    template_body = await _get_invitation_template(event.get("event_type") or "")

    try:
        png_bytes = render_invitation_png(
            guest_name=inv.get("guest_name", "") or "Qonağımız",
            event_name=event_name,
            event_date=event_date,
            event_time=event_time,
            event_location=event_location,
            body_template=template_body,
        )
    except FileNotFoundError as e:
        raise HTTPException(status_code=500, detail=f"Dəvətnamə şablonu tapılmadı: {e}")
    except Exception as e:
        logger.exception("Invitation render failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Dəvətnamə yaradıla bilmədi: {e}")

    safe_name = "".join(c if c.isalnum() else "_" for c in (inv.get("guest_name") or "guest"))[:32]
    filename = f"invitation_{safe_name}_{inv_id[:8]}.png"
    try:
        up = _cl_upload(png_bytes, filename=filename, folder="marsol/invitations", resource_type="image")
    except Exception as e:
        logger.exception("Cloudinary upload failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Cloudinary-ə yüklənmədi: {e}")

    url = up.get("url")
    public_id = up.get("public_id")

    # Compose WhatsApp pre-filled message
    phone = (inv.get("guest_phone") or "").strip()
    digits = "".join(c for c in phone if c.isdigit())
    msg_lines = [
        f"Hörmətli {inv.get('guest_name') or 'Qonağımız'},",
        f"Sizi \"{event_name}\" tədbirinə dəvət edirik.",
    ]
    if event_date or event_time:
        msg_lines.append(f"Tarix: {event_date} {event_time}".strip())
    if event_location:
        msg_lines.append(f"Ünvan: {event_location}")
    import urllib.parse as _up
    text = _up.quote("\n".join(msg_lines))
    whatsapp_link = f"https://wa.me/{digits}?text={text}" if digits else f"https://wa.me/?text={text}"

    # Persist
    await db.event_invitations.update_one(
        {"id": inv_id},
        {"$set": {
            "invitation_card_url": url,
            "invitation_card_public_id": public_id,
            "invitation_card_generated_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    return {"url": url, "public_id": public_id, "whatsapp_link": whatsapp_link, "filename": filename}


# ==================== CONTACT LISTS (SİYAHILAR) ====================

@api_router.get("/contact-lists")
async def get_contact_lists(current_user: dict = Depends(get_current_user)):
    lists = await db.contact_lists.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)
    for l in lists:
        l["contact_count"] = await db.contacts.count_documents({"list_id": l["id"]})
    return lists

@api_router.post("/contact-lists")
async def create_contact_list(data: dict, current_user: dict = Depends(check_permission("sales", "write"))):
    doc = {
        "id": str(uuid.uuid4()),
        "title": data.get("title", ""),
        "description": data.get("description", ""),
        "created_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.contact_lists.insert_one(doc)
    doc.pop("_id", None)
    doc["contact_count"] = 0
    return doc

@api_router.put("/contact-lists/{list_id}")
async def update_contact_list(list_id: str, data: dict, current_user: dict = Depends(check_permission("sales", "write"))):
    await db.contact_lists.update_one({"id": list_id}, {"$set": {k: v for k, v in data.items() if k not in ("id",)}})
    doc = await db.contact_lists.find_one({"id": list_id}, {"_id": 0})
    return doc

@api_router.delete("/contact-lists/{list_id}")
async def delete_contact_list(list_id: str, current_user: dict = Depends(check_permission("sales", "write"))):
    await db.contact_lists.delete_one({"id": list_id})
    await db.contacts.delete_many({"list_id": list_id})
    return {"message": "Siyahı silindi"}

@api_router.get("/contact-lists/{list_id}/contacts")
async def get_list_contacts(list_id: str, current_user: dict = Depends(get_current_user)):
    contacts = await db.contacts.find({"list_id": list_id}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    return contacts

@api_router.post("/contact-lists/{list_id}/contacts")
async def add_contact_to_list(list_id: str, data: dict, current_user: dict = Depends(check_permission("sales", "write"))):
    doc = {
        "id": str(uuid.uuid4()), "list_id": list_id,
        "name": data.get("name", ""), "surname": data.get("surname", ""),
        "company": data.get("company", ""), "position": data.get("position", ""),
        "phone": data.get("phone", ""), "email": data.get("email", ""),
        "notes": data.get("notes", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.contacts.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.post("/contact-lists/{list_id}/import")
async def import_contacts(list_id: str, data: dict, current_user: dict = Depends(check_permission("sales", "write"))):
    contacts = data.get("contacts", [])
    if not contacts:
        raise HTTPException(status_code=400, detail="Boş siyahı")
    docs = []
    for c in contacts:
        docs.append({
            "id": str(uuid.uuid4()), "list_id": list_id,
            "name": c.get("name", c.get("Ad", "")), "surname": c.get("surname", c.get("Soyad", "")),
            "company": c.get("company", c.get("Şirkət", "")), "position": c.get("position", c.get("Vəzifə", "")),
            "phone": c.get("phone", c.get("Telefon", "")), "email": c.get("email", c.get("Email", "")),
            "notes": c.get("notes", c.get("Qeyd", "")),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    await db.contacts.insert_many(docs)
    return {"message": f"{len(docs)} kontakt import edildi"}

@api_router.delete("/contact-lists/{list_id}/contacts/{contact_id}")
async def delete_contact(list_id: str, contact_id: str, current_user: dict = Depends(check_permission("sales", "write"))):
    await db.contacts.delete_one({"id": contact_id, "list_id": list_id})
    return {"message": "Kontakt silindi"}

@api_router.post("/contacts/{contact_id}/convert-to-lead")
async def convert_contact_to_lead(contact_id: str, current_user: dict = Depends(check_permission("sales", "write"))):
    contact = await db.contacts.find_one({"id": contact_id}, {"_id": 0})
    if not contact:
        raise HTTPException(status_code=404, detail="Kontakt tapılmadı")
    # Resolve the parent contact list name so we can surface it through every step
    src_list_name = ""
    if contact.get("list_id"):
        cl = await db.contact_lists.find_one({"id": contact["list_id"]}, {"_id": 0, "title": 1, "name": 1})
        if cl:
            src_list_name = cl.get("title") or cl.get("name", "")
    count = await db.sales_leads.count_documents({})
    lead = {
        "id": str(uuid.uuid4()), "lead_code": f"SB-{str(count+1).zfill(3)}",
        "company_name": contact.get("company", ""), "contact_name": f"{contact.get('name','')} {contact.get('surname','')}".strip(),
        "position": contact.get("position", ""), "phone": contact.get("phone", ""), "email": contact.get("email", ""),
        "source": f"Siyahı: {src_list_name}" if src_list_name else "Siyahıdan",
        "source_contact_list_id": contact.get("list_id", ""),
        "source_contact_list_name": src_list_name,
        "source_contact_id": contact_id,
        "sale_type": "", "status": "Yeni", "notes": contact.get("notes", ""),
        "curator": current_user.get("name", ""), "created_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.sales_leads.insert_one(lead)
    lead.pop("_id", None)
    return lead


# ==================== SALES LEADS (ŞİRKƏT BAZASI) ====================

LEAD_STATUSES = ["Yeni", "Əlaqə quruldu", "Görüş təyin edildi", "Təklif göndərildi", "Danışıqda", "Üzv oldu", "Satıldı", "İmtina"]

@api_router.get("/sales-leads")
async def get_sales_leads(
    status: Optional[str] = None,
    source: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status and status != "all":
        query["status"] = status
    if source and source != "all":
        query["source"] = source
    query = await apply_scope(query, current_user, "sales")
    leads = await db.sales_leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return leads

@api_router.get("/sales-leads/stats")
async def get_sales_leads_stats(current_user: dict = Depends(get_current_user)):
    query = await apply_scope({}, current_user, "sales")
    total = await db.sales_leads.count_documents(query)
    stats = {"total": total}
    for s in LEAD_STATUSES:
        q = await apply_scope({"status": s}, current_user, "sales")
        stats[s] = await db.sales_leads.count_documents(q)
    return stats

@api_router.post("/sales-leads")
async def create_sales_lead(data: dict, current_user: dict = Depends(check_permission("sales", "write"))):
    count = await db.sales_leads.count_documents({})
    lead_code = f"SB-{str(count + 1).zfill(3)}"
    doc = {
        "id": str(uuid.uuid4()),
        "lead_code": lead_code,
        "company_name": data.get("company_name", ""),
        "contact_name": data.get("contact_name", ""),
        "position": data.get("position", ""),
        "phone": data.get("phone", ""),
        "email": data.get("email", ""),
        "source": data.get("source", ""),
        "sale_type": data.get("sale_type", "Üzvlük"),
        "status": data.get("status", "Yeni"),
        "notes": data.get("notes", ""),
        "project_id": data.get("project_id", ""),
        "package": data.get("package", ""),
        "kv_m": data.get("kv_m"),
        "price_per_sqm": data.get("price_per_sqm"),
        "stand_number": data.get("stand_number", ""),
        "hall_number": data.get("hall_number", ""),
        "total_amount": data.get("total_amount"),
        "participant_count": data.get("participant_count"),
        "marsol_company": (data.get("marsol_company") or current_user.get("marsol_company") or "").strip(),
        "curator": data.get("curator") or current_user.get("name", ""),
        "created_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.sales_leads.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/sales-leads/{lead_id}")
async def update_sales_lead(lead_id: str, data: dict, current_user: dict = Depends(check_permission("sales", "write"))):
    update_data = {k: v for k, v in data.items() if k not in ("id", "lead_code")}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Check if status changed to "Üzv oldu"
    new_status = update_data.get("status")
    lead = await db.sales_leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead tapılmadı")
    await assert_scope_ownership(current_user, "sales", lead)
    
    result = await db.sales_leads.update_one({"id": lead_id}, {"$set": update_data})
    
    if new_status == "Üzv oldu" and lead.get("status") != "Üzv oldu":
        # Merge updates into lead view so newly-supplied fields (paid_amount,
        # total_amount, package, contract dates …) are reflected when we
        # spin up the matching `companies` record.
        merged = {**lead, **update_data}
        sale_type = merged.get("sale_type", "")
        if sale_type == "Üzvlük":
            # Check if company already exists
            existing = await db.companies.find_one({"brand_name": merged["company_name"]}, {"_id": 0})
            if not existing:
                new_display_id = await _next_company_display_id()
                total_amt = float(merged.get("total_amount") or 0)
                paid_amt = float(merged.get("paid_amount") or 0)
                company_doc = {
                    "id": str(uuid.uuid4()),
                    "display_id": new_display_id,
                    "brand_name": merged["company_name"],
                    "legal_name": merged["company_name"],
                    "sector": "",
                    "company_size": "",
                    "registration_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                    "address": "",
                    "bank_details": "",
                    "owner_name": merged.get("contact_name", ""),
                    "owner_phone": merged.get("phone", ""),
                    "owner_email": merged.get("email", ""),
                    "owner_social_links": "",
                    "co_founders": [],
                    "representative_name": "",
                    "representative_phone": "",
                    "representative_email": "",
                    "company_phone": merged.get("phone", ""),
                    "company_email": merged.get("email", ""),
                    "website": "",
                    "package": merged.get("package", ""),
                    "joined_project": "",  # leave empty so user picks in Şirkət bazası
                    "contract_start_date": merged.get("contract_start_date") or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                    "contract_end_date": merged.get("contract_end_date", ""),
                    "payment_amount": total_amt,
                    "total_amount": total_amt,
                    "paid_amount": paid_amt,
                    "debt_amount": max(total_amt - paid_amt, 0.0),
                    "payment_due_date": "",
                    "status": "Aktiv",
                    "sub_sector": "",
                    "marsol_representative": merged.get("curator", ""),
                    "source_lead_id": lead_id,
                    "source": merged.get("source", ""),
                    "source_contact_list_id": merged.get("source_contact_list_id", ""),
                    "source_contact_list_name": merged.get("source_contact_list_name", ""),
                    "source_contact_id": merged.get("source_contact_id", ""),
                    "curator": merged.get("curator", ""),
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.companies.insert_one(company_doc)
    
    doc = await db.sales_leads.find_one({"id": lead_id}, {"_id": 0})
    
    # Email notify on status transition to "Satıldı"/"Üzv oldu"
    if new_status in ("Satıldı", "Üzv oldu") and lead.get("status") not in ("Satıldı", "Üzv oldu"):
        curator_email = await _user_email_by_name(doc.get("curator", ""))
        amount = doc.get("total_amount") or 0
        body = f"""<p>🎉 Yeni satış bağlandı!</p>
        <table cellpadding='6' cellspacing='0' style='width:100%;border:1px solid #e2e8f0;border-radius:8px;font-size:13px'>
          <tr><td style='color:#64748b'>Şirkət</td><td style='font-weight:600'>{doc.get('company_name', '')}</td></tr>
          <tr><td style='color:#64748b'>Sahibkar</td><td>{doc.get('contact_name', '')}</td></tr>
          <tr><td style='color:#64748b'>Növ</td><td>{doc.get('sale_type', '')}</td></tr>
          <tr><td style='color:#64748b'>Status</td><td><span style='background:#dcfce7;color:#166534;padding:2px 8px;border-radius:4px'>{new_status}</span></td></tr>
          <tr><td style='color:#64748b'>Məbləğ</td><td style='font-weight:600;color:#166534'>{amount} AZN</td></tr>
          <tr><td style='color:#64748b'>Kurator</td><td>{doc.get('curator', '')}</td></tr>
        </table>"""
        await _email_notify_safe(
            title=f"Yeni satış: {doc.get('company_name', '')}",
            body_html=body,
            extra_recipients=[curator_email],
        )
    return doc
    return doc

@api_router.delete("/sales-leads/{lead_id}")
async def delete_sales_lead(lead_id: str, current_user: dict = Depends(check_permission("sales", "write"))):
    existing = await db.sales_leads.find_one({"id": lead_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Lead tapılmadı")
    await assert_scope_ownership(current_user, "sales", existing)
    result = await db.sales_leads.delete_one({"id": lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead tapılmadı")
    return {"message": "Lead silindi"}

@api_router.get("/sales-members")
async def get_sales_members(current_user: dict = Depends(get_current_user)):
    query = {"status": "Üzv oldu", "sale_type": "Üzvlük"}
    query = await apply_scope(query, current_user, "sales")
    members = await db.sales_leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return members

@api_router.get("/members")
async def get_members(
    package: Optional[str] = None,
    sector: Optional[str] = None,
    status: Optional[str] = None,
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if package and package != "all":
        query["package"] = package
    if sector and sector != "all":
        query["sector"] = sector
    if status and status != "all":
        query["status"] = status
    query = await apply_scope(query, current_user, "members")
    companies = await db.companies.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    if year is not None:
        companies = [c for c in companies if _company_covers_year(c, year)]
    # Annotate each company with the membership-period applicable to that year (or current)
    # and map the internal schema (brand_name, owner_name, …) to the field names
    # Members.jsx expects (company_name, director_name, …).
    now = datetime.now(timezone.utc)
    for c in companies:
        if year is not None:
            cs_y_cur = (c.get("contract_start_date") or "")[:4]
            ce_y_cur = (c.get("contract_end_date") or "")[:4]
            try:
                cs_yi = int(cs_y_cur) if cs_y_cur else 0
                ce_yi = int(ce_y_cur) if ce_y_cur else 9999
            except (ValueError, TypeError):
                cs_yi, ce_yi = 0, 9999
            if cs_yi <= year <= ce_yi:
                c["_period"] = {
                    "package": c.get("package", ""),
                    "contract_start": c.get("contract_start_date", ""),
                    "contract_end": c.get("contract_end_date", ""),
                    "is_current": True
                }
            else:
                for h in (c.get("membership_history") or []):
                    try:
                        hcs_y = int((h.get("contract_start") or "")[:4]) if h.get("contract_start") else 0
                        hce_y = int((h.get("contract_end") or "")[:4]) if h.get("contract_end") else 9999
                    except (ValueError, TypeError):
                        hcs_y, hce_y = 0, 9999
                    if hcs_y <= year <= hce_y:
                        c["_period"] = {**h, "is_current": False}
                        break
        # Field aliases for frontend compatibility
        c["company_name"] = c.get("brand_name", "")
        c["director_name"] = c.get("owner_name", "")
        c["director_phone"] = c.get("owner_phone", "")
        c["contact_person"] = c.get("representative_name", "")
        c["contact_position"] = c.get("representative_position", "")
        c["business_size"] = c.get("company_size", "")
        # Calculate days until expiry (negative if already expired)
        end_date_str = c.get("contract_end_date", "")
        c["days_until_expiry"] = None
        if end_date_str:
            try:
                end_dt = datetime.strptime(end_date_str, "%Y-%m-%d")
                c["days_until_expiry"] = (end_dt - now.replace(tzinfo=None)).days
            except (ValueError, TypeError):
                pass
    # Sort: expiring soonest first (None values go last)
    companies.sort(key=lambda m: (m.get("days_until_expiry") is None, m.get("days_until_expiry") if m.get("days_until_expiry") is not None else 9999))
    return companies

@api_router.get("/members/options/all")
async def get_members_options(current_user: dict = Depends(get_current_user)):
    packages_db = await db.packages.find({}, {"_id": 0}).to_list(100)
    sectors_db = await db.sectors.find({}, {"_id": 0}).to_list(100)
    projects_db = await db.projects.find({}, {"_id": 0}).to_list(100)
    users_db = await db.users.find({}, {"_id": 0, "name": 1}).to_list(500)
    return {
        "packages": [p["name"] for p in packages_db] if packages_db else ["Premium", "Business", "Business Plus"],
        "sectors": [s["name"] for s in sectors_db] if sectors_db else [],
        "statuses": ["Aktiv", "Qeyri-aktiv", "Gözləmədə"],
        "business_sizes": ["Böyük", "Orta", "Kiçik", "Mikro"],
        "curators": [u["name"] for u in users_db if u.get("name")],
        "projects": [p["name"] for p in projects_db] if projects_db else ["Üzvlük", "Sərgi", "Təlim/Proqram"],
        "contract_statuses": ["Gözləyir", "Bağlanıb", "Aktiv", "Bitib", "Ləğv edilib"],
    }

@api_router.post("/members")
async def create_member(data: dict, current_user: dict = Depends(check_permission("members", "write"))):
    new_display_id = await _next_company_display_id()
    doc = {
        "id": str(uuid.uuid4()),
        "display_id": new_display_id,
        "brand_name": data.get("company_name", ""),
        "legal_name": data.get("company_name", ""),
        "sector": data.get("sector", ""),
        "company_size": data.get("business_size", ""),
        "owner_name": data.get("director_name", ""),
        "owner_phone": data.get("director_phone", ""),
        "owner_email": data.get("email", ""),
        "representative_name": data.get("contact_person", ""),
        "company_phone": data.get("phone", ""),
        "package": data.get("package", ""),
        "joined_project": data.get("project", "Üzvlük"),
        "status": data.get("status", "Aktiv"),
        "curator": current_user.get("name", ""),
        "marsol_company": (current_user.get("marsol_company") or "").strip(),
        "registration_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.companies.insert_one(doc)
    doc.pop("_id", None)
    doc["company_name"] = doc["brand_name"]
    doc["director_name"] = doc["owner_name"]
    doc["business_size"] = doc["company_size"]
    return doc

@api_router.put("/members/{member_id}")
async def update_member(member_id: str, data: dict, current_user: dict = Depends(check_permission("members", "write"))):
    existing = await db.companies.find_one({"id": member_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Üzv tapılmadı")
    await assert_scope_ownership(current_user, "members", existing)
    update = {}
    field_map = {"company_name": "brand_name", "director_name": "owner_name", "director_phone": "owner_phone", "business_size": "company_size", "contact_person": "representative_name", "contact_position": "representative_position"}
    for k, v in data.items():
        if k in field_map:
            update[field_map[k]] = v
        elif k not in ("id",):
            update[k] = v
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.companies.update_one({"id": member_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Üzv tapılmadı")
    doc = await db.companies.find_one({"id": member_id}, {"_id": 0})
    doc["company_name"] = doc.get("brand_name", "")
    doc["director_name"] = doc.get("owner_name", "")
    doc["business_size"] = doc.get("company_size", "")
    return doc

@api_router.delete("/members/{member_id}")
async def delete_member(member_id: str, current_user: dict = Depends(check_permission("members", "write"))):
    existing = await db.companies.find_one({"id": member_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Üzv tapılmadı")
    await assert_scope_ownership(current_user, "members", existing)
    result = await db.companies.delete_one({"id": member_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Üzv tapılmadı")
    return {"message": "Üzv silindi"}

@api_router.post("/members/{member_id}/renew")
async def renew_member(member_id: str, data: dict, current_user: dict = Depends(check_permission("members", "write"))):
    """Archive current membership period to history, then start a new period.
    Body: {package, contract_start, contract_end, carry_over_quota?: bool}.
    """
    company = await db.companies.find_one({"id": member_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Üzv tapılmadı")
    await assert_scope_ownership(current_user, "members", company)

    # Snapshot current period's used_quota (counts invitations during current contract period)
    quotas = await get_package_quotas()
    cur_total_quota = quotas.get(company.get("package", ""), 0)
    cur_used_quota = await db.invitations.count_documents({
        "company_id": member_id,
        "obligation_deducted": True,
        "event_date": {
            "$gte": company.get("contract_start_date") or "0000-00-00",
            "$lte": company.get("contract_end_date") or "9999-99-99"
        }
    })

    history_entry = {
        "id": str(uuid.uuid4()),
        "package": company.get("package", ""),
        "contract_start": company.get("contract_start_date", ""),
        "contract_end": company.get("contract_end_date", ""),
        "total_quota": cur_total_quota,
        "used_quota": cur_used_quota,
        "remaining_quota": max(cur_total_quota - cur_used_quota, 0),
        "status": "Yenilənib",
        "archived_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Build update
    history = company.get("membership_history") or []
    history.append(history_entry)
    
    new_package = data.get("package", company.get("package", ""))
    update = {
        "membership_history": history,
        "package": new_package,
        "contract_start_date": data.get("contract_start", ""),
        "contract_end_date": data.get("contract_end", ""),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }

    # Carry-over (track as bonus quota)
    carry_over = bool(data.get("carry_over_quota"))
    if carry_over:
        update["bonus_quota"] = (company.get("bonus_quota") or 0) + history_entry["remaining_quota"]
    
    await db.companies.update_one({"id": member_id}, {"$set": update})
    updated = await db.companies.find_one({"id": member_id}, {"_id": 0})
    return updated

@api_router.post("/sales-leads/{lead_id}/create-meeting")
async def create_meeting_from_lead(lead_id: str, data: dict, current_user: dict = Depends(check_permission("sales", "write"))):
    lead = await db.sales_leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead tapılmadı")
    meeting_id = str(uuid.uuid4())
    meeting_doc = {
        "id": meeting_id,
        "employee": current_user.get("name", ""),
        "meeting_setter": data.get("meeting_setter", current_user.get("name", "")),
        "date": data.get("date", ""),
        "time": data.get("time", ""),
        "company": lead["company_name"],
        "contact_person": lead["contact_name"],
        "project": "",
        "meeting_type": data.get("meeting_type", "Müştəri görüşü"),
        "meeting_mode": data.get("meeting_mode", "Offline"),
        "department": "Satış",
        "location": data.get("location", ""),
        "result": "",
        "next_meeting": "",
        "notes": data.get("notes", ""),
        "reminders": [],
        "source_lead_id": lead_id,
        "created_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.meetings.insert_one(meeting_doc)
    meeting_doc.pop("_id", None)
    # Update lead status
    await db.sales_leads.update_one({"id": lead_id}, {"$set": {"status": "Görüş təyin edildi", "updated_at": datetime.now(timezone.utc).isoformat()}})
    return meeting_doc

# ==================== ASSEMBLIES (İCLAS) ====================

async def _sync_assembly_tasks(assembly_doc):
    """Sync assembly agenda tasks + general tasks to the tasks collection"""
    assembly_uuid = assembly_doc["id"]
    assembly_code = assembly_doc["assembly_code"]
    department = assembly_doc.get("department", "")
    deadline = assembly_doc.get("deadline", "")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    # Remove old tasks from this assembly
    await db.tasks.delete_many({"source": "assembly", "assembly_id": assembly_uuid})

    async def _create_task(task, related_label):
        task_title = task.get("title", "")
        responsibles = task.get("responsible_persons", [])
        assignees = task.get("assignees", [])
        # Backward compat for single values
        if not responsibles and task.get("responsible_person"):
            responsibles = [task["responsible_person"]]
        if not assignees and task.get("assignee"):
            assignees = [task["assignee"]]
        task_deadline = task.get("deadline", "") or deadline
        if task_title and (responsibles or assignees):
            count = await db.tasks.count_documents({})
            task_code = f"T-{str(count + 1).zfill(3)}"
            task_doc = {
                "id": str(uuid.uuid4()),
                "task_code": task_code,
                "task_name": f"[{assembly_code}] {task_title}",
                "department": department,
                "assignee": ", ".join(assignees) if assignees else ", ".join(responsibles),
                "responsible_person": ", ".join(responsibles),
                "priority": "Orta",
                "start_date": today,
                "end_date": task_deadline,
                "related_object_type": "İclas",
                "related_object_id": assembly_code,
                "related_object": related_label,
                "phase": "",
                "status": "Gözləyir",
                "notes": f"İclas: {assembly_code}",
                "source": "assembly",
                "assembly_id": assembly_uuid,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.tasks.insert_one(task_doc)

    # Agenda tasks
    for agenda in assembly_doc.get("agendas", []):
        agenda_title = agenda.get("title", "")
        for task in agenda.get("tasks", []):
            await _create_task(task, f"{assembly_code} - {agenda_title}")
    # General tasks (not linked to any agenda)
    for task in assembly_doc.get("general_tasks", []):
        await _create_task(task, f"{assembly_code} - Ümumi")

@api_router.get("/assemblies")
async def get_assemblies(
    department: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if department and department != "all":
        query["department"] = department
    if date_from or date_to:
        date_q = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to
        query["created_at"] = date_q
    query = await apply_scope(query, current_user, "assembly")
    assemblies = await db.assemblies.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return assemblies

@api_router.post("/assemblies")
async def create_assembly(data: dict, current_user: dict = Depends(check_permission("assembly", "write"))):
    count = await db.assemblies.count_documents({})
    assembly_code = f"IC-{str(count + 1).zfill(3)}"
    doc = {
        "id": str(uuid.uuid4()),
        "assembly_code": assembly_code,
        "department": data.get("department", ""),
        "purpose": data.get("purpose", ""),
        "agendas": data.get("agendas", []),
        "general_tasks": data.get("general_tasks", []),
        "discussion_topics": data.get("discussion_topics", []),
        "deadline": data.get("deadline", ""),
        "next_assembly_date": data.get("next_assembly_date", ""),
        "decisions": data.get("decisions", []),
        "created_by": current_user.get("name", ""),
        "marsol_company": (current_user.get("marsol_company") or "").strip(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.assemblies.insert_one(doc)
    doc.pop("_id", None)
    await _sync_assembly_tasks(doc)
    return doc

@api_router.put("/assemblies/{assembly_id}")
async def update_assembly(assembly_id: str, data: dict, current_user: dict = Depends(check_permission("assembly", "write"))):
    existing = await db.assemblies.find_one({"id": assembly_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="İclas tapılmadı")
    await assert_scope_ownership(current_user, "assembly", existing)
    update_data = {k: v for k, v in data.items() if k not in ("id", "assembly_code")}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.assemblies.update_one({"id": assembly_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="İclas tapılmadı")
    doc = await db.assemblies.find_one({"id": assembly_id}, {"_id": 0})
    await _sync_assembly_tasks(doc)
    return doc

@api_router.delete("/assemblies/{assembly_id}")
async def delete_assembly(assembly_id: str, current_user: dict = Depends(check_permission("assembly", "write"))):
    existing = await db.assemblies.find_one({"id": assembly_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="İclas tapılmadı")
    await assert_scope_ownership(current_user, "assembly", existing)
    result = await db.assemblies.delete_one({"id": assembly_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="İclas tapılmadı")
    await db.tasks.delete_many({"source": "assembly", "assembly_id": assembly_id})
    return {"message": "İclas silindi"}


# ==================== OPTIONS ====================

# ==================== MEMBERSHIP FORUM ====================

COMPANY_FORM_FIELDS = [
    # Şirkət məlumatları
    {"key": "legal_name", "label": "Hüquqi ad", "type": "text"},
    {"key": "voen", "label": "VÖEN", "type": "text"},
    {"key": "sector", "label": "Sektor", "type": "select"},
    {"key": "sub_sector", "label": "Alt sektor", "type": "select"},
    {"key": "company_size", "label": "Şirkət ölçüsü", "type": "select"},
    {"key": "employee_count", "label": "İşçi sayı", "type": "text"},
    {"key": "region", "label": "Region", "type": "text"},
    {"key": "registration_date", "label": "Şirkətin yaranma tarixi", "type": "date"},
    {"key": "address", "label": "Ünvan", "type": "textarea"},
    {"key": "company_phone", "label": "Şirkət telefonu", "type": "text"},
    {"key": "company_website", "label": "Veb sayt", "type": "text"},
    {"key": "social_links", "label": "Şirkət sosial media hesabları", "type": "array"},
    {"key": "bank_files", "label": "Bank rekvizitləri (fayl)", "type": "file"},
    {"key": "logo_url", "label": "Şirkət logosu", "type": "file"},
    # Referans
    {"key": "reference_source", "label": "Referans mənbəsi", "type": "text"},
    {"key": "reference_person_name", "label": "Referans şəxsin adı", "type": "text"},
    {"key": "reference_person_surname", "label": "Referans şəxsin soyadı", "type": "text"},
    {"key": "reference_person_position", "label": "Referans şəxsin vəzifəsi", "type": "text"},
    {"key": "reference_note", "label": "Referans qeydi", "type": "textarea"},
    # Sahibkar
    {"key": "owners", "label": "Sahibkarlar (ad, soyad, ata adı, vəzifə, telefon, email, doğum tarixi, vətəndaşlıq, təhsil, ixtisas, universitet, sosial media, uşaqlar, fəaliyyət sahələri)", "type": "owners"},
    # Əlaqədar şəxs
    {"key": "contact_first_name", "label": "Əlaqədar şəxs adı", "type": "text"},
    {"key": "contact_last_name", "label": "Əlaqədar şəxs soyadı", "type": "text"},
    {"key": "contact_position", "label": "Əlaqədar şəxs vəzifəsi", "type": "text"},
    {"key": "contact_phone", "label": "Əlaqədar şəxs telefonu", "type": "text"},
    {"key": "contact_email", "label": "Əlaqədar şəxs emaili", "type": "text"},
]

@api_router.get("/forum/fields")
async def get_forum_fields(current_user: dict = Depends(get_current_user)):
    """Get available form fields, which are enabled, and which are required."""
    settings = await db.setting_lists.find_one({"key": "forum_enabled_fields"}, {"_id": 0})
    enabled = settings.get("values", []) if settings else [f["key"] for f in COMPANY_FORM_FIELDS]
    req_doc = await db.setting_lists.find_one({"key": "forum_required_fields"}, {"_id": 0})
    required = req_doc.get("values", []) if req_doc else []
    # Also get custom fields for companies
    custom_fields = await db.custom_fields.find({"module": "companies"}, {"_id": 0}).to_list(100)
    all_fields = COMPANY_FORM_FIELDS.copy()
    for cf in custom_fields:
        all_fields.append({"key": f"custom_{cf['id']}", "label": cf.get("label", cf.get("name", "")), "custom": True})
    return {"fields": all_fields, "enabled": enabled, "required": required}

@api_router.put("/forum/fields")
async def update_forum_fields(data: dict, current_user: dict = Depends(check_permission("settings", "write"))):
    enabled = data.get("enabled", [])
    required = data.get("required", [])
    # Required must be a subset of enabled
    required = [r for r in required if r in enabled]
    await db.setting_lists.update_one({"key": "forum_enabled_fields"}, {"$set": {"values": enabled}}, upsert=True)
    await db.setting_lists.update_one({"key": "forum_required_fields"}, {"$set": {"values": required}}, upsert=True)
    return {"message": "Forum sahələri yeniləndi", "enabled": enabled, "required": required}


async def _get_all_options():
    """Helper: get dynamic options for forms - mirrors settings endpoints with fallbacks"""
    regions_db = await db.regions.find({}, {"_id": 0}).to_list(500)
    if not regions_db:
        regions_db = [{"name": n} for n in ["Bakı", "Sumqayıt", "Gəncə", "Lənkəran", "Mingəçevir", "Şəki", "Şirvan", "Naxçıvan", "Abşeron", "Digər"]]
    positions_db = await db.positions.find({}, {"_id": 0}).to_list(500)
    if not positions_db:
        positions_db = [{"name": n} for n in ["Direktor", "Təsisçi", "Baş direktor", "İcraçı direktor", "Kommersiya direktoru", "Maliyyə direktoru"]]
    activities_db = await db.activities.find({}, {"_id": 0}).to_list(500)
    if not activities_db:
        activities_db = [{"name": n} for n in ["Networking", "Təlim", "Sərgi", "Forum", "Mentorluq", "İş birliyi"]]
    company_sizes = await _get_setting_list("company_sizes", ["Böyük", "Orta", "Kiçik", "Mikro"])
    education_levels = await _get_setting_list("education_levels", ["Orta təhsil", "Sub bakalavr", "Bakalavr", "Magistratura", "Doktorantura"])
    return {
        "company_sizes": company_sizes,
        "regions": [r["name"] for r in regions_db],
        "positions": [p["name"] for p in positions_db],
        "education_levels": education_levels,
        "activities": [a["name"] for a in activities_db],
    }

@api_router.post("/forum/generate-link/{company_id}")
async def generate_forum_link(company_id: str, current_user: dict = Depends(get_current_user)):
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Şirkət tapılmadı")
    token = str(uuid.uuid4())[:12]
    await db.forum_tokens.update_one(
        {"company_id": company_id},
        {"$set": {"token": token, "company_id": company_id, "created_at": datetime.now(timezone.utc).isoformat(), "created_by": current_user.get("name", "")}},
        upsert=True
    )
    return {"token": token, "company_id": company_id}

# PUBLIC endpoints - no auth required
@api_router.get("/public/form/{token}")
async def get_public_form(token: str):
    form_token = await db.forum_tokens.find_one({"token": token}, {"_id": 0})
    if not form_token:
        raise HTTPException(status_code=404, detail="Form tapılmadı və ya vaxtı keçib")
    company = await db.companies.find_one({"id": form_token["company_id"]}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Şirkət tapılmadı")
    settings = await db.setting_lists.find_one({"key": "forum_enabled_fields"}, {"_id": 0})
    enabled = settings.get("values", []) if settings else [f["key"] for f in COMPANY_FORM_FIELDS]
    req_doc = await db.setting_lists.find_one({"key": "forum_required_fields"}, {"_id": 0})
    required = req_doc.get("values", []) if req_doc else []
    fields_info = {f["key"]: {"label": f["label"], "type": f.get("type", "text")} for f in COMPANY_FORM_FIELDS}
    custom_fields = await db.custom_fields.find({"module": "companies"}, {"_id": 0}).to_list(100)
    for cf in custom_fields:
        fields_info[f"custom_{cf['id']}"] = {"label": cf.get("label", cf.get("name", "")), "type": "text"}
    enabled_fields = [{"key": k, "label": fields_info[k]["label"], "type": fields_info[k]["type"], "required": k in required} for k in enabled if k in fields_info]
    # Get dynamic options for selects
    sectors_db = await db.sectors.find({}, {"_id": 0}).to_list(200)
    sub_sectors_db = await db.sub_sectors.find({}, {"_id": 0}).to_list(500)
    sub_map = {}
    for s in sub_sectors_db:
        sub_map.setdefault(s.get("sector", ""), []).append(s.get("name", ""))
    form_description = await db.setting_lists.find_one({"key": "forum_description"}, {"_id": 0})
    all_options = await _get_all_options()
    return {
        "company_name": company.get("brand_name", ""),
        "owner_phone": company.get("owner_phone", ""),
        "owner_name": company.get("owner_name", ""),
        "fields": enabled_fields,
        "current_values": {k: company.get(k, "") for k in enabled},
        "description": (form_description.get("values", [""])[0] if form_description else "Zəhmət olmasa şirkət məlumatlarını doldurun"),
        "options": {
            "sectors": [s.get("name", "") for s in sectors_db] + ["Digər"],
            "sub_sectors": sub_map,
            "company_sizes": all_options.get("company_sizes", ["Böyük", "Orta", "Kiçik", "Mikro"]),
            "regions": all_options.get("regions", []),
            "positions": all_options.get("positions", []),
            "education_levels": all_options.get("education_levels", []),
            "activities": all_options.get("activities", []),
        }
    }

@api_router.post("/public/form/{token}")
async def submit_public_form(token: str, data: dict):
    form_token = await db.forum_tokens.find_one({"token": token}, {"_id": 0})
    if not form_token:
        raise HTTPException(status_code=404, detail="Form tapılmadı")
    company_id = form_token["company_id"]
    settings = await db.setting_lists.find_one({"key": "forum_enabled_fields"}, {"_id": 0})
    enabled = settings.get("values", []) if settings else [f["key"] for f in COMPANY_FORM_FIELDS]
    req_doc = await db.setting_lists.find_one({"key": "forum_required_fields"}, {"_id": 0})
    required = [r for r in (req_doc.get("values", []) if req_doc else []) if r in enabled]
    field_labels = {f["key"]: f["label"] for f in COMPANY_FORM_FIELDS}
    # Validate required fields are filled
    missing = []
    for rk in required:
        val = data.get(rk, "")
        # consider empty if string is blank or list is empty
        is_blank = (val is None) or (isinstance(val, str) and not val.strip()) or (isinstance(val, list) and len(val) == 0)
        if is_blank:
            missing.append(field_labels.get(rk, rk))
    if missing:
        raise HTTPException(status_code=400, detail=f"Aşağıdakı məcburi sahələr doldurulmalıdır: {', '.join(missing)}")
    pending_data = {}
    for key in enabled:
        if key in data:
            pending_data[key] = data[key]
    if "owners" in data and "owners" in enabled:
        pending_data["owners"] = data["owners"]
    if pending_data:
        # Store as PENDING — admin must approve
        await db.companies.update_one({"id": company_id}, {"$set": {
            "pending_form_data": pending_data,
            "pending_form_submitted_at": datetime.now(timezone.utc).isoformat(),
            "pending_form_status": "Gözləyir",
        }})
        # Email notify admin + curator
        company = await db.companies.find_one({"id": company_id}, {"_id": 0})
        curator_email = await _user_email_by_name((company or {}).get("curator", ""))
        rows = "".join(f"<tr><td style='padding:4px 8px;color:#64748b'>{k}</td><td style='padding:4px 8px;color:#0f172a;font-weight:600'>{v}</td></tr>" for k, v in pending_data.items())
        body = f"""<p><strong>{(company or {}).get('brand_name', '')}</strong> şirkəti üzvlük formunu doldurub. Admin təsdiqi gözləyir.</p>
        <table cellpadding='0' cellspacing='0' style='width:100%;border:1px solid #e2e8f0;border-radius:8px;margin-top:8px;font-size:13px'>{rows}</table>
        <p style='margin-top:12px;color:#64748b;font-size:12px'>Sistemə daxil olub Şirkət Məlumatları → həmin şirkətə klikləyin və Təsdiqlə/Rədd et seçin.</p>"""
        await _email_notify_safe(
            title="Forum dəyişikliyi təsdiq gözləyir",
            body_html=body,
            extra_recipients=[curator_email],
        )
    return {"message": "Məlumatlar uğurla göndərildi. Admin təsdiqindən sonra məlumatlar yenilənəcək."}


@api_router.post("/companies/{company_id}/approve-form")
async def approve_form_submission(company_id: str, current_user: dict = Depends(check_permission("companies", "write"))):
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Şirkət tapılmadı")
    await assert_scope_ownership(current_user, "companies", company)
    pending = company.get("pending_form_data") or {}
    if not pending:
        raise HTTPException(status_code=400, detail="Təsdiq olunacaq məlumat yoxdur")
    update = {**pending, "form_submitted_at": company.get("pending_form_submitted_at") or datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()}
    await db.companies.update_one({"id": company_id}, {
        "$set": update,
        "$unset": {"pending_form_data": "", "pending_form_status": "", "pending_form_submitted_at": ""}
    })
    return {"message": "Təsdiqləndi və məlumatlar yeniləndi"}


@api_router.post("/companies/{company_id}/reject-form")
async def reject_form_submission(company_id: str, data: Optional[dict] = None, current_user: dict = Depends(check_permission("companies", "write"))):
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Şirkət tapılmadı")
    await assert_scope_ownership(current_user, "companies", company)
    if not company.get("pending_form_data"):
        raise HTTPException(status_code=400, detail="Təsdiq olunacaq məlumat yoxdur")
    reject_reason = (data or {}).get("reason", "")
    await db.companies.update_one({"id": company_id}, {
        "$unset": {"pending_form_data": "", "pending_form_submitted_at": ""},
        "$set": {"pending_form_status": "Rədd edildi", "pending_form_reject_reason": reject_reason, "updated_at": datetime.now(timezone.utc).isoformat()}
    })
    return {"message": "Forma rədd edildi"}



# ==================== ROLES ====================

@api_router.get("/roles")
async def get_roles(current_user: dict = Depends(get_current_user)):
    roles = await db.roles.find({}, {"_id": 0}).to_list(100)
    return roles

@api_router.post("/roles")
async def create_role(data: dict, current_user: dict = Depends(check_permission("settings", "write"))):
    existing = await db.roles.find_one({"name": data.get("name", "")})
    if existing:
        raise HTTPException(status_code=400, detail="Bu adda rol artıq mövcuddur")
    doc = {
        "id": str(uuid.uuid4()),
        "name": data.get("name", ""),
        "permissions": data.get("permissions", {}),
        "scopes": data.get("scopes", {}),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.roles.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/roles/{role_id}")
async def update_role(role_id: str, data: dict, current_user: dict = Depends(check_permission("settings", "write"))):
    update_data = {k: v for k, v in data.items() if k not in ("id",)}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.roles.update_one({"id": role_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Rol tapılmadı")
    doc = await db.roles.find_one({"id": role_id}, {"_id": 0})
    return doc

@api_router.delete("/roles/{role_id}")
async def delete_role(role_id: str, current_user: dict = Depends(check_permission("settings", "write"))):
    role = await db.roles.find_one({"id": role_id}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Rol tapılmadı")
    # Don't allow deleting role if users are assigned to it
    users_with_role = await db.users.count_documents({"role": role["name"]})
    if users_with_role > 0:
        raise HTTPException(status_code=400, detail=f"Bu rol {users_with_role} istifadəçiyə təyin edilib, əvvəl onların rolunu dəyişin")
    await db.roles.delete_one({"id": role_id})
    return {"message": "Rol silindi"}

@api_router.get("/my-permissions")
async def get_my_permissions(current_user: dict = Depends(get_current_user)):
    perms = await get_user_permissions(current_user)
    return {"role": current_user.get("role", ""), "permissions": perms}



async def _get_setting_list(key: str, defaults: list) -> list:
    doc = await db.setting_lists.find_one({"key": key}, {"_id": 0})
    if doc and doc.get("values"):
        return doc["values"]
    return defaults

@api_router.get("/settings/lists/{key}")
async def get_setting_list(key: str, current_user: dict = Depends(get_current_user)):
    doc = await db.setting_lists.find_one({"key": key}, {"_id": 0})
    if doc:
        return doc.get("values", [])
    return []

@api_router.put("/settings/lists/{key}")
async def update_setting_list(key: str, data: dict, current_user: dict = Depends(check_permission("settings", "write"))):
    values = data.get("values", [])
    await db.setting_lists.update_one({"key": key}, {"$set": {"key": key, "values": values, "updated_at": datetime.now(timezone.utc).isoformat()}}, upsert=True)
    return {"key": key, "values": values}

# ==== Manageable Lists Registry ====
MANAGEABLE_LISTS = [
    {"key": "company_sizes", "label": "Şirkət ölçüləri", "defaults": ["Böyük", "Orta", "Kiçik", "Mikro"], "group": "Şirkət"},
    {"key": "organization_forms", "label": "Təşkilat formaları", "defaults": ["MMC", "ASC", "QSC", "Fərdi sahibkar", "Fond", "İB", "Digər"], "group": "Şirkət"},
    {"key": "company_statuses", "label": "Şirkət statusları", "defaults": ["Aktiv", "Qeyri-aktiv", "Gözləmədə"], "group": "Şirkət"},
    {"key": "contract_statuses", "label": "Müqavilə statusları", "defaults": ["Aktiv", "Yeni", "Yeniləmə gözlənir", "Bitdi", "Ləğv edilib"], "group": "Şirkət"},
    {"key": "departments", "label": "Şöbələr", "defaults": ["Satış", "Marketing", "HR", "Maliyyə", "Layihə", "İT", "İdarəetmə"], "group": "HR"},
    {"key": "education_levels", "label": "Təhsil səviyyələri", "defaults": ["Orta təhsil", "Sub bakalavr", "Bakalavr", "Magistratura", "Doktorantura"], "group": "HR"},
    {"key": "citizenships", "label": "Vətəndaşlıqlar", "defaults": ["Azərbaycan", "Türkiyə", "Rusiya", "Gürcüstan", "Ukrayna", "Digər"], "group": "HR"},
    {"key": "employee_statuses", "label": "Əməkdaş statusları", "defaults": ["Aktiv", "Məzuniyyətdə", "Xəstələnib", "İşdən çıxıb"], "group": "HR"},
    {"key": "marital_statuses", "label": "Ailə vəziyyəti", "defaults": ["Subay", "Evli", "Boşanmış", "Dul"], "group": "HR"},
    {"key": "task_statuses", "label": "Tapşırıq statusları", "defaults": ["Gözləyir", "İcrada", "Tamamlandı", "Ləğv edildi"], "group": "Tapşırıqlar"},
    {"key": "priorities", "label": "Prioritetlər", "defaults": ["Yüksək", "Orta", "Aşağı"], "group": "Tapşırıqlar"},
    {"key": "lead_statuses", "label": "Lead statusları", "defaults": ["Yeni", "Əlaqə quruldu", "Görüş təyin edildi", "Təklif göndərildi", "Danışıqda", "Üzv oldu", "Satıldı", "İmtina"], "group": "Satış"},
    {"key": "reference_sources", "label": "Referans mənbələri", "defaults": ["Şirkət", "Şəxs", "Media", "Digər"], "group": "Satış"},
    {"key": "payment_methods", "label": "Ödəniş üsulları", "defaults": ["Nağd", "Köçürmə", "Kart", "Hissə-hissə"], "group": "Maliyyə"},
    {"key": "expense_types", "label": "Xərc növləri (ümumi)", "defaults": ["Əməliyyat", "Marketinq", "Layihə", "Texniki", "Satış", "Digər"], "group": "Maliyyə"},
    {"key": "event_types", "label": "Tədbir növləri", "defaults": ["Konfrans", "Seminar", "Təlim", "Sərgi", "Networking", "İclas"], "group": "Tədbirlər"},
    {"key": "invitation_response_statuses", "label": "Dəvət cavabları", "defaults": ["Gözləmədə", "Qatıldı", "Rədd etdi", "Cavab vermədi"], "group": "Tədbirlər"},
    {"key": "cities", "label": "Şəhərlər", "defaults": ["Bakı", "Sumqayıt", "Gəncə", "Mingəçevir", "Şirvan", "Naxçıvan", "Lənkəran", "Şəki", "Quba", "Qəbələ", "Şamaxı", "Xaçmaz", "İsmayıllı", "Yevlax", "Ağdam"], "group": "Təşkilatçılıq"},
    {"key": "layout_types", "label": "Düzülüş növləri", "defaults": ["Banket", "Teatr", "U-forma", "Klass", "Boardroom", "Kokteyl", "Yarımdairə", "Konfrans"], "group": "Təşkilatçılıq"},
]
MANAGEABLE_LIST_KEYS = [item["key"] for item in MANAGEABLE_LISTS]


@api_router.get("/settings/manageable-lists")
async def get_manageable_lists(current_user: dict = Depends(get_current_user)):
    """Return registry + current values for every managed dropdown list."""
    docs = await db.setting_lists.find({"key": {"$in": MANAGEABLE_LIST_KEYS}}, {"_id": 0}).to_list(500)
    by_key = {d["key"]: d.get("values") or [] for d in docs}
    out = []
    for item in MANAGEABLE_LISTS:
        values = by_key.get(item["key"])
        if not values:
            values = list(item["defaults"])
        out.append({**item, "values": values})
    return out

# ==== Global Notification Days Settings ====
NOTIFICATION_DEFAULTS = {
    "membership_warning_days": 10,   # üzvlük bitməsinə neçə gün qalmış xəbərdarlıq
    "contract_expiry_days": 30,      # müqavilə bitməsi xəbərdarlığı
    "birthday_advance_days": 1,      # ad günündən əvvəl xəbərdarlıq (gün)
    "debt_overdue_high_days": 30,    # neçə gündən sonra borc HIGH severity olur
    "meeting_reminder_high_days": 1, # görüş xatırlatması — neçə gün qalmış HIGH
    "meeting_reminder_medium_days": 3,
}

async def get_notification_settings():
    doc = await db.app_config.find_one({"key": "notification_settings"}, {"_id": 0})
    cfg = dict(NOTIFICATION_DEFAULTS)
    if doc and isinstance(doc.get("values"), dict):
        for k, v in doc["values"].items():
            try:
                cfg[k] = int(v)
            except (ValueError, TypeError):
                pass
    # Backward-compat: legacy `setting_lists.membership_warning_days`
    legacy = await db.setting_lists.find_one({"key": "membership_warning_days"}, {"_id": 0})
    if legacy and legacy.get("values"):
        try:
            cfg["membership_warning_days"] = int(legacy["values"][0])
        except (ValueError, IndexError, TypeError):
            pass
    return cfg

@api_router.get("/settings/notification-config")
async def get_notification_config(current_user: dict = Depends(get_current_user)):
    return await get_notification_settings()

# ==================== TASK AUTO-ARCHIVE SETTINGS ====================

TASK_ARCHIVE_DEFAULT_DAYS = 30

async def _get_task_archive_days() -> int:
    doc = await db.app_config.find_one({"key": "task_archive_config"}, {"_id": 0})
    if not doc:
        return TASK_ARCHIVE_DEFAULT_DAYS
    val = ((doc.get("values") or {}).get("auto_archive_days"))
    try:
        n = int(val)
        return max(0, n)
    except (TypeError, ValueError):
        return TASK_ARCHIVE_DEFAULT_DAYS


@api_router.get("/settings/task-archive-config")
async def get_task_archive_config(current_user: dict = Depends(get_current_user)):
    return {"auto_archive_days": await _get_task_archive_days()}


@api_router.put("/settings/task-archive-config")
async def update_task_archive_config(data: dict, current_user: dict = Depends(check_permission("settings", "write"))):
    try:
        days = max(0, int(data.get("auto_archive_days", TASK_ARCHIVE_DEFAULT_DAYS)))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="auto_archive_days tam ədəd olmalıdır")
    await db.app_config.update_one(
        {"key": "task_archive_config"},
        {"$set": {"key": "task_archive_config", "values": {"auto_archive_days": days},
                  "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    return {"auto_archive_days": days}


async def _auto_archive_completed_tasks() -> dict:
    """Archive tasks where status=Tamamlandı AND the most recent of
    (`completed_at` if present, otherwise `updated_at`, otherwise `end_date`,
    otherwise `created_at`) is older than `auto_archive_days` days ago.
    Returns {"archived": int}.
    """
    days = await _get_task_archive_days()
    if days <= 0:
        return {"archived": 0, "skipped_reason": "auto_archive disabled (days=0)"}
    cutoff_dt = datetime.now(timezone.utc) - timedelta(days=days)
    cutoff_iso = cutoff_dt.isoformat()
    now_iso = datetime.now(timezone.utc).isoformat()
    candidates = await db.tasks.find({"status": "Tamamlandı"}, {"_id": 0}).to_list(20000)
    to_archive = []
    for t in candidates:
        marker = (t.get("completed_at") or t.get("updated_at") or
                  (t.get("end_date") or "") or t.get("created_at") or "")
        # end_date is just YYYY-MM-DD (no time) — pad to full ISO for compare
        if len(marker) == 10:
            marker = marker + "T23:59:59+00:00"
        if marker and marker < cutoff_iso:
            to_archive.append(t)
    if not to_archive:
        return {"archived": 0}
    archive_docs = []
    for d in to_archive:
        archive_docs.append({**d, "archive_id": str(uuid.uuid4()),
                             "archived_at": now_iso, "archived_by": "system (auto)"})
    if archive_docs:
        await db.tasks_archive.insert_many(archive_docs)
        ids = [d["id"] for d in to_archive]
        await db.tasks.delete_many({"id": {"$in": ids}})
    return {"archived": len(to_archive)}


@api_router.post("/tasks/auto-archive")
async def trigger_auto_archive(current_user: dict = Depends(check_permission("tasks", "write"))):
    """Manual trigger for the auto-archive job (admin / write-capable user).
    Returns the number of tasks moved into the archive."""
    return await _auto_archive_completed_tasks()


@api_router.put("/settings/notification-config")
async def update_notification_config(data: dict, current_user: dict = Depends(check_permission("settings", "write"))):
    clean = {}
    for k in NOTIFICATION_DEFAULTS:
        if k in data:
            try:
                clean[k] = int(data[k])
            except (ValueError, TypeError):
                clean[k] = NOTIFICATION_DEFAULTS[k]
    await db.app_config.update_one(
        {"key": "notification_settings"},
        {"$set": {"key": "notification_settings", "values": clean, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    # Mirror membership_warning_days into legacy setting_lists for backward compat
    if "membership_warning_days" in clean:
        await db.setting_lists.update_one(
            {"key": "membership_warning_days"},
            {"$set": {"key": "membership_warning_days", "values": [clean["membership_warning_days"]]}},
            upsert=True,
        )
    return await get_notification_settings()

@api_router.get("/options/all")
async def get_all_options(current_user: dict = Depends(get_current_user)):
    # Get dynamic data from database
    packages_db = await db.packages.find({}, {"_id": 0}).to_list(100)
    projects_db = await db.projects.find({}, {"_id": 0}).to_list(100)
    sectors_db = await db.sectors.find({}, {"_id": 0}).to_list(100)
    users_db = await db.users.find({}, {"_id": 0, "id": 1, "name": 1, "role": 1, "department": 1}).to_list(500)
    sub_sectors_db = await db.sub_sectors.find({}, {"_id": 0}).to_list(500)
    positions_db = await db.positions.find({}, {"_id": 0}).to_list(200)
    activities_db = await db.activities.find({}, {"_id": 0}).to_list(200)
    regions_db = await db.regions.find({}, {"_id": 0}).to_list(200)
    marsol_companies_db = await db.marsol_companies.find({}, {"_id": 0}).to_list(100)
    
    # Use database values if exist, otherwise use defaults
    packages = [{"name": p["name"], "price": p.get("price", 0)} for p in packages_db] if packages_db else [{"name": "Premium", "price": 5000}, {"name": "Business", "price": 3000}, {"name": "Business Plus", "price": 4000}]
    projects = [p["name"] for p in projects_db] if projects_db else ["Üzvlük", "Sərgi", "Təlim/Proqram", "ICMA", "Sosial layihə"]
    sectors = [s["name"] for s in sectors_db] if sectors_db else ["İnşaat", "Təhsil", "Qida", "İKT", "Logistika", "Maliyyə", "Səhiyyə", "Turizm", "Kənd təsərrüfatı", "İstehsalat", "Pərakəndə satış", "Xidmət", "Digər"]
    marsol_representatives = [u["name"] for u in users_db if u.get("name")]
    sub_sectors = {}
    for ss in sub_sectors_db:
        sec = ss.get("sector", "")
        if sec not in sub_sectors:
            sub_sectors[sec] = []
        sub_sectors[sec].append(ss["name"])
    positions = [p["name"] for p in positions_db] if positions_db else ["Direktor", "Təsisçi", "Baş direktor", "İcraçı direktor", "Kommersiya direktoru", "Maliyyə direktoru"]
    activities = [a["name"] for a in activities_db] if activities_db else ["Networking", "Təlim", "Sərgi", "Forum", "Mentorluq", "İş birliyi"]
    regions = [r["name"] for r in regions_db] if regions_db else ["Bakı", "Sumqayıt", "Gəncə", "Lənkəran", "Mingəçevir", "Şəki", "Şirvan", "Naxçıvan", "Abşeron", "Digər"]
    marsol_companies = [m["name"] for m in marsol_companies_db] if marsol_companies_db else ["Marsol Group", "Marsol Events", "Marsol Media", "Marsol Academy"]
    
    return {
        "sectors": sectors,
        "packages": [p["name"] for p in packages],
        "packages_with_prices": packages,
        "company_sizes": await _get_setting_list("company_sizes", ["Böyük", "Orta", "Kiçik", "Mikro"]),
        "marsol_representatives": marsol_representatives,
        "projects": projects,
        "departments": await _get_setting_list("departments", ["Satış", "Marketing", "HR", "Maliyyə", "Layihə", "İT", "İdarəetmə"]),
        "meeting_types": await _get_setting_list("meeting_types", ["Satış görüşü", "Daxili iclas", "Müştəri görüşü", "Partnyor görüşü", "Təqdimat"]),
        "task_statuses": await _get_setting_list("task_statuses", ["Gözləyir", "İcrada", "Tamamlandı", "Ləğv edildi"]),
        "priorities": await _get_setting_list("priorities", ["Yüksək", "Orta", "Aşağı"]),
        "expense_categories": [
            {"name": "Əməliyyat xərcləri", "subcategories": ["Əmək haqqı", "Bonus", "Ofis icarəsi", "Kommunal", "Ofis xərcləri"]},
            {"name": "Marketinq xərcləri", "subcategories": ["Sosial Media reklamı", "Outdoor reklam", "Promo materiallar"]},
            {"name": "Layihə xərcləri", "subcategories": ["Məkan icarəsi", "Texniki avadanlıq", "Aparıcı", "Musiqi", "Çap materialları"]},
            {"name": "Texniki xərclər", "subcategories": ["Hosting", "Domen", "Proqram təminatı", "İT xidmətləri"]},
            {"name": "Satış xərcləri", "subcategories": ["Müştəri görüş xərcləri", "Hədiyyə"]},
            {"name": "Digər xərclər", "subcategories": ["Cərimələr", "Hüquqi xidmətlər"]}
        ],
        "expense_types": await _get_setting_list("expense_types", ["Əməliyyat", "Marketinq", "Layihə", "Texniki", "Satış", "Digər"]),
        "reference_sources": await _get_setting_list("reference_sources", ["Şirkət", "Şəxs", "Media", "Digər"]),
        "statuses": await _get_setting_list("company_statuses", ["Aktiv", "Qeyri-aktiv", "Gözləmədə"]),
        "company_statuses": await _get_setting_list("company_statuses", ["Aktiv", "Qeyri-aktiv", "Gözləmədə"]),
        "organization_forms": await _get_setting_list("organization_forms", ["MMC", "ASC", "QSC", "Fərdi sahibkar", "Fond", "İB", "Digər"]),
        "contract_statuses": await _get_setting_list("contract_statuses", ["Aktiv", "Yeni", "Yeniləmə gözlənir", "Bitdi", "Ləğv edilib"]),
        "citizenships": await _get_setting_list("citizenships", ["Azərbaycan", "Türkiyə", "Rusiya", "Gürcüstan", "Ukrayna", "Digər"]),
        "employee_statuses": await _get_setting_list("employee_statuses", ["Aktiv", "Məzuniyyətdə", "Xəstələnib", "İşdən çıxıb"]),
        "marital_statuses": await _get_setting_list("marital_statuses", ["Subay", "Evli", "Boşanmış", "Dul"]),
        "payment_methods": await _get_setting_list("payment_methods", ["Nağd", "Köçürmə", "Kart", "Hissə-hissə"]),
        "invitation_response_statuses": await _get_setting_list("invitation_response_statuses", ["Gözləmədə", "Qatıldı", "Rədd etdi", "Cavab vermədi"]),
        "sub_sectors": sub_sectors,
        "positions": positions,
        "activities": activities,
        "regions": regions,
        "marsol_companies": marsol_companies,
        "education_levels": await _get_setting_list("education_levels", ["Orta təhsil", "Sub bakalavr", "Bakalavr", "Magistratura", "Doktorantura"]),
        "event_types": await _get_setting_list("event_types", EVENT_TYPES),
        "package_quotas": await get_package_quotas(),
        "lead_sources": await _get_setting_list("lead_sources", ["Marketing", "Referans", "Sosial media", "Veb sayt", "Sərgi", "Soyuq zəng", "Digər"]),
        "lead_statuses": await _get_setting_list("lead_statuses", ["Yeni", "Əlaqə quruldu", "Görüş təyin edildi", "Təklif göndərildi", "Danışıqda", "Üzv oldu", "Satıldı", "İmtina"]),
        "sale_types": await _get_setting_list("sale_types", ["Üzvlük", "Sərgi stendi", "Tur (Daxili)", "Tur (Xarici)", "Təlim", "Digər"]),
    }

# Get companies for select dropdown (simplified)
@api_router.get("/options/companies")
async def get_companies_for_select(current_user: dict = Depends(get_current_user)):
    companies = await db.companies.find({}, {"_id": 0, "id": 1, "brand_name": 1, "owner_name": 1, "package": 1, "joined_project": 1, "sector": 1, "sub_sector": 1, "owner_phone": 1, "company_phone": 1, "status": 1}).to_list(1000)
    return companies

# ==================== PACKAGES MANAGEMENT ====================

@api_router.get("/settings/packages")
async def get_packages(current_user: dict = Depends(get_current_user)):
    packages = await db.packages.find({}, {"_id": 0}).to_list(100)
    if not packages:
        # Return default packages
        return [
            {"id": "1", "name": "Premium", "description": "Premium üzvlük paketi", "price": 5000, "invitation_count": 12},
            {"id": "2", "name": "Business", "description": "Business üzvlük paketi", "price": 3000, "invitation_count": 15},
            {"id": "3", "name": "Business Plus", "description": "Business Plus üzvlük paketi", "price": 4000, "invitation_count": 25},
            {"id": "4", "name": "Sponsor", "description": "Sponsor paketi", "price": 8000, "invitation_count": 40}
        ]
    return packages

@api_router.post("/settings/packages")
async def create_package(package_data: dict, current_user: dict = Depends(get_current_user)):
    package_id = str(uuid.uuid4())
    package_doc = {
        "id": package_id,
        "name": package_data.get("name"),
        "description": package_data.get("description", ""),
        "price": package_data.get("price", 0),
        "invitation_count": package_data.get("invitation_count", 0),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.packages.insert_one(package_doc)
    package_doc.pop("_id", None)
    return package_doc

@api_router.put("/settings/packages/{package_id}")
async def update_package(package_id: str, package_data: dict, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in package_data.items() if v is not None}
    result = await db.packages.update_one({"id": package_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Paket tapılmadı")
    package = await db.packages.find_one({"id": package_id}, {"_id": 0})
    return package

@api_router.delete("/settings/packages/{package_id}")
async def delete_package(package_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.packages.delete_one({"id": package_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Paket tapılmadı")
    return {"message": "Paket silindi"}


# ==================== PACKAGE SERVICES (XİDMƏTLƏR) ====================
@api_router.get("/settings/packages/{package_id}/services")
async def list_package_services(package_id: str, current_user: dict = Depends(get_current_user)):
    pkg = await db.packages.find_one({"id": package_id}, {"_id": 0, "id": 1, "services": 1})
    if not pkg:
        raise HTTPException(status_code=404, detail="Paket tapılmadı")
    services = pkg.get("services") or []
    services.sort(key=lambda s: (s.get("sort_order") or 0, s.get("name") or ""))
    return services


@api_router.post("/settings/packages/{package_id}/services")
async def add_package_service(package_id: str, data: dict, current_user: dict = Depends(check_permission("settings", "write"))):
    pkg = await db.packages.find_one({"id": package_id}, {"_id": 0})
    if not pkg:
        raise HTTPException(status_code=404, detail="Paket tapılmadı")
    services = pkg.get("services") or []
    new_svc = {
        "id": str(uuid.uuid4()),
        "name": (data.get("name") or "").strip(),
        "description": (data.get("description") or "").strip(),
        "value": (data.get("value") or "").strip(),
        "included": bool(data.get("included", True)),
        "sort_order": int(data.get("sort_order") if data.get("sort_order") is not None else len(services)),
    }
    if not new_svc["name"]:
        raise HTTPException(status_code=400, detail="Xidmət adı boş ola bilməz")
    services.append(new_svc)
    await db.packages.update_one({"id": package_id}, {"$set": {"services": services}})
    return new_svc


@api_router.put("/settings/packages/{package_id}/services/{service_id}")
async def update_package_service(package_id: str, service_id: str, data: dict, current_user: dict = Depends(check_permission("settings", "write"))):
    pkg = await db.packages.find_one({"id": package_id}, {"_id": 0})
    if not pkg:
        raise HTTPException(status_code=404, detail="Paket tapılmadı")
    services = pkg.get("services") or []
    found = False
    for s in services:
        if s.get("id") == service_id:
            for k in ("name", "description", "value"):
                if k in data and data[k] is not None:
                    s[k] = str(data[k]).strip()
            if "included" in data:
                s["included"] = bool(data["included"])
            if "sort_order" in data and data["sort_order"] is not None:
                try:
                    s["sort_order"] = int(data["sort_order"])
                except (ValueError, TypeError):
                    pass
            found = True
            break
    if not found:
        raise HTTPException(status_code=404, detail="Xidmət tapılmadı")
    await db.packages.update_one({"id": package_id}, {"$set": {"services": services}})
    return next(s for s in services if s.get("id") == service_id)


@api_router.delete("/settings/packages/{package_id}/services/{service_id}")
async def delete_package_service(package_id: str, service_id: str, current_user: dict = Depends(check_permission("settings", "write"))):
    pkg = await db.packages.find_one({"id": package_id}, {"_id": 0})
    if not pkg:
        raise HTTPException(status_code=404, detail="Paket tapılmadı")
    services = [s for s in (pkg.get("services") or []) if s.get("id") != service_id]
    await db.packages.update_one({"id": package_id}, {"$set": {"services": services}})
    return {"message": "Xidmət silindi"}


@api_router.post("/settings/packages/services/seed")
async def seed_package_services(current_user: dict = Depends(check_permission("settings", "write"))):
    """Seed the 2026 partnership package service catalog from the official brochure."""
    # Service catalog: each row gives the service definition + per-package value/inclusion.
    # Values: bool=True means a plain check, bool=False means X (not included),
    # string means a custom value (e.g. "15", "1 dəfə", "limitsiz", or descriptive text).
    catalog = [
        {"name": "Marsol Plus - Mobil Tətbiqi", "description": "Şirkət məlumatı, şəxsi profil, satınalma, kampaniya, elanlar və s.", "premium": True, "business": True, "business_plus": True},
        {"name": "B2B görüşlərə dəvət", "description": "İşgüzar səhər yeməyi, ofis-istehsalat ziyarəti, axşam ziyafəti və s.", "premium": "15", "business": "20", "business_plus": "30"},
        {"name": "Şirkətin ofis və istehsalatına ziyarət", "description": "", "premium": True, "business": True, "business_plus": True},
        {"name": "Video Müsahibə", "description": "Müsahibə Facebook, Instagram, YouTube və Marsol.az-da yayımlanır", "premium": False, "business": "1 dəfə", "business_plus": "2 dəfə"},
        {"name": "İşgüzar səhər yeməyində ətraflı təqdimat", "description": "Video təqdimat + 5 dəq. çıxış imkanı", "premium": False, "business": False, "business_plus": "1 dəfə"},
        {"name": "Region sahibkarları ilə onlayn görüşlər", "description": "", "premium": True, "business": True, "business_plus": True},
        {"name": "Region konfrans və tədbirlərə dəvət", "description": "", "premium": False, "business": True, "business_plus": True},
        {"name": "Rəsmi konfrans çıxış imkanı", "description": "", "premium": True, "business": True, "business_plus": True},
        {"name": "Partnyorlarla əlaqəyə dəstək", "description": "", "premium": True, "business": True, "business_plus": True},
        {"name": "Kampaniya və endirimlər", "description": "Partnyorların kampaniya və endirimlərdən yararlanma imkanı", "premium": True, "business": True, "business_plus": True},
        {"name": "E-mail göndərilməsi", "description": "Kampaniya, xidmət və məlumat xarakterli məktubların e-mail üzərindən partnyorlara göndərilməsi", "premium": True, "business": True, "business_plus": True},
        {"name": "Marsol.az partnyor səhifəsi", "description": "Partnyorlarımız bölümündə şirkətiniz haqqında məlumatın yerləşdirilməsi", "premium": True, "business": True, "business_plus": "Xüsusi paylaşım, ətraflı məlumat"},
        {"name": "Instagram paylaşımı", "description": "Marsol Group ilə əməkdaşlıq posterinin həftəyə bölməsində paylaşımı", "premium": "5 dəfə", "business": "7 dəfə", "business_plus": "limitsiz"},
        {"name": "Facebook paylaşımı", "description": "MARSOL Group ilə əməkdaşlıq posterinin paylaşımı", "premium": False, "business": False, "business_plus": True},
        {"name": "Partnyorların məkan açılışlarının təşkili", "description": "2 həftə öncədən məlumat verilməlidir", "premium": True, "business": True, "business_plus": True},
        {"name": "Dövlət qurumlarıyla görüşlər", "description": "Müəyyən mövzularda təşkil olunan müzakirə görüşlərində iştirak", "premium": True, "business": True, "business_plus": True},
        {"name": "Sosial fəaliyyətlərə dəvət *", "description": "İdman oyunları, intellektual yarışlar, ölkədaxili və ölkəxarici turlar və s. (ulduz - əlavə ödəniş)", "premium": True, "business": True, "business_plus": True},
        {"name": "Marsol Academy *", "description": "Müxtəlif mövzularda təşkil olunan təlimlərdə iştirak imkanı (ulduz - əlavə ödəniş)", "premium": True, "business": True, "business_plus": True},
        {"name": "İş Adamları Cəmiyyətinin Milli Assosiasiyası (İCMA) üzvlüyünə qəbul", "description": "İCMA üzvlüyünün xidmətləri ilə yaxından tanış olmaq üçün QR kodu oxudun.", "premium": True, "business": True, "business_plus": True},
    ]

    # Map package name (case-insensitive) to slot key in catalog
    name_map = {
        "premium": "premium",
        "business": "business",
        "business plus": "business_plus",
        "business+": "business_plus",
        "businessplus": "business_plus",
    }

    packages = await db.packages.find({}, {"_id": 0}).to_list(100)
    if not packages:
        # Create the 4 default packages so seed has something to attach to
        defaults = [
            {"id": str(uuid.uuid4()), "name": "Premium", "description": "Premium üzvlük paketi", "price": 2000, "invitation_count": 12, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Business", "description": "Business üzvlük paketi", "price": 2800, "invitation_count": 15, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Business Plus", "description": "Business+ üzvlük paketi", "price": 4500, "invitation_count": 25, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Sponsor", "description": "Sponsor paketi", "price": 8000, "invitation_count": 40, "created_at": datetime.now(timezone.utc).isoformat()},
        ]
        await db.packages.insert_many([{**p} for p in defaults])
        packages = await db.packages.find({}, {"_id": 0}).to_list(100)

    summary = []
    for pkg in packages:
        slot = name_map.get((pkg.get("name") or "").strip().lower())
        if not slot:
            continue  # Skip non-standard packages like Sponsor
        services = []
        for idx, row in enumerate(catalog):
            raw = row[slot]
            included = bool(raw)
            value = raw if isinstance(raw, str) else ""
            services.append({
                "id": str(uuid.uuid4()),
                "name": row["name"],
                "description": row["description"],
                "value": value,
                "included": included,
                "sort_order": idx,
            })
        await db.packages.update_one({"id": pkg["id"]}, {"$set": {"services": services}})
        summary.append({"package": pkg.get("name"), "service_count": len(services)})

    return {"seeded": summary, "catalog_size": len(catalog)}


# ==================== SERVICE USAGE TRACKING ====================
async def _resolve_company_package(company: dict):
    """Return (package_doc, services) for the company's current package, or (None, [])."""
    if not company:
        return None, []
    pkg_name = (company.get("package") or "").strip()
    if not pkg_name:
        return None, []
    pkg = await db.packages.find_one({"name": pkg_name}, {"_id": 0})
    if not pkg:
        return None, []
    return pkg, pkg.get("services") or []


async def auto_track_service_usage(company_name: str, service_name_keywords, *, used_date: str = "", quantity: int = 1, notes: str = "", related_object_type: str = "", related_object_id: str = "", created_by: str = "system"):
    """Create a service_usage record by matching a company by brand_name and a service
    by checking if any of the given keywords are contained in the service name.

    Idempotent: same (company_id, service_id, related_object_type, related_object_id)
    triple is upserted instead of duplicated.
    """
    if not company_name:
        return None
    company = await db.companies.find_one({"brand_name": company_name}, {"_id": 0})
    if not company:
        return None
    pkg, services = await _resolve_company_package(company)
    if not services:
        return None
    keywords = [k.lower() for k in (service_name_keywords if isinstance(service_name_keywords, (list, tuple)) else [service_name_keywords])]
    matched = next(
        (s for s in services if s.get("included") and any(kw in (s.get("name") or "").lower() for kw in keywords)),
        None,
    )
    if not matched:
        return None
    # Idempotency: skip if a record already links to the same source object
    if related_object_type and related_object_id:
        dup = await db.service_usage.find_one({
            "company_id": company["id"],
            "service_id": matched["id"],
            "related_object_type": related_object_type,
            "related_object_id": related_object_id,
        }, {"_id": 0, "id": 1})
        if dup:
            return dup
    doc = {
        "id": str(uuid.uuid4()),
        "company_id": company["id"],
        "package_id": pkg.get("id") if pkg else None,
        "package_name": pkg.get("name") if pkg else (company.get("package") or ""),
        "service_id": matched["id"],
        "service_name": matched["name"],
        "quantity": max(int(quantity or 1), 1),
        "used_date": (used_date or datetime.now(timezone.utc).strftime("%Y-%m-%d"))[:10],
        "notes": notes,
        "related_object_type": related_object_type,
        "related_object_id": related_object_id,
        "auto": True,
        "created_by": created_by,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.service_usage.insert_one(doc)
    doc.pop("_id", None)
    return doc


# Mapping of meeting_type → list of service-name keywords to match the package catalog.
_MEETING_TYPE_TO_SERVICE_KEYWORDS = {
    "B2B": ["b2b"],
    "Səhər yeməyi": ["səhər yeməyi", "b2b"],
    "İşgüzar səhər yeməyi": ["səhər yeməyi", "b2b"],
    "Ofis ziyarət": ["ofis", "ziyarət"],
    "Ofis ziyarəti": ["ofis", "ziyarət"],
    "Şirkət ziyarəti": ["ofis", "ziyarət"],
    "Axşam ziyafəti": ["b2b"],
    "Region görüşü": ["region sahibkar", "region"],
    "Onlayn görüş": ["region sahibkar"],
    "Dövlət qurumu": ["dövlət"],
    "Dövlət görüşü": ["dövlət"],
    "Dövlət qurumu görüşü": ["dövlət"],
    "Konfrans": ["rəsmi konfrans", "konfrans"],
    "Rəsmi konfrans": ["rəsmi konfrans"],
    "Marsol Academy": ["marsol academy"],
    "Akademiya": ["marsol academy"],
    "Sosial fəaliyyət": ["sosial fəaliyyət"],
}


async def _auto_track_meeting_service(meeting_doc: dict):
    """Best-effort auto-tracking when a meeting is created/updated."""
    try:
        company_name = meeting_doc.get("company") or ""
        if not company_name:
            return
        meeting_type = (meeting_doc.get("meeting_type") or "").strip()
        keywords = _MEETING_TYPE_TO_SERVICE_KEYWORDS.get(meeting_type)
        if not keywords:
            # Try fuzzy match on lowercased meeting_type
            for k, v in _MEETING_TYPE_TO_SERVICE_KEYWORDS.items():
                if meeting_type and k.lower() in meeting_type.lower():
                    keywords = v
                    break
        if not keywords:
            return
        await auto_track_service_usage(
            company_name=company_name,
            service_name_keywords=keywords,
            used_date=meeting_doc.get("date", ""),
            notes=f"Avto: {meeting_type} — {meeting_doc.get('contact_person') or meeting_doc.get('employee') or ''}".strip(" -"),
            related_object_type="meeting",
            related_object_id=meeting_doc.get("id", ""),
            created_by=meeting_doc.get("created_by") or "auto",
        )
    except Exception as exc:
        logger.warning(f"auto_track_meeting_service failed: {exc}")


def _parse_quota(value: str):
    """Extract first numeric value from `value` (e.g. '15', '5 dəfə', '2 dəfə'). None if none."""
    if not value:
        return None
    val = str(value).strip().lower()
    if val in ("limitsiz", "limitsiz."):
        return None  # treated as unlimited; surfaced in `unlimited` flag
    digits = ""
    for ch in val:
        if ch.isdigit():
            digits += ch
        elif digits:
            break
    return int(digits) if digits else None


@api_router.get("/companies/{company_id}/service-usage")
async def list_service_usage(company_id: str, service_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    company = await db.companies.find_one({"id": company_id}, {"_id": 0, "id": 1})
    if not company:
        raise HTTPException(status_code=404, detail="Şirkət tapılmadı")
    query = {"company_id": company_id}
    if service_id:
        query["service_id"] = service_id
    rows = await db.service_usage.find(query, {"_id": 0}).sort("used_date", -1).to_list(2000)
    return rows


@api_router.post("/companies/{company_id}/service-usage")
async def create_service_usage(company_id: str, data: dict, current_user: dict = Depends(check_permission("members", "write"))):
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Şirkət tapılmadı")
    pkg, services = await _resolve_company_package(company)
    service_id = data.get("service_id")
    service_name = (data.get("service_name") or "").strip()
    matched = next((s for s in services if s.get("id") == service_id or s.get("name") == service_name), None)
    if matched:
        service_id = matched["id"]
        service_name = matched["name"]
    elif not service_name:
        raise HTTPException(status_code=400, detail="Xidmət adı və ya ID lazımdır")

    try:
        quantity = int(data.get("quantity") or 1)
    except (ValueError, TypeError):
        quantity = 1
    if quantity <= 0:
        quantity = 1

    doc = {
        "id": str(uuid.uuid4()),
        "company_id": company_id,
        "package_id": pkg.get("id") if pkg else None,
        "package_name": pkg.get("name") if pkg else (company.get("package") or ""),
        "service_id": service_id,
        "service_name": service_name,
        "quantity": quantity,
        "used_date": (data.get("used_date") or datetime.now(timezone.utc).strftime("%Y-%m-%d"))[:10],
        "notes": (data.get("notes") or "").strip(),
        "related_object_type": data.get("related_object_type") or "",
        "related_object_id": data.get("related_object_id") or "",
        "auto": bool(data.get("auto", False)),
        "created_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.service_usage.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/service-usage/{usage_id}")
async def update_service_usage(usage_id: str, data: dict, current_user: dict = Depends(check_permission("members", "write"))):
    existing = await db.service_usage.find_one({"id": usage_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Qeyd tapılmadı")
    update = {}
    for k in ("service_name", "service_id", "notes", "used_date"):
        if k in data and data[k] is not None:
            update[k] = str(data[k]).strip() if k != "service_id" else data[k]
    if "quantity" in data and data["quantity"] is not None:
        try:
            q = int(data["quantity"])
            update["quantity"] = q if q > 0 else 1
        except (ValueError, TypeError):
            pass
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.service_usage.update_one({"id": usage_id}, {"$set": update})
    doc = await db.service_usage.find_one({"id": usage_id}, {"_id": 0})
    return doc


@api_router.delete("/service-usage/{usage_id}")
async def delete_service_usage(usage_id: str, current_user: dict = Depends(check_permission("members", "write"))):
    existing = await db.service_usage.find_one({"id": usage_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Qeyd tapılmadı")
    await db.service_usage.delete_one({"id": usage_id})
    return {"message": "Qeyd silindi"}


@api_router.get("/companies/{company_id}/service-stats")
async def company_service_stats(company_id: str, current_user: dict = Depends(get_current_user)):
    """Return per-service usage stats for a company.

    Each row: {service_id, name, value, included, quota, used, remaining, unlimited, last_used, history_count}.
    """
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Şirkət tapılmadı")
    pkg, services = await _resolve_company_package(company)
    usage_rows = await db.service_usage.find({"company_id": company_id}, {"_id": 0}).to_list(5000)

    # Group usage by service_id (or by name when id missing)
    grouped = {}
    for u in usage_rows:
        key = u.get("service_id") or u.get("service_name")
        bucket = grouped.setdefault(key, {"used": 0, "last_used": "", "history": []})
        bucket["used"] += int(u.get("quantity") or 1)
        if (u.get("used_date") or "") > (bucket["last_used"] or ""):
            bucket["last_used"] = u.get("used_date") or ""
        bucket["history"].append(u)

    rows = []
    seen_ids = set()
    for s in services:
        sid = s.get("id")
        seen_ids.add(sid)
        bucket = grouped.get(sid) or grouped.get(s.get("name")) or {"used": 0, "last_used": "", "history": []}
        quota_raw = (s.get("value") or "").strip().lower()
        unlimited = quota_raw == "limitsiz"
        quota = _parse_quota(s.get("value")) if not unlimited else None
        if quota is None and not unlimited and s.get("included"):
            # boolean-style service (just "yes/no")
            pass
        used = int(bucket["used"])
        remaining = None
        if quota is not None:
            remaining = max(quota - used, 0)
        rows.append({
            "service_id": sid,
            "name": s.get("name"),
            "description": s.get("description", ""),
            "value": s.get("value", ""),
            "included": s.get("included", False),
            "quota": quota,
            "unlimited": unlimited,
            "used": used,
            "remaining": remaining,
            "last_used": bucket["last_used"],
            "history_count": len(bucket["history"]),
            "sort_order": s.get("sort_order", 0),
        })

    # Append any usage records that don't match a current service (e.g. legacy / package change)
    for key, bucket in grouped.items():
        if key in seen_ids:
            continue
        # Find a name from one of the entries
        name = next((h.get("service_name") for h in bucket["history"] if h.get("service_name")), str(key))
        rows.append({
            "service_id": key,
            "name": name,
            "description": "",
            "value": "",
            "included": False,
            "quota": None,
            "unlimited": False,
            "used": bucket["used"],
            "remaining": None,
            "last_used": bucket["last_used"],
            "history_count": len(bucket["history"]),
            "sort_order": 9999,
            "legacy": True,
        })

    rows.sort(key=lambda r: (r.get("sort_order") or 0, r.get("name") or ""))
    return {
        "package_name": pkg.get("name") if pkg else (company.get("package") or ""),
        "package_id": pkg.get("id") if pkg else None,
        "services": rows,
    }


@api_router.get("/dashboard/service-usage-stats")
async def dashboard_service_usage_stats(month: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Top services used this month + total usage counts.

    Optional `month` param in YYYY-MM format; defaults to current month.
    """
    if month:
        try:
            datetime.strptime(month + "-01", "%Y-%m-%d")
            target = month
        except (ValueError, TypeError):
            target = datetime.now(timezone.utc).strftime("%Y-%m")
    else:
        target = datetime.now(timezone.utc).strftime("%Y-%m")
    start = f"{target}-01"
    # Last day of target month
    yyyy, mm = target.split("-")
    next_month = (datetime.strptime(start, "%Y-%m-%d") + timedelta(days=32)).replace(day=1).strftime("%Y-%m-%d")
    rows = await db.service_usage.find(
        {"used_date": {"$gte": start, "$lt": next_month}},
        {"_id": 0, "service_id": 1, "service_name": 1, "quantity": 1, "company_id": 1, "package_name": 1},
    ).to_list(10000)
    bucket = {}
    for r in rows:
        key = r.get("service_id") or r.get("service_name") or "unknown"
        b = bucket.setdefault(key, {
            "service_id": r.get("service_id"),
            "service_name": r.get("service_name") or "",
            "total_quantity": 0,
            "company_count": set(),
        })
        b["total_quantity"] += int(r.get("quantity") or 1)
        if r.get("company_id"):
            b["company_count"].add(r["company_id"])
    summary = []
    for v in bucket.values():
        summary.append({
            "service_id": v["service_id"],
            "service_name": v["service_name"],
            "total_quantity": v["total_quantity"],
            "company_count": len(v["company_count"]),
        })
    summary.sort(key=lambda x: x["total_quantity"], reverse=True)
    return {"month": target, "top_services": summary[:10], "total_records": len(rows)}

# ==================== PROJECTS MANAGEMENT ====================

@api_router.get("/settings/projects")
async def get_projects(current_user: dict = Depends(get_current_user)):
    projects = await db.projects.find({}, {"_id": 0}).to_list(100)
    if not projects:
        # Return default projects
        return [
            {"id": "1", "name": "Üzvlük", "description": "Üzvlük layihəsi"},
            {"id": "2", "name": "Sərgi", "description": "Sərgi layihəsi"},
            {"id": "3", "name": "Təlim/Proqram", "description": "Təlim və proqramlar"},
            {"id": "4", "name": "ICMA", "description": "ICMA layihəsi"},
            {"id": "5", "name": "Sosial layihə", "description": "Sosial layihələr"}
        ]
    return projects

@api_router.post("/settings/projects")
async def create_project(project_data: dict, current_user: dict = Depends(get_current_user)):
    project_id = str(uuid.uuid4())
    project_doc = {
        "id": project_id,
        "name": project_data.get("name"),
        "description": project_data.get("description", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.projects.insert_one(project_doc)
    project_doc.pop("_id", None)
    return project_doc

@api_router.put("/settings/projects/{project_id}")
async def update_project(project_id: str, project_data: dict, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in project_data.items() if v is not None}
    result = await db.projects.update_one({"id": project_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Layihə tapılmadı")
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    return project

@api_router.delete("/settings/projects/{project_id}")
async def delete_project(project_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.projects.delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Layihə tapılmadı")
    return {"message": "Layihə silindi"}

# ==================== FILE UPLOAD ====================

@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Legacy upload — now backed by Cloudinary so files survive Render restarts.

    Returns the same shape as before (url + filename) plus Cloudinary metadata.
    """
    raw = await file.read()
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail=f"Fayl çox böyükdür (max {MAX_UPLOAD_BYTES // (1024*1024)} MB)")
    try:
        result = _cl_upload(raw, file.filename or "file", "marsol/files")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Yükləmə alınmadı: {str(e)[:200]}")
    return {
        "url": result["url"],
        "filename": file.filename,
        "stored_name": result["public_id"],
        "public_id": result["public_id"],
        "resource_type": result["resource_type"],
    }

@api_router.post("/public/upload")
async def public_upload_file(file: UploadFile = File(...)):
    """Public unauthenticated upload (used by public forms). Cloudinary-backed."""
    raw = await file.read()
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail=f"Fayl çox böyükdür (max {MAX_UPLOAD_BYTES // (1024*1024)} MB)")
    try:
        result = _cl_upload(raw, file.filename or "file", "marsol/files")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Yükləmə alınmadı: {str(e)[:200]}")
    return {"url": result["url"], "filename": file.filename, "public_id": result["public_id"]}

# ==================== BRANDING (LOGOLAR) ====================

DEFAULT_LOGO = "https://customer-assets.emergentagent.com/job_03e89fda-1599-48f3-846d-f1d3e818b1fa/artifacts/h0q248dw_Marsol.png"

@api_router.get("/public/branding")
async def get_public_branding():
    """Public read of branding config (used by Login page, PublicForm, etc)."""
    doc = await db.app_config.find_one({"key": "branding"}, {"_id": 0}) or {}
    return {
        "sidebar_logo_url": doc.get("sidebar_logo_url") or "",
        "main_logo_url": doc.get("main_logo_url") or DEFAULT_LOGO,
    }

@api_router.get("/settings/branding")
async def get_branding(current_user: dict = Depends(get_current_user)):
    doc = await db.app_config.find_one({"key": "branding"}, {"_id": 0}) or {}
    return {
        "sidebar_logo_url": doc.get("sidebar_logo_url") or "",
        "main_logo_url": doc.get("main_logo_url") or DEFAULT_LOGO,
    }

@api_router.put("/settings/branding")
async def update_branding(data: dict, current_user: dict = Depends(check_permission("settings", "write"))):
    update = {
        "key": "branding",
        "sidebar_logo_url": (data.get("sidebar_logo_url") or "").strip(),
        "main_logo_url": (data.get("main_logo_url") or "").strip(),
        "updated_by": current_user.get("name", ""),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.app_config.update_one({"key": "branding"}, {"$set": update}, upsert=True)
    return {
        "sidebar_logo_url": update["sidebar_logo_url"],
        "main_logo_url": update["main_logo_url"] or DEFAULT_LOGO,
    }


# ==================== SUB-SECTORS ====================

@api_router.get("/settings/sub-sectors")
async def get_sub_sectors(sector: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if sector:
        query["sector"] = sector
    subs = await db.sub_sectors.find(query, {"_id": 0}).to_list(500)
    return subs

@api_router.post("/settings/sub-sectors")
async def create_sub_sector(data: dict, current_user: dict = Depends(get_current_user)):
    sub_id = str(uuid.uuid4())
    doc = {"id": sub_id, "name": data.get("name"), "sector": data.get("sector"), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.sub_sectors.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/settings/sub-sectors/{sub_id}")
async def update_sub_sector(sub_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    result = await db.sub_sectors.update_one({"id": sub_id}, {"$set": {k: v for k, v in data.items() if v is not None}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Alt sektor tapılmadı")
    return await db.sub_sectors.find_one({"id": sub_id}, {"_id": 0})

@api_router.delete("/settings/sub-sectors/{sub_id}")
async def delete_sub_sector(sub_id: str, current_user: dict = Depends(get_current_user)):
    await db.sub_sectors.delete_one({"id": sub_id})
    return {"message": "Alt sektor silindi"}

# ==================== POSITIONS (VƏZİFƏLƏR) ====================

@api_router.get("/settings/positions")
async def get_positions(current_user: dict = Depends(get_current_user)):
    positions = await db.positions.find({}, {"_id": 0}).to_list(200)
    if not positions:
        return [{"id": str(i), "name": n} for i, n in enumerate(["Direktor", "Təsisçi", "Baş direktor", "İcraçı direktor", "Kommersiya direktoru", "Maliyyə direktoru"])]
    return positions

@api_router.post("/settings/positions")
async def create_position(data: dict, current_user: dict = Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), "name": data.get("name"), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.positions.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/settings/positions/{pos_id}")
async def delete_position(pos_id: str, current_user: dict = Depends(get_current_user)):
    await db.positions.delete_one({"id": pos_id})
    return {"message": "Vəzifə silindi"}

# ==================== ACTIVITIES (FƏALİYYƏTLƏR) ====================

@api_router.get("/settings/activities")
async def get_activities(current_user: dict = Depends(get_current_user)):
    acts = await db.activities.find({}, {"_id": 0}).to_list(200)
    if not acts:
        return [{"id": str(i), "name": n} for i, n in enumerate(["Networking", "Təlim", "Sərgi", "Forum", "Mentorluq", "İş birliyi"])]
    return acts

@api_router.post("/settings/activities")
async def create_activity(data: dict, current_user: dict = Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), "name": data.get("name"), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.activities.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/settings/activities/{act_id}")
async def delete_activity(act_id: str, current_user: dict = Depends(get_current_user)):
    await db.activities.delete_one({"id": act_id})
    return {"message": "Fəaliyyət silindi"}

# ==================== REGIONS (REGİONLAR) ====================

@api_router.get("/settings/regions")
async def get_regions(current_user: dict = Depends(get_current_user)):
    regions = await db.regions.find({}, {"_id": 0}).to_list(200)
    if not regions:
        return [{"id": str(i), "name": n} for i, n in enumerate(["Bakı", "Sumqayıt", "Gəncə", "Lənkəran", "Mingəçevir", "Şəki", "Şirvan", "Naxçıvan", "Abşeron", "Digər"])]
    return regions

@api_router.post("/settings/regions")
async def create_region(data: dict, current_user: dict = Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), "name": data.get("name"), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.regions.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/settings/regions/{region_id}")
async def delete_region(region_id: str, current_user: dict = Depends(get_current_user)):
    await db.regions.delete_one({"id": region_id})
    return {"message": "Region silindi"}

# ==================== MARSOL COMPANIES (MÜƏSSİSƏLƏR) ====================

@api_router.get("/settings/marsol-companies")
async def get_marsol_companies(current_user: dict = Depends(get_current_user)):
    items = await db.marsol_companies.find({}, {"_id": 0}).to_list(100)
    if not items:
        return [{"id": str(i), "name": n} for i, n in enumerate(["Marsol Group", "Marsol Events", "Marsol Media", "Marsol Academy"])]
    return items

@api_router.post("/settings/marsol-companies")
async def create_marsol_company(data: dict, current_user: dict = Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), "name": data.get("name"), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.marsol_companies.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.post("/settings/tenant/repair")
async def repair_tenant_assignments(current_user: dict = Depends(get_current_user)):
    """Admin-only: re-run the smart tenant backfill. Repairs legacy records
    whose tenant got mis-assigned to the default by looking up the creator's
    actual müəssisə. Useful if the production data shows hidden legacy items.
    """
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Yalnız admin")
    # Clear the v2 flag so the smart re-backfill re-scans every record.
    collections = ["tasks", "meetings", "companies", "sales_leads", "assemblies", "project_events", "notes", "files"]
    for c in collections:
        await db[c].update_many({}, {"$unset": {"tenant_backfill_v2": ""}})
    first = await db.marsol_companies.find_one({}, {"_id": 0, "name": 1})
    default_name = (first.get("name") if first else None) or "Marsol Group"
    await _smart_retenant_records(default_name)
    # Count of records per tenant after repair
    summary = {}
    for c in collections:
        pipeline = [{"$group": {"_id": "$marsol_company", "count": {"$sum": 1}}}]
        rows = await db[c].aggregate(pipeline).to_list(50)
        summary[c] = {r["_id"] or "_none_": r["count"] for r in rows}
    return {"message": "Tenant repair tamamlandı", "summary": summary}

@api_router.delete("/settings/marsol-companies/{item_id}")
async def delete_marsol_company(item_id: str, current_user: dict = Depends(get_current_user)):
    await db.marsol_companies.delete_one({"id": item_id})
    return {"message": "Müəssisə silindi"}

# ==================== SECTORS MANAGEMENT ====================

@api_router.get("/settings/sectors")
async def get_sectors(current_user: dict = Depends(get_current_user)):
    sectors = await db.sectors.find({}, {"_id": 0}).to_list(100)
    if not sectors:
        defaults = ["İnşaat", "Təhsil", "Qida", "İKT", "Logistika", "Maliyyə", "Səhiyyə", "Turizm", "Kənd təsərrüfatı", "İstehsalat", "Pərakəndə satış", "Xidmət", "Digər"]
        return [{"id": str(i+1), "name": s} for i, s in enumerate(defaults)]
    return sectors

@api_router.post("/settings/sectors")
async def create_sector(sector_data: dict, current_user: dict = Depends(get_current_user)):
    sector_id = str(uuid.uuid4())
    sector_doc = {
        "id": sector_id,
        "name": sector_data.get("name"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.sectors.insert_one(sector_doc)
    sector_doc.pop("_id", None)
    return sector_doc

@api_router.put("/settings/sectors/{sector_id}")
async def update_sector(sector_id: str, sector_data: dict, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in sector_data.items() if v is not None}
    result = await db.sectors.update_one({"id": sector_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sektor tapılmadı")
    sector = await db.sectors.find_one({"id": sector_id}, {"_id": 0})
    return sector

@api_router.delete("/settings/sectors/{sector_id}")
async def delete_sector(sector_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.sectors.delete_one({"id": sector_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sektor tapılmadı")
    return {"message": "Sektor silindi"}

# ==================== CUSTOM FIELDS ====================

@api_router.get("/settings/custom-fields")
async def get_custom_fields(module: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if module and module != "all":
        query["module"] = module
    fields = await db.custom_fields.find(query, {"_id": 0}).to_list(500)
    return fields

@api_router.post("/settings/custom-fields")
async def create_custom_field(field_data: dict, current_user: dict = Depends(get_current_user)):
    field_id = str(uuid.uuid4())
    field_doc = {
        "id": field_id,
        "module": field_data.get("module"),
        "sub_tab": field_data.get("sub_tab", ""),
        "field_name": field_data.get("field_name"),
        "field_label": field_data.get("field_label", ""),
        "field_type": field_data.get("field_type", "text"),
        "options": field_data.get("options", []),
        "required": field_data.get("required", False),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.custom_fields.insert_one(field_doc)
    field_doc.pop("_id", None)
    return field_doc

@api_router.put("/settings/custom-fields/{field_id}")
async def update_custom_field(field_id: str, field_data: dict, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in field_data.items() if v is not None}
    result = await db.custom_fields.update_one({"id": field_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sahə tapılmadı")
    field = await db.custom_fields.find_one({"id": field_id}, {"_id": 0})
    return field

@api_router.delete("/settings/custom-fields/{field_id}")
async def delete_custom_field(field_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.custom_fields.delete_one({"id": field_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sahə tapılmadı")
    return {"message": "Sahə silindi"}

# ==================== USER MANAGEMENT ====================

@api_router.get("/settings/users")
async def get_users(current_user: dict = Depends(get_current_user)):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(500)
    return users

@api_router.post("/settings/users")
async def create_user(user_data: dict, current_user: dict = Depends(get_current_user)):
    existing = await db.users.find_one({"email": user_data.get("email")})
    if existing:
        raise HTTPException(status_code=400, detail="Bu email artıq mövcuddur")
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "email": user_data.get("email"),
        "name": user_data.get("name"),
        "password": hash_password(user_data.get("password", "123456")),
        "role": user_data.get("role", "user"),
        "department": user_data.get("department", ""),
        "marsol_company": user_data.get("marsol_company", ""),
        "phone": user_data.get("phone", ""),
        "status": user_data.get("status", "Aktiv"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    user_doc.pop("_id", None)
    user_doc.pop("password", None)
    return user_doc

@api_router.put("/settings/users/{user_id}")
async def update_user(user_id: str, user_data: dict, current_user: dict = Depends(get_current_user)):
    update_data = {}
    for k, v in user_data.items():
        if v is not None and k != "password":
            update_data[k] = v
    if "password" in user_data and user_data["password"]:
        update_data["password"] = hash_password(user_data["password"])
    if not update_data:
        raise HTTPException(status_code=400, detail="Yenilənəcək məlumat yoxdur")
    result = await db.users.update_one({"id": user_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="İstifadəçi tapılmadı")
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    return user

@api_router.delete("/settings/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if user_id == current_user.get("id"):
        raise HTTPException(status_code=400, detail="Özünüzü silə bilməzsiniz")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="İstifadəçi tapılmadı")
    return {"message": "İstifadəçi silindi"}

# ==================== SALES (SATIŞ) ====================

@api_router.get("/sales/leads")
async def get_leads(stage: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if stage and stage != "all":
        query["stage"] = stage
    leads = await db.sales_leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return leads

@api_router.post("/sales/leads")
async def create_lead(lead_data: dict, current_user: dict = Depends(get_current_user)):
    lead_id = str(uuid.uuid4())
    lead_doc = {
        "id": lead_id,
        "company_name": lead_data.get("company_name", ""),
        "contact_person": lead_data.get("contact_person", ""),
        "phone": lead_data.get("phone", ""),
        "email": lead_data.get("email", ""),
        "source": lead_data.get("source", ""),
        "stage": lead_data.get("stage", "Yeni Lead"),
        "assigned_to": lead_data.get("assigned_to", ""),
        "expected_amount": lead_data.get("expected_amount", 0),
        "package": lead_data.get("package", ""),
        "project": lead_data.get("project", ""),
        "notes": lead_data.get("notes", ""),
        "priority": lead_data.get("priority", "Orta"),
        "next_action_date": lead_data.get("next_action_date", ""),
        "created_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.sales_leads.insert_one(lead_doc)
    lead_doc.pop("_id", None)
    return lead_doc

@api_router.put("/sales/leads/{lead_id}")
async def update_lead(lead_id: str, lead_data: dict, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in lead_data.items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.sales_leads.update_one({"id": lead_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lead tapılmadı")
    lead = await db.sales_leads.find_one({"id": lead_id}, {"_id": 0})
    return lead

@api_router.delete("/sales/leads/{lead_id}")
async def delete_lead(lead_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.sales_leads.delete_one({"id": lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead tapılmadı")
    return {"message": "Lead silindi"}

@api_router.get("/sales/stats")
async def get_sales_stats(current_user: dict = Depends(get_current_user)):
    stages = ["Yeni Lead", "Əlaqə", "Təklif", "Danışıq", "Uğurlu", "Uğursuz"]
    stats = {}
    for stage in stages:
        count = await db.sales_leads.count_documents({"stage": stage})
        pipeline = [{"$match": {"stage": stage}}, {"$group": {"_id": None, "total": {"$sum": "$expected_amount"}}}]
        amount_data = await db.sales_leads.aggregate(pipeline).to_list(1)
        stats[stage] = {"count": count, "amount": amount_data[0]["total"] if amount_data else 0}
    return stats

# ==================== MESSAGES (MESAJLAR) ====================

@api_router.get("/messages/conversations")
async def get_conversations(current_user: dict = Depends(get_current_user)):
    user_id = current_user["id"]
    conversations = await db.conversations.find(
        {"participants": user_id}, {"_id": 0}
    ).sort("last_message_at", -1).to_list(100)
    return conversations

@api_router.post("/messages/conversations")
async def create_conversation(data: dict, current_user: dict = Depends(get_current_user)):
    """Create a 1-1 or group conversation.

    Body: { participant_id?: str, participant_ids?: [str], name?: str }
    - participant_id (legacy): single user → 1-1 chat (reuses existing if any)
    - participant_ids: array of users → group chat (always creates a new doc; group `name` recommended)
    """
    conv_id = str(uuid.uuid4())
    me = current_user["id"]
    raw_ids = data.get("participant_ids")
    if isinstance(raw_ids, list) and len(raw_ids) > 0:
        # Group / multi-participant flow
        unique = [pid for pid in dict.fromkeys(raw_ids) if pid and pid != me]
        if not unique:
            raise HTTPException(status_code=400, detail="Ən az 1 iştirakçı seçilməlidir")
        participants = [me] + unique
        is_group = len(participants) > 2 or bool(data.get("name"))
        # For 1-1, reuse existing conversation
        if not is_group:
            existing = await db.conversations.find_one({
                "participants": {"$all": participants, "$size": 2},
                "is_group": {"$ne": True},
            })
            if existing:
                existing.pop("_id", None)
                return existing
    else:
        # Legacy single-recipient path
        participant_id = data.get("participant_id")
        if not participant_id:
            raise HTTPException(status_code=400, detail="participant_id və ya participant_ids tələb olunur")
        participants = [me, participant_id]
        is_group = False
        existing = await db.conversations.find_one({"participants": {"$all": participants, "$size": 2}, "is_group": {"$ne": True}})
        if existing:
            existing.pop("_id", None)
            return existing

    # Build participant_names map
    user_docs = await db.users.find({"id": {"$in": participants}}, {"_id": 0, "id": 1, "name": 1}).to_list(50)
    names_map = {u["id"]: u.get("name", "") for u in user_docs}
    # Sanity: must have at least the current user resolved
    if me not in names_map:
        names_map[me] = current_user.get("name", "")

    now_iso = datetime.now(timezone.utc).isoformat()
    conv_doc = {
        "id": conv_id,
        "participants": participants,
        "participant_names": names_map,
        "is_group": is_group,
        "name": (data.get("name") or "").strip() if is_group else "",
        "created_by": me,
        "last_message": "",
        "last_message_at": now_iso,
        "created_at": now_iso,
    }
    await db.conversations.insert_one(conv_doc)
    conv_doc.pop("_id", None)
    return conv_doc


@api_router.get("/messages/unread-count")
async def messages_unread_count(current_user: dict = Depends(get_current_user)):
    """Return per-conversation unread message counts and total for the sidebar badge.

    NOTE: This route MUST be declared before the dynamic `/messages/{conversation_id}`
    one — otherwise FastAPI matches it as conversation_id='unread-count'.
    """
    me = current_user["id"]
    convs = await db.conversations.find({"participants": me}, {"_id": 0, "id": 1}).to_list(500)
    reads = {r["conversation_id"]: r.get("last_read_at", "") for r in await db.message_reads.find({"user_id": me}, {"_id": 0}).to_list(500)}
    per_conv = {}
    total = 0
    for c in convs:
        cid = c["id"]
        last_read = reads.get(cid, "")
        q = {"conversation_id": cid, "sender_id": {"$ne": me}}
        if last_read:
            q["created_at"] = {"$gt": last_read}
        cnt = await db.messages.count_documents(q)
        if cnt:
            per_conv[cid] = cnt
            total += cnt
    return {"total": total, "per_conversation": per_conv}


@api_router.get("/messages/{conversation_id}")
async def get_messages(conversation_id: str, current_user: dict = Depends(get_current_user)):
    messages = await db.messages.find(
        {"conversation_id": conversation_id}, {"_id": 0}
    ).sort("created_at", 1).to_list(500)
    return messages

@api_router.post("/messages/{conversation_id}")
async def send_message(conversation_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    msg_id = str(uuid.uuid4())
    # Validate participant
    conv = await db.conversations.find_one({"id": conversation_id, "participants": current_user["id"]}, {"_id": 0})
    if not conv:
        raise HTTPException(status_code=404, detail="Söhbət tapılmadı")
    text = (data.get("text") or "").strip()
    attachment = data.get("attachment")  # { url, name, mime_type, bytes, resource_type }
    if not text and not attachment:
        raise HTTPException(status_code=400, detail="Mesaj və ya fayl tələb olunur")
    msg_doc = {
        "id": msg_id,
        "conversation_id": conversation_id,
        "sender_id": current_user["id"],
        "sender_name": current_user["name"],
        "text": text,
        "attachment": attachment if isinstance(attachment, dict) else None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.messages.insert_one(msg_doc)
    msg_doc.pop("_id", None)

    # Update conversation preview
    preview = text or (attachment.get("name") if isinstance(attachment, dict) else "📎 Fayl")
    await db.conversations.update_one(
        {"id": conversation_id},
        {"$set": {"last_message": preview[:120], "last_message_at": msg_doc["created_at"]}}
    )

    # In-app notifications to all participants except the sender
    sender_name = current_user.get("name", "")
    conv_label = conv.get("name") or sender_name
    push_recipients = []
    for pid in conv.get("participants", []):
        if pid == current_user["id"]:
            continue
        user = await db.users.find_one({"id": pid}, {"_id": 0, "name": 1})
        if not user:
            continue
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "type": "message",
            "title": f"Yeni mesaj — {conv_label}",
            "body": (preview[:160] if preview else "Yeni mesaj"),
            "conversation_id": conversation_id,
            "sender_name": sender_name,
            "recipient_name": user.get("name", ""),
            "is_read": False,
            "created_at": msg_doc["created_at"],
        })
        if user.get("name"):
            push_recipients.append(user["name"])
    if push_recipients:
        _safe_push(
            push_recipients,
            f"Yeni mesaj — {conv_label}",
            preview[:160] if preview else "Yeni mesaj",
            link=f"/messages?conversation={conversation_id}",
            data={"type": "message", "conversation_id": conversation_id},
        )
    return msg_doc


@api_router.delete("/messages/{conversation_id}/message/{message_id}")
async def delete_message(conversation_id: str, message_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a single message. Only the sender (or admin) may delete it."""
    msg = await db.messages.find_one({"id": message_id, "conversation_id": conversation_id}, {"_id": 0})
    if not msg:
        raise HTTPException(status_code=404, detail="Mesaj tapılmadı")
    is_admin = (current_user.get("role") or "").lower() == "admin"
    if msg.get("sender_id") != current_user["id"] and not is_admin:
        raise HTTPException(status_code=403, detail="Yalnız mesaj göndərən və ya admin silə bilər")
    await db.messages.delete_one({"id": message_id})
    # Re-compute last message preview for the conversation
    last = await db.messages.find({"conversation_id": conversation_id}, {"_id": 0}).sort("created_at", -1).limit(1).to_list(1)
    if last:
        last_doc = last[0]
        preview = last_doc.get("text") or (last_doc.get("attachment") or {}).get("name") or "📎 Fayl"
        await db.conversations.update_one(
            {"id": conversation_id},
            {"$set": {"last_message": preview[:120], "last_message_at": last_doc.get("created_at", "")}}
        )
    else:
        await db.conversations.update_one({"id": conversation_id}, {"$set": {"last_message": "", "last_message_at": ""}})
    return {"deleted": True}


@api_router.put("/messages/conversations/{conversation_id}")
async def update_conversation(conversation_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Edit a group conversation: rename and/or update participants.
    Only the creator (or admin) of a group may modify it."""
    conv = await db.conversations.find_one({"id": conversation_id}, {"_id": 0})
    if not conv:
        raise HTTPException(status_code=404, detail="Söhbət tapılmadı")
    if not conv.get("is_group"):
        raise HTTPException(status_code=400, detail="Yalnız qrup söhbətləri redaktə oluna bilər")
    is_admin = (current_user.get("role") or "").lower() == "admin"
    if conv.get("created_by") != current_user["id"] and not is_admin:
        raise HTTPException(status_code=403, detail="Yalnız qrupu yaradan və ya admin redaktə edə bilər")
    update = {}
    if "name" in data:
        update["name"] = (data.get("name") or "").strip()
    if "participant_ids" in data and isinstance(data["participant_ids"], list):
        creator = conv.get("created_by")
        unique = [pid for pid in dict.fromkeys(data["participant_ids"]) if pid]
        if creator not in unique:
            unique.insert(0, creator)
        update["participants"] = unique
        user_docs = await db.users.find({"id": {"$in": unique}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
        update["participant_names"] = {u["id"]: u.get("name", "") for u in user_docs}
    if update:
        await db.conversations.update_one({"id": conversation_id}, {"$set": update})
    conv_updated = await db.conversations.find_one({"id": conversation_id}, {"_id": 0})
    return conv_updated


@api_router.delete("/messages/conversations/{conversation_id}")
async def delete_conversation(conversation_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a conversation entirely (only creator or admin)."""
    conv = await db.conversations.find_one({"id": conversation_id}, {"_id": 0})
    if not conv:
        raise HTTPException(status_code=404, detail="Söhbət tapılmadı")
    is_admin = (current_user.get("role") or "").lower() == "admin"
    if conv.get("created_by") != current_user["id"] and not is_admin:
        raise HTTPException(status_code=403, detail="Yalnız qrupu yaradan və ya admin silə bilər")
    await db.messages.delete_many({"conversation_id": conversation_id})
    await db.conversations.delete_one({"id": conversation_id})
    await db.message_reads.delete_many({"conversation_id": conversation_id})
    return {"deleted": True}


@api_router.post("/messages/{conversation_id}/mark-read")
async def mark_conversation_read(conversation_id: str, current_user: dict = Depends(get_current_user)):
    """Record the last-read timestamp for the current user in this conversation."""
    now = datetime.now(timezone.utc).isoformat()
    await db.message_reads.update_one(
        {"conversation_id": conversation_id, "user_id": current_user["id"]},
        {"$set": {"last_read_at": now}},
        upsert=True,
    )
    return {"marked": True, "at": now}


# ==================== PACKAGE QUOTA CONFIG ====================

DEFAULT_PACKAGE_QUOTAS = {
    "Premium": 12,
    "Business": 15,
    "Business Plus": 25,
    "Business+": 25,
    "Sponsor": 40,
}

async def get_package_quotas():
    packages = await db.packages.find({}, {"_id": 0, "name": 1, "invitation_count": 1}).to_list(100)
    if packages:
        quotas = {}
        for p in packages:
            name = p.get("name", "")
            count = p.get("invitation_count", 0)
            if name and count:
                quotas[name] = count
        if quotas:
            return quotas
    return DEFAULT_PACKAGE_QUOTAS

EVENT_TYPES = ["Breakfast", "Ofis ziyarəti", "Mafia", "Sosial fəaliyyət", "Təlim", "B2B görüş"]

# ==================== EVENTS (FƏALİYYƏTLƏR/GÖRÜŞLƏR) ====================

@api_router.get("/events")
async def get_events(event_type: Optional[str] = None, status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if event_type and event_type != "all":
        query["event_type"] = event_type
    if status and status != "all":
        query["status"] = status
    events = await db.events.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return events

@api_router.post("/events")
async def create_event(data: dict, current_user: dict = Depends(get_current_user)):
    event_id = str(uuid.uuid4())
    event_doc = {
        "id": event_id,
        "name": data.get("name", ""),
        "event_type": data.get("event_type", ""),
        "date": data.get("date", ""),
        "time": data.get("time", ""),
        "venue": data.get("venue", ""),
        "location_link": data.get("location_link", ""),
        "participant_limit": data.get("participant_limit", 0),
        "host_company_id": data.get("host_company_id", ""),
        "host_company_name": data.get("host_company_name", ""),
        "status": data.get("status", "Planlaşdırılır"),
        "notes": data.get("notes", ""),
        "created_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.events.insert_one(event_doc)
    event_doc.pop("_id", None)
    return event_doc

@api_router.get("/events/{event_id}")
async def get_event(event_id: str, current_user: dict = Depends(get_current_user)):
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Fəaliyyət tapılmadı")
    return event

@api_router.put("/events/{event_id}")
async def update_event(event_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in data.items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.events.update_one({"id": event_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Fəaliyyət tapılmadı")
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    return event

@api_router.delete("/events/{event_id}")
async def delete_event(event_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.events.delete_one({"id": event_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fəaliyyət tapılmadı")
    await db.invitations.delete_many({"event_id": event_id})
    return {"message": "Fəaliyyət silindi"}

@api_router.get("/events/types/list")
async def get_event_types(current_user: dict = Depends(get_current_user)):
    return EVENT_TYPES

# ==================== INVITATIONS (DƏVƏTLƏR) ====================

@api_router.get("/invitations")
async def get_invitations(event_id: Optional[str] = None, company_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if event_id:
        query["event_id"] = event_id
    if company_id:
        query["company_id"] = company_id
    invitations = await db.invitations.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    # Enrich with display_id of the related company so every invitations table
    # can show the company ID alongside its name (consistency across the app).
    company_ids = [i.get("company_id") for i in invitations if i.get("company_id")]
    if company_ids:
        companies = await db.companies.find(
            {"id": {"$in": list(set(company_ids))}},
            {"_id": 0, "id": 1, "display_id": 1},
        ).to_list(2000)
        id_map = {c["id"]: c.get("display_id", "") for c in companies}
        for inv in invitations:
            inv["company_display_id"] = id_map.get(inv.get("company_id", ""), "")
    return invitations

@api_router.post("/invitations")
async def create_invitation(data: dict, current_user: dict = Depends(get_current_user)):
    inv_id = str(uuid.uuid4())
    inv_doc = {
        "id": inv_id,
        "event_id": data.get("event_id", ""),
        "event_name": data.get("event_name", ""),
        "event_type": data.get("event_type", ""),
        "event_date": data.get("event_date", ""),
        "company_id": data.get("company_id", ""),
        "company_name": data.get("company_name", ""),
        "call_status": "Gözləyir",
        "participation_status": "",
        "obligation_deducted": False,
        "called_by": "",
        "called_at": "",
        "notes": data.get("notes", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.invitations.insert_one(inv_doc)
    inv_doc.pop("_id", None)
    return inv_doc

@api_router.post("/invitations/bulk")
async def create_bulk_invitations(data: dict, current_user: dict = Depends(get_current_user)):
    event_id = data.get("event_id", "")
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Fəaliyyət tapılmadı")
    company_ids = data.get("company_ids", [])
    created = []
    for cid in company_ids:
        existing = await db.invitations.find_one({"event_id": event_id, "company_id": cid})
        if existing:
            continue
        company = await db.companies.find_one({"id": cid}, {"_id": 0, "brand_name": 1})
        inv_id = str(uuid.uuid4())
        inv_doc = {
            "id": inv_id,
            "event_id": event_id,
            "event_name": event.get("name", ""),
            "event_type": event.get("event_type", ""),
            "event_date": event.get("date", ""),
            "company_id": cid,
            "company_name": company.get("brand_name", "") if company else "",
            "call_status": "Gözləyir",
            "participation_status": "",
            "obligation_deducted": False,
            "called_by": "",
            "called_at": "",
            "notes": "",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.invitations.insert_one(inv_doc)
        inv_doc.pop("_id", None)
        created.append(inv_doc)
    return {"created": len(created), "invitations": created}

@api_router.put("/invitations/{inv_id}/call")
async def update_invitation_call(inv_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    call_status = data.get("call_status", "")
    participation_status = data.get("participation_status", "")
    update = {
        "call_status": call_status,
        "called_by": current_user.get("name", ""),
        "called_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    if call_status == "Cavab verdi":
        update["participation_status"] = participation_status
        update["obligation_deducted"] = True
    elif call_status == "Cavab vermədi":
        update["participation_status"] = ""
        update["obligation_deducted"] = False
    result = await db.invitations.update_one({"id": inv_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dəvət tapılmadı")
    inv = await db.invitations.find_one({"id": inv_id}, {"_id": 0})
    return inv

@api_router.put("/invitations/{inv_id}/notes")
async def update_invitation_notes(inv_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Persist a free-form note (e.g. reason for non-attendance) on an invitation."""
    notes = (data.get("notes") or "").strip()
    result = await db.invitations.update_one(
        {"id": inv_id},
        {"$set": {"notes": notes, "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dəvət tapılmadı")
    inv = await db.invitations.find_one({"id": inv_id}, {"_id": 0})
    return inv


@api_router.delete("/invitations/{inv_id}")
async def delete_invitation(inv_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.invitations.delete_one({"id": inv_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dəvət tapılmadı")
    return {"message": "Dəvət silindi"}


@api_router.post("/invitations/{inv_id}/generate-card")
async def generate_company_invitation_card(
    inv_id: str,
    payload: dict | None = None,
    current_user: dict = Depends(get_current_user),
):
    """Render a personalised invitation PNG for a company-based invitation.

    Looks up `db.invitations` (by `id`), resolves the linked company for the
    guest name (owner_name / representative_name override possible via
    payload.guest_name) and phone (payload.phone overrides), reads the event
    details from `db.events`, renders + uploads to Cloudinary, and returns
    `{ url, public_id, whatsapp_link }`. `payload` is optional: pass
    `{ phone: "+994...", guest_name: "Custom Name" }` to override defaults.
    """
    inv = await db.invitations.find_one({"id": inv_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Dəvət tapılmadı")
    event = await db.events.find_one({"id": inv.get("event_id", "")}, {"_id": 0}) or {}
    company = await db.companies.find_one({"id": inv.get("company_id", "")}, {"_id": 0}) or {}

    payload = payload or {}
    guest_name = (
        payload.get("guest_name")
        or company.get("owner_name")
        or company.get("representative_name")
        or inv.get("company_name", "")
        or "Qonağımız"
    )
    phone = (
        payload.get("phone")
        or company.get("owner_phone")
        or company.get("company_phone")
        or company.get("representative_phone")
        or ""
    )

    event_name = event.get("name") or inv.get("event_name") or "Tədbir"
    event_date = _fmt_dd_mm_yyyy(event.get("date") or inv.get("event_date") or "")
    event_time = event.get("time", "")
    event_location = event.get("venue") or event.get("location_link") or ""

    template_body = await _get_invitation_template(event.get("event_type") or "")

    try:
        png_bytes = render_invitation_png(
            guest_name=guest_name,
            event_name=event_name,
            event_date=event_date,
            event_time=event_time,
            event_location=event_location,
            body_template=template_body,
        )
    except Exception as e:
        logger.exception("Invitation render failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Dəvətnamə yaradıla bilmədi: {e}")

    safe_name = "".join(c if c.isalnum() else "_" for c in guest_name)[:32] or "guest"
    filename = f"invitation_{safe_name}_{inv_id[:8]}.png"
    try:
        up = _cl_upload(png_bytes, filename=filename, folder="marsol/invitations", resource_type="image")
    except Exception as e:
        logger.exception("Cloudinary upload failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Cloudinary-ə yüklənmədi: {e}")

    url = up.get("url")
    public_id = up.get("public_id")

    digits = "".join(c for c in (phone or "") if c.isdigit())
    msg_lines = [
        f"Hörmətli {guest_name},",
        f"Sizi \"{event_name}\" tədbirinə dəvət edirik.",
    ]
    if event_date or event_time:
        msg_lines.append(f"Tarix: {event_date} {event_time}".strip())
    if event_location:
        msg_lines.append(f"Ünvan: {event_location}")
    import urllib.parse as _up
    text = _up.quote("\n".join(msg_lines))
    whatsapp_link = f"https://wa.me/{digits}?text={text}" if digits else f"https://wa.me/?text={text}"

    await db.invitations.update_one(
        {"id": inv_id},
        {"$set": {
            "invitation_card_url": url,
            "invitation_card_public_id": public_id,
            "invitation_card_generated_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    return {"url": url, "public_id": public_id, "whatsapp_link": whatsapp_link, "filename": filename}


# ==================== INVITATION TEMPLATES ====================

@api_router.get("/invitation-templates")
async def list_invitation_templates(current_user: dict = Depends(get_current_user)):
    """Return all event-type templates merged with defaults so the UI can show
    every supported event type even if it has not been customised yet."""
    saved = {d["event_type"]: d for d in await db.invitation_templates.find({}, {"_id": 0}).to_list(200)}
    # Pull event_types from the manageable list 'event_types' for completeness
    event_types = list(DEFAULT_INVITATION_TEMPLATES.keys())
    et_list_doc = await db.setting_lists.find_one({"key": "event_types"}, {"_id": 0})
    if et_list_doc and et_list_doc.get("values"):
        for v in et_list_doc["values"]:
            if v not in event_types:
                event_types.append(v)
    out = []
    for et in event_types:
        if et == "default":
            continue
        s = saved.get(et)
        out.append({
            "event_type": et,
            "body": (s or {}).get("body") or DEFAULT_INVITATION_TEMPLATES.get(et) or DEFAULT_INVITATION_TEMPLATES["default"],
            "is_default": s is None,
            "updated_at": (s or {}).get("updated_at"),
            "updated_by": (s or {}).get("updated_by"),
        })
    # Always expose the catch-all 'default' template at the end so admins can edit it
    s = saved.get("default")
    out.append({
        "event_type": "default",
        "body": (s or {}).get("body") or DEFAULT_INVITATION_TEMPLATES["default"],
        "is_default": s is None,
        "updated_at": (s or {}).get("updated_at"),
        "updated_by": (s or {}).get("updated_by"),
    })
    return out


@api_router.put("/invitation-templates/{event_type}")
async def update_invitation_template(event_type: str, data: dict, current_user: dict = Depends(get_current_user)):
    _admin_only(current_user)
    body = (data.get("body") or "").strip()
    if not body:
        raise HTTPException(status_code=400, detail="Şablon mətni boş ola bilməz")
    doc = {
        "event_type": event_type,
        "body": body,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user.get("name", ""),
    }
    await db.invitation_templates.update_one(
        {"event_type": event_type},
        {"$set": doc},
        upsert=True,
    )
    return doc


@api_router.delete("/invitation-templates/{event_type}")
async def reset_invitation_template(event_type: str, current_user: dict = Depends(get_current_user)):
    """Reset to default (delete the customised override)."""
    _admin_only(current_user)
    await db.invitation_templates.delete_one({"event_type": event_type})
    return {"event_type": event_type, "body": DEFAULT_INVITATION_TEMPLATES.get(event_type) or DEFAULT_INVITATION_TEMPLATES["default"], "is_default": True}


# ==================== OBLIGATIONS (ÖHDƏLİKLƏR) ====================

async def _get_company_obligation(company: dict, year: Optional[int] = None) -> dict:
    """Compute obligation stats. If year given, use period covering that year (current or from history) and only count invitations whose event_date is in that year."""
    quotas = await get_package_quotas()
    company_id = company.get("id", "")

    # Find the applicable period
    current = {
        "package": company.get("package", ""),
        "contract_start": company.get("contract_start_date", ""),
        "contract_end": company.get("contract_end_date", ""),
    }
    history = company.get("membership_history", []) or []
    
    def period_covers_year(period, y):
        try:
            cs = period.get("contract_start") or period.get("contract_start_date") or ""
            ce = period.get("contract_end") or period.get("contract_end_date") or ""
            cs_y = int(cs[:4]) if cs else 0
            ce_y = int(ce[:4]) if ce else 9999
            return cs_y <= y <= ce_y
        except (ValueError, TypeError):
            return False

    period = current
    if year is not None:
        # Try current first; else search history
        if not period_covers_year(current, year):
            for h in history:
                if period_covers_year(h, year):
                    period = {
                        "package": h.get("package", ""),
                        "contract_start": h.get("contract_start", ""),
                        "contract_end": h.get("contract_end", ""),
                    }
                    break
    
    package = period["package"]
    total_quota = quotas.get(package, 0)
    start_date = period["contract_start"]
    end_date = period["contract_end"]
    
    now = datetime.now(timezone.utc)
    days_remaining = 365
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            days_remaining = max((end_dt - now.replace(tzinfo=None)).days, 0)
        except (ValueError, TypeError):
            pass
    
    # Build invitation filter — by year if given, else by current period dates
    inv_filter = {"company_id": company_id, "obligation_deducted": True}
    inv_all_filter = {"company_id": company_id}
    if year is not None:
        # Match invitations whose event_date is in that calendar year
        date_re = f"^{year}-"
        inv_filter["event_date"] = {"$regex": date_re}
        inv_all_filter["event_date"] = {"$regex": date_re}
    
    used_quota = await db.invitations.count_documents(inv_filter)
    remaining = max(total_quota - used_quota, 0)
    priority_score = 0
    if days_remaining > 0 and remaining > 0:
        priority_score = remaining * (365 / max(days_remaining, 1))
    elif days_remaining == 0 and remaining > 0:
        priority_score = remaining * 1000
    total_invited = await db.invitations.count_documents(inv_all_filter)
    total_attended = await db.invitations.count_documents({**inv_all_filter, "participation_status": "Qatılır"})
    total_declined = await db.invitations.count_documents({**inv_all_filter, "participation_status": "Qatılmır"})
    total_no_answer = await db.invitations.count_documents({**inv_all_filter, "call_status": "Cavab vermədi"})
    return {
        "company_id": company_id,
        "display_id": company.get("display_id", ""),
        "company_name": company.get("brand_name", ""),
        "owner_name": company.get("owner_name", ""),
        "owner_phone": company.get("owner_phone", ""),
        "company_phone": company.get("company_phone", ""),
        "package": package,
        "total_quota": total_quota,
        "used_quota": used_quota,
        "remaining_quota": remaining,
        "contract_start_date": start_date,
        "contract_end_date": end_date,
        "days_remaining": days_remaining,
        "priority_score": round(priority_score, 2),
        "total_invited": total_invited,
        "total_attended": total_attended,
        "total_declined": total_declined,
        "total_no_answer": total_no_answer,
        "status": company.get("status", "Aktiv"),
    }

def _company_covers_year(company: dict, year: int) -> bool:
    """True if any membership period (current or in history) covers the given calendar year."""
    def cov(cs, ce):
        try:
            cs_y = int(cs[:4]) if cs else 0
            ce_y = int(ce[:4]) if ce else 9999
            return cs_y <= year <= ce_y
        except (ValueError, TypeError):
            return False
    if cov(company.get("contract_start_date", ""), company.get("contract_end_date", "")):
        return True
    for h in (company.get("membership_history") or []):
        if cov(h.get("contract_start", ""), h.get("contract_end", "")):
            return True
    return False

async def _bulk_invitation_stats(company_ids: List[str], year: Optional[int] = None) -> Dict[str, Dict[str, int]]:
    """Fetch invitation counts for many companies in ONE aggregation. Returns
    {company_id: {used_quota, total_invited, total_attended, total_declined, total_no_answer}}.
    """
    if not company_ids:
        return {}
    match: Dict[str, Any] = {"company_id": {"$in": company_ids}}
    if year is not None:
        match["event_date"] = {"$regex": f"^{year}-"}
    pipeline = [
        {"$match": match},
        {"$group": {
            "_id": "$company_id",
            "total_invited": {"$sum": 1},
            "used_quota": {"$sum": {"$cond": [{"$eq": ["$obligation_deducted", True]}, 1, 0]}},
            "total_attended": {"$sum": {"$cond": [{"$eq": ["$participation_status", "Qatılır"]}, 1, 0]}},
            "total_declined": {"$sum": {"$cond": [{"$eq": ["$participation_status", "Qatılmır"]}, 1, 0]}},
            "total_no_answer": {"$sum": {"$cond": [{"$eq": ["$call_status", "Cavab vermədi"]}, 1, 0]}},
        }},
    ]
    rows = await db.invitations.aggregate(pipeline).to_list(5000)
    return {r["_id"]: r for r in rows}


def _build_company_obligation(company: dict, quotas: Dict[str, int], stats: Dict[str, int], year: Optional[int] = None) -> dict:
    """Pure helper — combines pre-fetched quota + stats into the obligation dict."""
    company_id = company.get("id", "")
    current = {
        "package": company.get("package", ""),
        "contract_start": company.get("contract_start_date", ""),
        "contract_end": company.get("contract_end_date", ""),
    }
    history = company.get("membership_history", []) or []

    def period_covers_year(period, y):
        try:
            cs = period.get("contract_start") or period.get("contract_start_date") or ""
            ce = period.get("contract_end") or period.get("contract_end_date") or ""
            cs_y = int(cs[:4]) if cs else 0
            ce_y = int(ce[:4]) if ce else 9999
            return cs_y <= y <= ce_y
        except (ValueError, TypeError):
            return False

    period = current
    if year is not None and not period_covers_year(current, year):
        for h in history:
            if period_covers_year(h, year):
                period = {
                    "package": h.get("package", ""),
                    "contract_start": h.get("contract_start", ""),
                    "contract_end": h.get("contract_end", ""),
                }
                break

    package = period["package"]
    total_quota = quotas.get(package, 0)
    start_date = period["contract_start"]
    end_date = period["contract_end"]

    now = datetime.now(timezone.utc)
    days_remaining = 365
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            days_remaining = max((end_dt - now.replace(tzinfo=None)).days, 0)
        except (ValueError, TypeError):
            pass

    used_quota = int(stats.get("used_quota", 0) or 0)
    total_invited = int(stats.get("total_invited", 0) or 0)
    total_attended = int(stats.get("total_attended", 0) or 0)
    total_declined = int(stats.get("total_declined", 0) or 0)
    total_no_answer = int(stats.get("total_no_answer", 0) or 0)

    # Manual overrides — if the company has an obligation_overrides[<year_key>]
    # sub-document (set via the Excel import endpoint), prefer its values over
    # the aggregated invitations stats. Lets users migrate historical data
    # without backfilling 1000s of invitation rows.
    overrides_doc = company.get("obligation_overrides") or {}
    year_key = str(year) if year is not None else "all"
    ov = overrides_doc.get(year_key) or {}
    if "total_quota" in ov:
        total_quota = int(ov["total_quota"] or 0)
    if "used_quota" in ov:
        used_quota = int(ov["used_quota"] or 0)
    if "total_invited" in ov:
        total_invited = int(ov["total_invited"] or 0)
    if "total_attended" in ov:
        total_attended = int(ov["total_attended"] or 0)
    if "total_declined" in ov:
        total_declined = int(ov["total_declined"] or 0)
    if "total_no_answer" in ov:
        total_no_answer = int(ov["total_no_answer"] or 0)

    remaining = max(total_quota - used_quota, 0)
    priority_score = 0
    if days_remaining > 0 and remaining > 0:
        priority_score = remaining * (365 / max(days_remaining, 1))
    elif days_remaining == 0 and remaining > 0:
        priority_score = remaining * 1000

    return {
        "company_id": company_id,
        "display_id": company.get("display_id", ""),
        "company_name": company.get("brand_name", ""),
        "owner_name": company.get("owner_name", ""),
        "owner_phone": company.get("owner_phone", ""),
        "company_phone": company.get("company_phone", ""),
        "package": package,
        "total_quota": total_quota,
        "used_quota": used_quota,
        "remaining_quota": remaining,
        "contract_start_date": start_date,
        "contract_end_date": end_date,
        "days_remaining": days_remaining,
        "priority_score": round(priority_score, 2),
        "total_invited": total_invited,
        "total_attended": total_attended,
        "total_declined": total_declined,
        "total_no_answer": total_no_answer,
        "status": company.get("status", "Aktiv"),
    }


@api_router.get("/obligations/dashboard")
async def get_obligations_dashboard(year: Optional[int] = None, current_user: dict = Depends(get_current_user)):
    companies = await db.companies.find({"status": "Aktiv"}, {"_id": 0}).to_list(2000)
    if year is not None:
        companies = [c for c in companies if _company_covers_year(c, year)]
    # Bulk: fetch quotas once + 1 aggregation for invitation stats across ALL companies.
    quotas = await get_package_quotas()
    stats_map = await _bulk_invitation_stats([c.get("id", "") for c in companies], year=year)
    obligations = [
        _build_company_obligation(c, quotas, stats_map.get(c.get("id", ""), {}), year=year)
        for c in companies
    ]
    obligations.sort(key=lambda x: x["priority_score"], reverse=True)
    total = len(obligations)
    not_invited = sum(1 for o in obligations if o["total_invited"] == 0)
    under_invited = sum(1 for o in obligations if 0 < o["total_invited"] < o["total_quota"] and o["remaining_quota"] > 0)
    fully_served = sum(1 for o in obligations if o["remaining_quota"] == 0)
    urgent = sum(1 for o in obligations if o["priority_score"] > 50)
    return {
        "obligations": obligations,
        "stats": {
            "total": total,
            "not_invited": not_invited,
            "under_invited": under_invited,
            "fully_served": fully_served,
            "urgent": urgent,
        }
    }

@api_router.get("/obligations/company/{company_id}")
async def get_company_obligation(company_id: str, year: Optional[int] = None, current_user: dict = Depends(get_current_user)):
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Şirkət tapılmadı")
    # Reuse the bulk pipeline so override values written by the Excel import
    # show up here too (the legacy helper bypassed obligation_overrides).
    quotas = await get_package_quotas()
    stats_map = await _bulk_invitation_stats([company_id], year=year)
    obl = _build_company_obligation(company, quotas, stats_map.get(company_id, {}), year=year)
    inv_filter = {"company_id": company_id}
    if year is not None:
        inv_filter["event_date"] = {"$regex": f"^{year}-"}
    invitations = await db.invitations.find(inv_filter, {"_id": 0}).sort("event_date", -1).to_list(500)
    type_breakdown = {}
    for inv in invitations:
        et = inv.get("event_type", "Digər")
        if et not in type_breakdown:
            type_breakdown[et] = {"invited": 0, "attended": 0, "declined": 0, "no_answer": 0}
        type_breakdown[et]["invited"] += 1
        if inv.get("participation_status") == "Qatılır":
            type_breakdown[et]["attended"] += 1
        elif inv.get("participation_status") == "Qatılmır":
            type_breakdown[et]["declined"] += 1
        if inv.get("call_status") == "Cavab vermədi":
            type_breakdown[et]["no_answer"] += 1
    obl["invitations"] = invitations
    obl["type_breakdown"] = type_breakdown
    return obl


def _parse_az_date(value) -> str:
    """Parse a date-ish value (DD/MM/YYYY, DD.MM.YYYY, ISO, datetime, Excel
    serial via openpyxl) into ISO 'YYYY-MM-DD'. Returns '' for empty/invalid."""
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return ""
    # Already ISO?
    iso_m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if iso_m:
        y, mo, da = iso_m.groups()
        return f"{y}-{int(mo):02d}-{int(da):02d}"
    m = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})$", s)
    if m:
        dd, mo, yy = m.groups()
        if len(yy) == 2:
            yy = "20" + yy
        try:
            return f"{int(yy):04d}-{int(mo):02d}-{int(dd):02d}"
        except (ValueError, TypeError):
            return ""
    return ""


@api_router.post("/obligations/import-excel")
async def import_obligations_excel(payload: dict, current_user: dict = Depends(check_permission("obligations", "write"))):
    """Bulk-update companies based on the same xlsx schema we export from
    /obligations: rows is a list of dicts with keys 'Şirkət', 'Ad', 'Soyad',
    'Paket', 'Müqavilə başlama', 'Müqavilə bitmə'. Matches by brand_name."""
    import unicodedata
    rows = payload.get("rows") or []
    if not isinstance(rows, list) or not rows:
        raise HTTPException(status_code=400, detail="rows boş ola bilməz")

    def _norm_key(s: str) -> str:
        """NFC-normalise + casefold + strip — so that 'Şirkət', ' şirkət ' and
        a decomposed-unicode 'Şirkət' all collapse to the same key.
        Also strips combining marks (e.g. the dot-above produced when 'İ'
        casefolds to 'i̇') so Azerbaijani capital İ matches latin 'i'."""
        if s is None:
            return ""
        x = unicodedata.normalize("NFC", str(s)).strip().casefold()
        # Decompose, drop combining marks, recompose
        x = unicodedata.normalize("NFD", x)
        x = "".join(ch for ch in x if not unicodedata.combining(ch))
        return unicodedata.normalize("NFC", x)

    # Synonym map: every accepted alias for each canonical field. Lookups go
    # via a row's normalised keys so column-header drift (extra spaces,
    # accent decomposition, case differences) doesn't silently drop data.
    SYNONYMS = {
        "company_id": ["id", "şirkət id", "sirket id", "company_id"],
        "brand": ["şirkət", "sirket", "şirket", "company", "müəssisə", "muessise"],
        "first": ["ad", "first_name", "first name", "owner_first", "owner first name"],
        "last":  ["soyad", "last_name", "last name", "owner_last", "owner last name"],
        "owner_full": ["sahibkar", "sahibkar adı", "owner", "owner_name", "ad soyad"],
        "package": ["paket", "package"],
        "start": ["müqavilə başlama", "muqavile baslama", "contract start", "contract_start", "başlama", "baslama"],
        "end":   ["müqavilə bitmə", "muqavile bitme", "contract end", "contract_end", "bitmə", "bitme"],
        # Aggregated obligation metrics — stored as a per-year override on the company doc
        "total_quota": ["ümumi kvota", "umumi kvota", "total_quota", "kvota"],
        "used_quota": ["istifadə olunan", "istifade olunan", "istifadə olunan kvota", "used_quota"],
        "total_invited": ["cəmi dəvət", "cemi devet", "total_invited", "dəvət sayı"],
        "total_attended": ["qatıldı", "qatildi", "total_attended"],
        "total_declined": ["rədd etdi", "redd etdi", "total_declined"],
        "total_no_answer": ["cavab vermədi", "cavab vermedi", "total_no_answer"],
    }
    SYN_NORM = {k: [_norm_key(a) for a in v] for k, v in SYNONYMS.items()}

    def _pick(row_dict_norm: Dict[str, Any], field: str):
        for alias in SYN_NORM[field]:
            v = row_dict_norm.get(alias)
            if v is not None and v != "":
                return v
        return ""

    updated = 0
    skipped = 0
    errors: List[Dict[str, Any]] = []

    # Optional top-level year scope — when the user exports filtered by year
    # they should be able to import metrics that apply ONLY to that year.
    # Omit (or pass null) for "all years".
    raw_year = payload.get("year")
    year_key = str(int(raw_year)) if raw_year not in (None, "", 0) else "all"

    def _to_int(v) -> Optional[int]:
        if v is None or v == "":
            return None
        try:
            return int(float(str(v).replace(",", ".").strip()))
        except (ValueError, TypeError):
            return None

    for i, row in enumerate(rows, start=1):
        try:
            if not isinstance(row, dict):
                skipped += 1
                errors.append({"row": i, "reason": "Sətir obyekt deyil"})
                continue
            # Re-key the row using normalised keys so we're resilient to
            # accent/whitespace/case drift in column headers.
            row_norm = {_norm_key(k): v for k, v in row.items()}
            # Prefer matching by display_id (stable, never changes) — fall back to brand_name.
            display_id = str(_pick(row_norm, "company_id") or "").strip()
            brand = str(_pick(row_norm, "brand") or "").strip()
            company = None
            if display_id:
                company = await db.companies.find_one({"display_id": display_id}, {"_id": 0})
            if not company and brand:
                company = await db.companies.find_one({"brand_name": brand}, {"_id": 0})
            if not company:
                skipped += 1
                errors.append({"row": i, "reason": f"Şirkət tapılmadı: {display_id or brand or '(boş)'}"})
                continue
            update: Dict[str, Any] = {}
            first = str(_pick(row_norm, "first") or "").strip()
            last = str(_pick(row_norm, "last") or "").strip()
            if first or last:
                update["owner_name"] = (first + " " + last).strip()
            else:
                # Backwards compat — old xlsx exports used a single 'Sahibkar' column.
                full = str(_pick(row_norm, "owner_full") or "").strip()
                if full:
                    update["owner_name"] = full
            pkg = str(_pick(row_norm, "package") or "").strip()
            if pkg:
                update["package"] = pkg
            start_iso = _parse_az_date(_pick(row_norm, "start"))
            if start_iso:
                update["contract_start_date"] = start_iso
            end_iso = _parse_az_date(_pick(row_norm, "end"))
            if end_iso:
                update["contract_end_date"] = end_iso

            # ---- Obligation metric overrides ----
            # used_quota / total_invited / total_attended / total_declined /
            # total_no_answer / total_quota live inside an obligation_overrides
            # sub-doc keyed by year (or 'all'). _build_company_obligation reads
            # these and prefers them over the invitations aggregation, so
            # imported numbers appear immediately on the dashboard.
            metric_keys = ("total_quota", "used_quota", "total_invited",
                           "total_attended", "total_declined", "total_no_answer")
            new_override: Dict[str, int] = {}
            for mk in metric_keys:
                v = _to_int(_pick(row_norm, mk))
                if v is not None:
                    new_override[mk] = max(0, v)
            if new_override:
                existing = (company.get("obligation_overrides") or {}).get(year_key, {})
                merged = {**existing, **new_override}
                update[f"obligation_overrides.{year_key}"] = merged

            if not update:
                skipped += 1
                continue
            update["updated_at"] = datetime.now(timezone.utc).isoformat()
            await db.companies.update_one({"id": company["id"]}, {"$set": update})
            updated += 1
        except Exception as ex:  # noqa: BLE001
            skipped += 1
            errors.append({"row": i, "reason": str(ex)})

    return {
        "updated": updated,
        "skipped": skipped,
        "total": len(rows),
        "errors": errors[:50],
        # diagnostic — first row's actual keys help main agent debug column-name drift
        "sample_keys": list(rows[0].keys()) if rows and isinstance(rows[0], dict) else [],
    }

# ==================== AUTO-SUGGEST (AVTO-TƏKLİF) ====================

@api_router.post("/events/{event_id}/auto-suggest")
async def auto_suggest_companies(event_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Fəaliyyət tapılmadı")
    count = data.get("count", event.get("participant_limit", 20))
    exclude_ids = data.get("exclude_ids", [])
    already_invited = await db.invitations.find({"event_id": event_id}, {"_id": 0, "company_id": 1}).to_list(5000)
    already_invited_ids = {inv["company_id"] for inv in already_invited}
    companies = await db.companies.find({"status": "Aktiv"}, {"_id": 0}).to_list(2000)
    candidates = []
    for c in companies:
        cid = c.get("id", "")
        if cid in already_invited_ids or cid in exclude_ids:
            continue
        obl = await _get_company_obligation(c)
        if obl["remaining_quota"] <= 0:
            continue
        obl["sector"] = c.get("sector", "")
        obl["sub_sector"] = c.get("sub_sector", "")
        candidates.append(obl)
    candidates.sort(key=lambda x: x["priority_score"], reverse=True)
    # Sector conflict filter: only 1 company per sub_sector (or sector if no sub_sector)
    selected = []
    used_sub_sectors = set()
    for c in candidates:
        sub = c.get("sub_sector") or ""
        sector = c.get("sector") or ""
        conflict_key = sub.strip().lower() if sub.strip() else sector.strip().lower()
        if conflict_key and conflict_key in used_sub_sectors:
            continue
        selected.append(c)
        if conflict_key:
            used_sub_sectors.add(conflict_key)
        if len(selected) >= count:
            break
    return {"suggestions": selected, "total_candidates": len(candidates)}

@api_router.post("/events/{event_id}/check-sector-conflict")
async def check_sector_conflict(event_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    company_id = data.get("company_id", "")
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        return {"conflict": False}
    company_sub = (company.get("sub_sector") or "").strip().lower()
    company_sector = (company.get("sector") or "").strip().lower()
    check_key = company_sub if company_sub else company_sector
    if not check_key:
        return {"conflict": False}
    invited = await db.invitations.find({"event_id": event_id}, {"_id": 0, "company_id": 1}).to_list(500)
    for inv in invited:
        inv_company = await db.companies.find_one({"id": inv["company_id"]}, {"_id": 0, "sector": 1, "sub_sector": 1, "brand_name": 1})
        if not inv_company:
            continue
        inv_sub = (inv_company.get("sub_sector") or "").strip().lower()
        inv_sector = (inv_company.get("sector") or "").strip().lower()
        inv_key = inv_sub if inv_sub else inv_sector
        if inv_key == check_key:
            return {
                "conflict": True,
                "conflicting_company": inv_company.get("brand_name", ""),
                "conflict_type": "alt sektor" if company_sub else "sektor",
                "conflict_value": company.get("sub_sector") or company.get("sector", "")
            }
    return {"conflict": False}

# ==================== NOTIFICATIONS (BİLDİRİŞLƏR) ====================

@api_router.get("/notifications")
async def get_notifications(current_user: dict = Depends(get_current_user)):
    notifications = []
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    cfg = await get_notification_settings()

    # 0. Stored direct notifications (task_assigned, message, generic). Show ONLY
    # those addressed to the current user (by name) over the last 30 days.
    me_name = current_user.get("name", "")
    if me_name:
        cutoff = (now - timedelta(days=30)).isoformat()
        stored = await db.notifications.find(
            {
                "recipient_name": me_name,
                "created_at": {"$gte": cutoff},
                "type": {"$in": ["task_assigned", "task_updated", "task_comment", "message", "note_shared"]},
            },
            {"_id": 0},
        ).sort("created_at", -1).to_list(200)
        for n in stored:
            notifications.append({
                "id": n.get("id", ""),
                "type": n.get("type", "info"),
                "severity": "medium",
                "title": n.get("title", ""),
                "message": n.get("body", ""),
                "task_id": n.get("task_id", ""),
                "conversation_id": n.get("conversation_id", ""),
                "date": (n.get("created_at") or "")[:10] or today,
            })

    # Scope helper for computed notifications: non-admin users with scope='own'
    # should only see notifications tied to records they curate / are assigned to.
    is_admin = (current_user.get("role") or "").lower() == "admin"
    scopes = await get_user_scopes(current_user) if not is_admin else {}

    def _user_owns_company(c: dict) -> bool:
        if is_admin:
            return True
        sc = scopes.get("companies", "own")
        if sc == "all":
            return True
        return c.get("curator") == me_name or c.get("created_by") == me_name

    def _user_owns_meeting(m: dict) -> bool:
        if is_admin:
            return True
        sc = scopes.get("meetings", "own")
        if sc == "all":
            return True
        if m.get("created_by") == me_name or m.get("meeting_setter") == me_name:
            return True
        if m.get("employee") == me_name:
            return True
        if me_name in (m.get("participant_names") or []):
            return True
        return False
    
    # 1. Overdue debts (borclu şirkətlər)
    debtors = await db.companies.find({"debt_amount": {"$gt": 0}}, {"_id": 0}).to_list(500)
    for c in debtors:
        if not _user_owns_company(c):
            continue
        days_overdue = 0
        if c.get("payment_due_date"):
            try:
                due = datetime.strptime(c["payment_due_date"], "%Y-%m-%d")
                diff = (now.replace(tzinfo=None) - due).days
                if diff > 0:
                    days_overdue = diff
            except (ValueError, TypeError):
                pass
        if days_overdue > 0:
            notifications.append({
                "id": f"debt-{c['id']}",
                "type": "debt_overdue",
                "severity": "high" if days_overdue > cfg["debt_overdue_high_days"] else "medium",
                "title": f"Gecikmiş ödəniş: {c['brand_name']}",
                "message": f"{c['debt_amount']:,.0f} AZN borc — {days_overdue} gün gecikib",
                "company_id": c["id"],
                "date": today
            })
        elif (c.get("debt_amount") or 0) > 0:
            notifications.append({
                "id": f"debt-pending-{c['id']}",
                "type": "debt_pending",
                "severity": "low",
                "title": f"Ödənilməmiş borc: {c['brand_name']}",
                "message": f"{c['debt_amount']:,.0f} AZN borc qalıb",
                "company_id": c["id"],
                "date": today
            })

    # 2. Contract expiring soon (müqavilə xitamı yaxınlaşan)
    all_companies = await db.companies.find({"contract_end_date": {"$ne": ""}}, {"_id": 0}).to_list(500)
    for c in all_companies:
        if not _user_owns_company(c):
            continue
        end_str = c.get("contract_end_date")
        if not end_str:
            continue
        try:
            end_date = datetime.strptime(end_str, "%Y-%m-%d")
            diff = (end_date - now.replace(tzinfo=None)).days
            if diff < 0:
                notifications.append({
                    "id": f"contract-expired-{c['id']}",
                    "type": "contract_expired",
                    "severity": "high",
                    "title": f"Müqavilə bitib: {c.get('brand_name', '')}",
                    "message": f"Müqavilə {abs(diff)} gün əvvəl bitib ({end_str})",
                    "company_id": c["id"],
                    "date": today
                })
            elif diff <= cfg["contract_expiry_days"]:
                notifications.append({
                    "id": f"contract-expiring-{c['id']}",
                    "type": "contract_expiring",
                    "severity": "medium",
                    "title": f"Müqavilə bitir: {c.get('brand_name', '')}",
                    "message": f"{diff} gün sonra bitəcək ({end_str})",
                    "company_id": c["id"],
                    "date": today
                })
        except (ValueError, TypeError):
            pass

    # 3.5 Pending form submissions (üzvlük formu təsdiq gözləyir)
    pending_form_companies = await db.companies.find(
        {"pending_form_data": {"$exists": True, "$ne": {}}},
        {"_id": 0}
    ).to_list(500)
    for c in pending_form_companies:
        if not (c.get("pending_form_data") or {}):
            continue
        if not _user_owns_company(c):
            continue
        notifications.append({
            "id": f"form-pending-{c['id']}",
            "type": "form_submission",
            "severity": "medium",
            "title": f"Forum dəyişikliyi: {c.get('brand_name', '')}",
            "message": f"Üzvlük forumu doldurulub, təsdiq gözləyir.",
            "company_id": c["id"],
            "date": (c.get("pending_form_submitted_at") or "")[:10] or today
        })

    # 4. Meeting reminders (görüş xatırlatmaları)
    meeting_reminders = await db.notifications.find({"type": "reminder"}, {"_id": 0}).to_list(500)
    # Build a meeting lookup map to test ownership per reminder
    rem_meeting_ids = [r.get("meeting_id") for r in meeting_reminders if r.get("meeting_id")]
    if rem_meeting_ids and not is_admin:
        rem_meetings_map = {m["id"]: m for m in await db.meetings.find({"id": {"$in": rem_meeting_ids}}, {"_id": 0}).to_list(500)}
    else:
        rem_meetings_map = {}
    for r in meeting_reminders:
        if not is_admin and scopes.get("meetings", "all") != "all":
            mid = r.get("meeting_id")
            m = rem_meetings_map.get(mid) if mid else None
            if m and not _user_owns_meeting(m):
                continue
            # If there's no meeting record (orphan reminder), prefer privacy → skip for non-admins
            if not m:
                continue
        severity = "low"
        rem_date = r.get("reminder_date", "")
        if rem_date:
            try:
                rd = datetime.strptime(rem_date, "%Y-%m-%d")
                diff = (rd - now.replace(tzinfo=None)).days
                if diff < 0:
                    severity = "high"
                elif diff <= cfg["meeting_reminder_high_days"]:
                    severity = "high"
                elif diff <= cfg["meeting_reminder_medium_days"]:
                    severity = "medium"
            except (ValueError, TypeError):
                pass
        notifications.append({
            "id": r.get("id", ""),
            "type": "reminder",
            "severity": severity,
            "title": r.get("title", "Görüş xatırlatması"),
            "message": r.get("message", ""),
            "meeting_id": r.get("meeting_id", ""),
            "reminder_date": rem_date,
            "reminder_time": r.get("reminder_time", ""),
            "is_read": r.get("is_read", False),
            "date": rem_date or today
        })

    # 4. Membership expiry warnings (üzvlük bitmə xəbərdarlığı)
    warning_days = cfg["membership_warning_days"]
    expiring = await db.companies.find({"contract_end_date": {"$ne": ""}}, {"_id": 0, "id": 1, "brand_name": 1, "contract_end_date": 1, "curator": 1}).to_list(500)
    for c in expiring:
        if not _user_owns_company(c):
            continue
        end_str = c.get("contract_end_date")
        if not end_str:
            continue
        try:
            end_dt = datetime.strptime(end_str, "%Y-%m-%d")
            diff = (end_dt - now.replace(tzinfo=None)).days
            if 0 < diff <= warning_days:
                notifications.append({
                    "id": f"expiry-{c['id']}",
                    "type": "membership_expiry",
                    "severity": "high" if diff <= 3 else "medium",
                    "title": f"Üzvlük bitir: {c.get('brand_name', '')}",
                    "message": f"{diff} gün sonra bitəcək ({end_str})",
                    "company_id": c["id"],
                    "date": today
                })
            elif diff <= 0:
                notifications.append({
                    "id": f"expired-{c['id']}",
                    "type": "membership_expired",
                    "severity": "high",
                    "title": f"Üzvlük bitib: {c.get('brand_name', '')}",
                    "message": f"{abs(diff)} gün əvvəl bitib ({end_str})",
                    "company_id": c["id"],
                    "date": today
                })
        except (ValueError, TypeError):
            pass

    # 5. Birthday reminders (configured advance days + on the day) — for ALL contacts:
    # company owners, representatives, contact_persons, employees, and
    # contact_lists entries. We compare month+day so the reminder fires every
    # year regardless of birth year.
    today_md = now.strftime("%m-%d")
    advance_days = max(int(cfg.get("birthday_advance_days") or 0), 0)
    advance_md_set = set()
    for d in range(1, advance_days + 1):
        advance_md_set.add((now + timedelta(days=d)).strftime("%m-%d"))

    def _md(date_str):
        if not date_str:
            return None
        try:
            return datetime.strptime(date_str[:10], "%Y-%m-%d").strftime("%m-%d")
        except (ValueError, TypeError):
            return None

    def _bday_notif(person_name, role, source_type, source_id, bday_str, comp_id=None, company_name=None):
        md = _md(bday_str)
        if md is None:
            return None
        if md == today_md:
            when = "Bu gün"
            sev = "high"
            anchor_year = now.year
        elif md in advance_md_set:
            # Calculate which day-ahead this is
            try:
                days_ahead = next((d for d in range(1, advance_days + 1) if (now + timedelta(days=d)).strftime("%m-%d") == md), 1)
            except (ValueError, TypeError):
                days_ahead = 1
            when = "Sabah" if days_ahead == 1 else f"{days_ahead} gün sonra"
            sev = "medium"
            anchor_year = (now + timedelta(days=days_ahead)).year
        else:
            return None
        msg = f"{role} — {bday_str}"
        if company_name:
            msg = f"{role} ({company_name}) — {bday_str}"
        return {
            "id": f"bday-{anchor_year}-{source_type}-{source_id}-{md}",
            "type": "birthday",
            "severity": sev,
            "title": f"🎂 {when} ad günü: {person_name}",
            "message": msg,
            "company_id": comp_id,
            "company_name": company_name,
            "date": today,
        }

    # 5a. Owners + representatives + contact_person on companies
    bday_companies = await db.companies.find(
        {},
        {
            "_id": 0, "id": 1, "brand_name": 1,
            "owner_name": 1, "owner_birth_date": 1,
            "owners": 1,
            "representative_name": 1, "representative_birth_date": 1,
            "contact_first_name": 1, "contact_last_name": 1, "contact_birth_date": 1,
        },
    ).to_list(2000)
    for c in bday_companies:
        if not _user_owns_company(c):
            continue
        cid = c["id"]
        cname = c.get("brand_name") or ""
        # Single-owner shortcut
        if c.get("owner_birth_date"):
            n = _bday_notif(c.get("owner_name", ""), "Sahibkar", "owner", cid, c["owner_birth_date"], cid, cname)
            if n: notifications.append(n)
        # Multiple-owners array
        for idx, o in enumerate(c.get("owners") or []):
            if o.get("birth_date"):
                pname = f"{o.get('first_name','')} {o.get('last_name','')}".strip() or o.get('name','')
                n = _bday_notif(pname, "Sahibkar", f"owner-{cid}", str(idx), o["birth_date"], cid, cname)
                if n: notifications.append(n)
        if c.get("representative_birth_date"):
            n = _bday_notif(c.get("representative_name", ""), "Nümayəndə", "rep", cid, c["representative_birth_date"], cid, cname)
            if n: notifications.append(n)
        if c.get("contact_birth_date"):
            full = f"{c.get('contact_first_name','')} {c.get('contact_last_name','')}".strip()
            n = _bday_notif(full, "Əlaqəli şəxs", "contact", cid, c["contact_birth_date"], cid, cname)
            if n: notifications.append(n)

    # 5b. Employees
    employees = await db.employees.find(
        {"birth_date": {"$exists": True, "$ne": ""}},
        {"_id": 0, "id": 1, "first_name": 1, "last_name": 1, "birth_date": 1},
    ).to_list(1000)
    for e in employees:
        full = f"{e.get('first_name','')} {e.get('last_name','')}".strip()
        n = _bday_notif(full, "Əməkdaş", "emp", e["id"], e["birth_date"], None, "Marsol Group")
        if n: notifications.append(n)

    # 5c. Contact list entries
    contacts = await db.contacts.find(
        {"birthday": {"$exists": True, "$ne": ""}},
        {"_id": 0, "id": 1, "name": 1, "surname": 1, "company": 1, "birthday": 1},
    ).to_list(2000)
    for ct in contacts:
        full = f"{ct.get('name','')} {ct.get('surname','')}".strip()
        comp_label = ct.get("company") or ""
        n = _bday_notif(full, "Kontakt", "ctc", ct["id"], ct["birthday"], None, comp_label)
        if n: notifications.append(n)

    # Sort by severity
    severity_order = {"high": 0, "medium": 1, "low": 2}
    notifications.sort(key=lambda x: severity_order.get(x["severity"], 3))

    # Mark notifications that the user already dismissed/read.
    user_id = current_user.get("id") or current_user.get("email", "")
    read_ids = set()
    try:
        rows = await db.notification_reads.find({"user_id": user_id}, {"_id": 0, "notification_id": 1}).to_list(5000)
        read_ids = {r["notification_id"] for r in rows if r.get("notification_id")}
    except Exception:
        pass
    for n in notifications:
        n["read"] = n["id"] in read_ids
    unread_count = sum(1 for n in notifications if not n["read"])
    high_unread = sum(1 for n in notifications if not n["read"] and n["severity"] == "high")

    return {
        "notifications": notifications,
        "count": unread_count,
        "total_count": len(notifications),
        "high_count": high_unread,
    }


@api_router.post("/notifications/mark-read")
async def mark_notifications_read(payload: dict, current_user: dict = Depends(get_current_user)):
    """Mark one (id) or several (ids) computed notifications as read for the current user.
    The notifications themselves are computed on the fly, so we persist only the
    read receipts in `notification_reads`."""
    ids = payload.get("ids")
    if not ids and payload.get("id"):
        ids = [payload["id"]]
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="id və ya ids tələb olunur")
    user_id = current_user.get("id") or current_user.get("email", "")
    now_iso = datetime.now(timezone.utc).isoformat()
    ops = [
        {"user_id": user_id, "notification_id": nid, "read_at": now_iso}
        for nid in ids
    ]
    for op in ops:
        await db.notification_reads.update_one(
            {"user_id": op["user_id"], "notification_id": op["notification_id"]},
            {"$set": op},
            upsert=True,
        )
    return {"marked": len(ops)}


@api_router.post("/notifications/mark-all-read")
async def mark_all_notifications_read(current_user: dict = Depends(get_current_user)):
    """Mark all currently visible notifications as read."""
    # Re-compute notification ids; cheaper than caching since /notifications is fast.
    fake_payload = await get_notifications(current_user=current_user)
    ids = [n["id"] for n in fake_payload.get("notifications", [])]
    if not ids:
        return {"marked": 0}
    return await mark_notifications_read({"ids": ids}, current_user=current_user)


@api_router.post("/notifications/admin-reset-all")
async def admin_reset_all_notifications(current_user: dict = Depends(get_current_user)):
    """ADMIN-ONLY: clear notifications for ALL users system-wide.
    Deletes all stored notifications, mark-read tracking, and dispatched-email tracking."""
    _admin_only(current_user)
    n1 = await db.notifications.delete_many({})
    n2 = await db.notification_reads.delete_many({})
    n3 = await db.notification_emails.delete_many({}) if "notification_emails" in await db.list_collection_names() else None
    return {
        "deleted_notifications": getattr(n1, "deleted_count", 0),
        "deleted_reads": getattr(n2, "deleted_count", 0),
        "deleted_dispatched_emails": getattr(n3, "deleted_count", 0) if n3 else 0,
    }



@api_router.post("/notifications/dispatch-emails")
async def dispatch_notification_emails(current_user: dict = Depends(get_current_user)):
    """Send email for any new computed notifications (idempotent by id).
    Called periodically by frontend (e.g. on dashboard load).
    Each unique notification id gets sent at most once."""
    full = await get_notifications(current_user)
    notifications = full.get("notifications", [])
    if not notifications:
        return {"sent": 0, "skipped": 0}
    # Find which ones already dispatched
    ids = [n["id"] for n in notifications]
    already = await db.notification_email_log.find({"id": {"$in": ids}}, {"_id": 0, "id": 1}).to_list(2000)
    already_ids = {a["id"] for a in already}
    sent = 0
    for n in notifications:
        if n["id"] in already_ids:
            continue
        # Resolve curator email if company-related
        curator_email = None
        if n.get("company_id"):
            comp = await db.companies.find_one({"id": n["company_id"]}, {"_id": 0, "curator": 1})
            curator_email = await _user_email_by_name((comp or {}).get("curator", ""))
        severity_label = {"high": "🔴 Yüksək", "medium": "🟡 Orta", "low": "🟢 Aşağı"}.get(n.get("severity", "low"), "")
        body = f"""<p>{severity_label} prioritetli sistem bildirişi:</p>
        <table cellpadding='6' cellspacing='0' style='width:100%;border:1px solid #e2e8f0;border-radius:8px;font-size:13px'>
          <tr><td style='color:#64748b'>Başlıq</td><td style='font-weight:600'>{n.get('title', '')}</td></tr>
          <tr><td style='color:#64748b'>Mesaj</td><td>{n.get('message', '')}</td></tr>
          <tr><td style='color:#64748b'>Tarix</td><td>{n.get('date', '')}</td></tr>
        </table>
        <p style='margin-top:12px;color:#64748b;font-size:12px'>Sistemə daxil olub aydınlaşdırın.</p>"""
        await _email_notify_safe(
            title=n.get("title", "Bildiriş"),
            body_html=body,
            extra_recipients=[curator_email],
        )
        # Log every dispatch attempt so we don't retry on subsequent fetches
        await db.notification_email_log.insert_one({
            "id": n["id"],
            "type": n.get("type", ""),
            "sent_at": datetime.now(timezone.utc).isoformat(),
        })
        sent += 1
    return {"sent": sent, "skipped": len(notifications) - sent}


# ==================== SMS (LSIM Quick SMS) ====================

def _admin_only(current_user: dict):
    role = (current_user.get("role") or "").lower()
    if role != "admin":
        raise HTTPException(status_code=403, detail="Admin icazəsi tələb olunur")


async def _require_sms_permission(current_user: dict, level: str = "read"):
    """Allow admin OR any role with sms permission >= level."""
    if (current_user.get("role") or "").lower() == "admin":
        return
    perms = await get_user_permissions(current_user)
    user_level = perms.get("sms", "none")
    if level == "read" and user_level in ("read", "write"):
        return
    if level == "write" and user_level == "write":
        return
    raise HTTPException(status_code=403, detail="SMS modulu üçün icazəniz yoxdur")


def _collect_company_phones(company: dict) -> List[Dict[str, str]]:
    """Return list of {phone, name} extracted from a company doc.
    Prefers owner_phone, then representative_phone, then company_phone.
    Includes co-owners with phones."""
    out = []
    if not company:
        return out
    seen = set()
    def _add(phone, role_name):
        if phone and phone not in seen:
            seen.add(phone)
            out.append({"phone": phone, "name": role_name})
    _add(company.get("owner_phone"), company.get("owner_name") or "Sahibkar")
    _add(company.get("representative_phone"), company.get("representative_name") or "Nümayəndə")
    _add(company.get("company_phone"), company.get("brand_name") or "Şirkət")
    for o in (company.get("owners") or []):
        nm = f"{o.get('first_name','')} {o.get('last_name','')}".strip() or o.get("name", "")
        _add(o.get("phone"), nm or "Sahibkar")
    return out


@api_router.get("/sms/balance")
async def sms_balance(current_user: dict = Depends(get_current_user)):
    await _require_sms_permission(current_user, "read")
    return await sms_service.lsim_check_balance()


@api_router.get("/sms/templates")
async def sms_templates(current_user: dict = Depends(get_current_user)):
    await _require_sms_permission(current_user, "read")
    out = {}
    for k in ("event_reminder", "birthday"):
        out[k] = await sms_service.get_template(db, k)
    return out


@api_router.put("/sms/templates/{key}")
async def sms_update_template(key: str, data: dict, current_user: dict = Depends(get_current_user)):
    await _require_sms_permission(current_user, "write")
    if key not in ("event_reminder", "birthday"):
        raise HTTPException(status_code=400, detail="Yanlış template açarı")
    text = (data.get("text") or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Mətn boş ola bilməz")
    return await sms_service.set_template(db, key, text, current_user.get("name", ""))


@api_router.post("/sms/send")
async def sms_send_single(data: dict, current_user: dict = Depends(get_current_user)):
    """Send single SMS. Body: {phone, text, sender?}."""
    await _require_sms_permission(current_user, "write")
    phone = (data.get("phone") or "").strip()
    text = (data.get("text") or "").strip()
    if not phone or not text:
        raise HTTPException(status_code=400, detail="phone və text tələb olunur")
    sender = (data.get("sender") or sms_service.LSIM_SENDER).strip()
    msisdn = sms_service.normalize_phone(phone)
    res = await sms_service.lsim_send_sms(phone, text, sender)
    log = await sms_service.log_sms(
        db, phone=phone, msisdn=msisdn, text=text, sender=sender,
        category="manual", ok=res.get("ok", False),
        transid=res.get("transid"), error_code=res.get("error_code"),
        error_message=res.get("error_message"),
        recipient_type="manual", recipient_name=data.get("recipient_name", ""),
        sent_by=current_user.get("name", ""),
    )
    return {**res, "log_id": log["id"]}


@api_router.post("/sms/bulk")
async def sms_send_bulk(data: dict, current_user: dict = Depends(get_current_user)):
    """Send SMS to many recipients.
    Body: { text, sender?, recipients: [{phone, name?, type?, id?}] }
    OR: { text, sender?, recipient_type: 'companies'|'members'|'contacts',
          ids: [...] }  // server will resolve phones.
    """
    await _require_sms_permission(current_user, "write")
    text = (data.get("text") or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Mətn tələb olunur")
    sender = (data.get("sender") or sms_service.LSIM_SENDER).strip()

    recipients: List[Dict[str, Any]] = []
    explicit = data.get("recipients") or []
    if explicit:
        for r in explicit:
            ph = (r.get("phone") or "").strip()
            if ph:
                recipients.append({
                    "phone": ph,
                    "name": r.get("name", ""),
                    "type": r.get("type", "manual"),
                    "id": r.get("id", ""),
                })
    rtype = data.get("recipient_type")
    ids = data.get("ids") or []
    if rtype and ids:
        if rtype == "companies":
            comps = await db.companies.find({"id": {"$in": ids}}, {"_id": 0}).to_list(2000)
            for c in comps:
                for ph in _collect_company_phones(c):
                    recipients.append({"phone": ph["phone"], "name": ph["name"],
                                       "type": "company", "id": c.get("id", "")})
        elif rtype == "contacts":
            cts = await db.contacts.find({"id": {"$in": ids}},
                                          {"_id": 0, "id": 1, "name": 1, "surname": 1, "phone": 1}).to_list(2000)
            for ct in cts:
                if ct.get("phone"):
                    recipients.append({
                        "phone": ct["phone"],
                        "name": f"{ct.get('name','')} {ct.get('surname','')}".strip(),
                        "type": "contact", "id": ct["id"],
                    })
        elif rtype == "members":
            # Members = companies with status Aktiv
            comps = await db.companies.find({"id": {"$in": ids}, "status": "Aktiv"}, {"_id": 0}).to_list(2000)
            for c in comps:
                for ph in _collect_company_phones(c):
                    recipients.append({"phone": ph["phone"], "name": ph["name"],
                                       "type": "member", "id": c.get("id", "")})

    # Deduplicate by msisdn
    seen = set()
    unique: List[Dict[str, Any]] = []
    for r in recipients:
        msisdn = sms_service.normalize_phone(r["phone"])
        if not msisdn or msisdn in seen:
            continue
        seen.add(msisdn)
        unique.append({**r, "msisdn": msisdn})

    if not unique:
        raise HTTPException(status_code=400, detail="Heç bir yararlı telefon nömrəsi tapılmadı")

    sent_ok = 0
    sent_fail = 0
    failures: List[Dict[str, Any]] = []
    for r in unique:
        ctx = {"name": r.get("name", "")}
        msg = sms_service.render_template(text, ctx) if "{" in text else text
        res = await sms_service.lsim_send_sms(r["phone"], msg, sender)
        await sms_service.log_sms(
            db, phone=r["phone"], msisdn=r["msisdn"], text=msg, sender=sender,
            category="bulk", ok=res.get("ok", False),
            transid=res.get("transid"), error_code=res.get("error_code"),
            error_message=res.get("error_message"),
            recipient_type=r.get("type", ""), recipient_id=r.get("id", ""),
            recipient_name=r.get("name", ""), sent_by=current_user.get("name", ""),
        )
        if res.get("ok"):
            sent_ok += 1
        else:
            sent_fail += 1
            failures.append({"phone": r["phone"], "name": r.get("name", ""),
                              "error": res.get("error_message")})
    return {"total": len(unique), "sent": sent_ok, "failed": sent_fail, "failures": failures[:50]}


@api_router.get("/sms/logs")
async def sms_get_logs(category: Optional[str] = None,
                        status: Optional[str] = None,
                        limit: int = 200,
                        current_user: dict = Depends(get_current_user)):
    await _require_sms_permission(current_user, "read")
    q: Dict[str, Any] = {}
    if category:
        q["category"] = category
    if status:
        q["status"] = status
    items = await db.sms_logs.find(q, {"_id": 0}).sort("created_at", -1).to_list(min(limit, 1000))
    total = await db.sms_logs.count_documents(q)
    return {"items": items, "total": total}


@api_router.get("/sms/logs/stats")
async def sms_log_stats(current_user: dict = Depends(get_current_user)):
    await _require_sms_permission(current_user, "read")
    total = await db.sms_logs.count_documents({})
    sent = await db.sms_logs.count_documents({"status": "sent"})
    failed = await db.sms_logs.count_documents({"status": "failed"})
    today_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    today = await db.sms_logs.count_documents({"created_at": {"$gte": today_iso}})
    by_category = {}
    for cat in ("manual", "bulk", "event_reminder", "birthday"):
        by_category[cat] = await db.sms_logs.count_documents({"category": cat})
    return {"total": total, "sent": sent, "failed": failed, "today": today, "by_category": by_category}


@api_router.post("/sms/dispatch-daily")
async def sms_dispatch_daily(current_user: dict = Depends(get_current_user)):
    """Send: (a) tomorrow's event reminders to invited companies' phones,
    (b) today's birthday SMS to owners/reps/employees/contacts.
    Idempotent — uses sms_logs to skip already-sent items in the same calendar day.
    """
    await _require_sms_permission(current_user, "write")
    now = datetime.now(timezone.utc)
    today_iso = now.strftime("%Y-%m-%d")
    tomorrow = (now + timedelta(days=1)).strftime("%Y-%m-%d")
    today_md = now.strftime("%m-%d")

    template_event = await sms_service.get_template(db, "event_reminder")
    template_bday = await sms_service.get_template(db, "birthday")
    sender = sms_service.LSIM_SENDER

    sent_event = 0
    sent_bday = 0
    skipped = 0
    failed = 0

    async def _already_sent(category: str, key: str) -> bool:
        return await db.sms_logs.count_documents({
            "category": category,
            "related_object_id": key,
            "created_at": {"$gte": today_iso},
        }) > 0

    # ===== A) Event reminders for events scheduled tomorrow =====
    events = await db.events.find({"date": tomorrow}, {"_id": 0}).to_list(500)
    for ev in events:
        invs = await db.invitations.find({"event_id": ev["id"]}, {"_id": 0}).to_list(2000)
        for inv in invs:
            if (inv.get("participation_status") or "").lower() == "imtina":
                continue
            comp_id = inv.get("company_id")
            if not comp_id:
                continue
            comp = await db.companies.find_one({"id": comp_id}, {"_id": 0})
            phones = _collect_company_phones(comp)
            for ph in phones:
                key = f"{ev['id']}:{comp_id}:{ph['phone']}"
                if await _already_sent("event_reminder", key):
                    skipped += 1
                    continue
                ctx = {
                    "name": ph.get("name", ""),
                    "event_name": ev.get("name", ""),
                    "event_type": ev.get("event_type", ""),
                    "date": ev.get("date", ""),
                    "time": ev.get("time", ""),
                    "venue": ev.get("venue", ""),
                    "company": comp.get("brand_name", "") if comp else "",
                }
                msg = sms_service.render_template(template_event, ctx)
                res = await sms_service.lsim_send_sms(ph["phone"], msg, sender)
                await sms_service.log_sms(
                    db, phone=ph["phone"], msisdn=sms_service.normalize_phone(ph["phone"]),
                    text=msg, sender=sender, category="event_reminder",
                    ok=res.get("ok", False), transid=res.get("transid"),
                    error_code=res.get("error_code"), error_message=res.get("error_message"),
                    recipient_type="company", recipient_id=comp_id, recipient_name=ph.get("name", ""),
                    sent_by=current_user.get("name", "") + " (auto)",
                    related_object_id=key, related_object_type="event",
                )
                if res.get("ok"):
                    sent_event += 1
                else:
                    failed += 1

    # ===== B) Birthday SMS today =====
    def _md(d):
        if not d:
            return None
        try:
            return datetime.strptime(str(d)[:10], "%Y-%m-%d").strftime("%m-%d")
        except (ValueError, TypeError):
            return None

    async def _send_bday(person_name: str, phone: str, kind: str, key: str):
        nonlocal sent_bday, skipped, failed
        if not phone:
            return
        if await _already_sent("birthday", key):
            skipped += 1
            return
        ctx = {"name": person_name}
        msg = sms_service.render_template(template_bday, ctx)
        res = await sms_service.lsim_send_sms(phone, msg, sender)
        await sms_service.log_sms(
            db, phone=phone, msisdn=sms_service.normalize_phone(phone),
            text=msg, sender=sender, category="birthday",
            ok=res.get("ok", False), transid=res.get("transid"),
            error_code=res.get("error_code"), error_message=res.get("error_message"),
            recipient_type=kind, recipient_name=person_name,
            sent_by=current_user.get("name", "") + " (auto)",
            related_object_id=key, related_object_type="birthday",
        )
        if res.get("ok"):
            sent_bday += 1
        else:
            failed += 1

    # Owners / reps / contact_person on companies
    bday_companies = await db.companies.find({}, {"_id": 0}).to_list(2000)
    for c in bday_companies:
        cid = c.get("id", "")
        if _md(c.get("owner_birth_date")) == today_md and c.get("owner_phone"):
            await _send_bday(c.get("owner_name", "") or "Sahibkar", c["owner_phone"],
                              "owner", f"owner:{cid}:{today_md}")
        for idx, o in enumerate(c.get("owners") or []):
            if _md(o.get("birth_date")) == today_md and o.get("phone"):
                pname = f"{o.get('first_name','')} {o.get('last_name','')}".strip() or o.get("name", "Sahibkar")
                await _send_bday(pname, o["phone"], "owner", f"owner-co:{cid}:{idx}:{today_md}")
        if _md(c.get("representative_birth_date")) == today_md and c.get("representative_phone"):
            await _send_bday(c.get("representative_name", "") or "Nümayəndə", c["representative_phone"],
                              "representative", f"rep:{cid}:{today_md}")

    # Employees
    emps = await db.employees.find({"birth_date": {"$exists": True, "$ne": ""}}, {"_id": 0}).to_list(1000)
    for e in emps:
        if _md(e.get("birth_date")) == today_md and e.get("phone"):
            full = f"{e.get('first_name','')} {e.get('last_name','')}".strip()
            await _send_bday(full or "Əməkdaş", e["phone"], "employee", f"emp:{e.get('id','')}:{today_md}")

    # Contacts (siyahılardakı)
    cts = await db.contacts.find({"birthday": {"$exists": True, "$ne": ""}},
                                  {"_id": 0, "id": 1, "name": 1, "surname": 1, "phone": 1, "birthday": 1}).to_list(2000)
    for ct in cts:
        if _md(ct.get("birthday")) == today_md and ct.get("phone"):
            full = f"{ct.get('name','')} {ct.get('surname','')}".strip()
            await _send_bday(full or "Kontakt", ct["phone"], "contact", f"ctc:{ct.get('id','')}:{today_md}")

    return {
        "event_reminders_sent": sent_event,
        "birthday_sent": sent_bday,
        "skipped": skipped,
        "failed": failed,
    }


# ==================== MARKETING (Mailchimp + SMS Campaigns) ====================

@api_router.get("/marketing/mailchimp/ping")
async def mailchimp_ping(current_user: dict = Depends(get_current_user)):
    _admin_only(current_user)
    return await mailchimp_service.ping()


@api_router.get("/marketing/mailchimp/audiences")
async def mailchimp_audiences(current_user: dict = Depends(get_current_user)):
    _admin_only(current_user)
    return await mailchimp_service.list_audiences()


@api_router.post("/marketing/mailchimp/audiences/{audience_id}/sync-companies")
async def mailchimp_sync_companies(audience_id: str, data: dict = None, current_user: dict = Depends(get_current_user)):
    """Push selected companies (or all active members) to a Mailchimp audience as subscribers."""
    _admin_only(current_user)
    data = data or {}
    ids = data.get("ids") or []
    query = {"id": {"$in": ids}} if ids else {"status": "Aktiv"}
    companies = await db.companies.find(query, {"_id": 0}).to_list(2000)
    sent = 0
    failed = 0
    skipped = 0
    failures = []
    for c in companies:
        email = c.get("contact_email") or c.get("owner_email") or ""
        if not email:
            skipped += 1
            continue
        first = c.get("contact_first_name") or c.get("owner_first_name") or c.get("owner_name", "").split(" ")[0] if c.get("owner_name") else ""
        last = c.get("contact_last_name") or c.get("owner_last_name") or " ".join((c.get("owner_name") or "").split(" ")[1:]) if c.get("owner_name") else ""
        company_name = c.get("brand_name") or c.get("legal_name", "")
        res = await mailchimp_service.upsert_member(audience_id, email, first or "", last or "", company_name)
        if res.get("ok"):
            sent += 1
        else:
            failed += 1
            failures.append({"email": email, "company": company_name, "error": res.get("error")})
    return {"total": len(companies), "synced": sent, "failed": failed, "skipped_no_email": skipped, "failures": failures[:30]}


@api_router.get("/marketing/mailchimp/campaigns")
async def mailchimp_list_campaigns(status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    _admin_only(current_user)
    return await mailchimp_service.list_campaigns(count=50, status=status)


@api_router.get("/marketing/mailchimp/campaigns/{campaign_id}/report")
async def mailchimp_campaign_report(campaign_id: str, current_user: dict = Depends(get_current_user)):
    _admin_only(current_user)
    return await mailchimp_service.get_campaign_report(campaign_id)


@api_router.post("/marketing/mailchimp/campaigns")
async def mailchimp_create_send_campaign(data: dict, current_user: dict = Depends(get_current_user)):
    """Body: {audience_id, subject, title, from_name, reply_to, preview_text, html, send_now}."""
    _admin_only(current_user)
    audience_id = (data.get("audience_id") or "").strip()
    subject = (data.get("subject") or "").strip()
    title = (data.get("title") or subject or "Marsol Kampaniya").strip()
    from_name = (data.get("from_name") or "Marsol Group").strip()
    reply_to = (data.get("reply_to") or "").strip()
    html = data.get("html") or ""
    send_now = bool(data.get("send_now"))
    if not audience_id or not subject or not html or not reply_to:
        raise HTTPException(status_code=400, detail="audience_id, subject, html, reply_to tələb olunur")
    create_res = await mailchimp_service.create_campaign(audience_id, subject, title, from_name, reply_to, data.get("preview_text", ""))
    if not create_res.get("ok"):
        raise HTTPException(status_code=400, detail=create_res.get("error") or "Kampaniya yaradıla bilmədi")
    campaign_id = create_res["data"].get("id")
    content_res = await mailchimp_service.set_campaign_content(campaign_id, html)
    if not content_res.get("ok"):
        raise HTTPException(status_code=400, detail=content_res.get("error") or "Məzmun yüklənə bilmədi")
    sent = False
    if send_now:
        send_res = await mailchimp_service.send_campaign(campaign_id)
        if not send_res.get("ok"):
            raise HTTPException(status_code=400, detail=send_res.get("error") or "Göndərmə baş tutmadı")
        sent = True
    # log
    await db.email_campaigns.insert_one({
        "id": str(uuid.uuid4()),
        "mailchimp_id": campaign_id,
        "subject": subject,
        "title": title,
        "audience_id": audience_id,
        "html": html,
        "from_name": from_name,
        "reply_to": reply_to,
        "send_now": send_now,
        "status": "sent" if sent else "draft",
        "created_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    return {"ok": True, "campaign_id": campaign_id, "status": "sent" if sent else "draft"}


@api_router.post("/marketing/email/bulk")
async def marketing_email_bulk(data: dict, current_user: dict = Depends(get_current_user)):
    """Send a bulk email through Mailchimp to selected internal recipients.

    Body: { audience_id, subject, html, from_name?, reply_to?, preview_text?, send_now?,
            recipient_type: 'companies'|'members'|'contacts'|'project_leads', ids: [...] }

    Flow: 1) upsert recipients to the chosen Mailchimp audience, 2) create a
    static segment containing those recipients, 3) create a regular campaign
    targeted at that segment, 4) optionally send. The audience grows naturally
    as new emails are added — but the campaign reaches ONLY the selected ones.
    """
    _admin_only(current_user)
    audience_id = (data.get("audience_id") or "").strip()
    subject = (data.get("subject") or "").strip()
    html = data.get("html") or ""
    if not audience_id or not subject or not html:
        raise HTTPException(status_code=400, detail="audience_id, subject və html tələb olunur")
    from_name = (data.get("from_name") or "Marsol Group").strip()
    reply_to = (data.get("reply_to") or "info@marsol.az").strip()
    preview_text = (data.get("preview_text") or "").strip()
    send_now = bool(data.get("send_now"))

    # Resolve recipients
    recipients: List[Dict[str, str]] = []
    explicit = data.get("recipients") or []
    for r in explicit:
        em = (r.get("email") or "").strip()
        if em:
            recipients.append({"email": em, "name": r.get("name", ""), "company": r.get("company", ""), "type": "manual", "id": ""})
    rtype = data.get("recipient_type")
    ids = data.get("ids") or []
    if rtype and ids:
        if rtype in ("companies", "members"):
            q = {"id": {"$in": ids}}
            if rtype == "members":
                q["status"] = "Aktiv"
            for c in await db.companies.find(q, {"_id": 0}).to_list(2000):
                for em in [c.get("contact_email"), c.get("owner_email")]:
                    if em:
                        recipients.append({
                            "email": em,
                            "name": c.get("contact_name") or c.get("owner_name", ""),
                            "company": c.get("brand_name") or c.get("legal_name", ""),
                            "type": rtype, "id": c.get("id", ""),
                        })
        elif rtype == "contacts":
            for ct in await db.contacts.find({"id": {"$in": ids}}, {"_id": 0}).to_list(2000):
                if ct.get("email"):
                    recipients.append({
                        "email": ct["email"],
                        "name": f"{ct.get('name','')} {ct.get('surname','')}".strip(),
                        "company": ct.get("company", ""),
                        "type": "contact", "id": ct["id"],
                    })
        elif rtype == "project_leads":
            for ld in await db.sales_leads.find({"id": {"$in": ids}}, {"_id": 0}).to_list(2000):
                if ld.get("email"):
                    recipients.append({
                        "email": ld["email"],
                        "name": ld.get("contact_name") or "",
                        "company": ld.get("company_name", ""),
                        "type": "project_lead", "id": ld["id"],
                    })

    # Deduplicate by email
    seen = set()
    unique = []
    for r in recipients:
        em = (r["email"] or "").strip().lower()
        if "@" not in em or em in seen:
            continue
        seen.add(em)
        unique.append({**r, "email": em})
    if not unique:
        raise HTTPException(status_code=400, detail="Heç bir yararlı email tapılmadı")
    if len(unique) > 5000:
        raise HTTPException(status_code=400, detail="Tək kampaniyada maksimum 5000 alıcı")

    # 1) Upsert each recipient to the audience
    upsert_failed = []
    for r in unique:
        first, last = "", ""
        if r.get("name"):
            parts = r["name"].split(" ", 1)
            first = parts[0]
            last = parts[1] if len(parts) > 1 else ""
        res = await mailchimp_service.upsert_member(audience_id, r["email"], first, last, r.get("company", ""))
        if not res.get("ok"):
            upsert_failed.append({"email": r["email"], "error": res.get("error")})

    # If everyone failed (e.g. invalid audience), abort
    if len(upsert_failed) >= len(unique):
        raise HTTPException(status_code=400, detail=f"Heç bir alıcı Mailchimp-ə əlavə edilə bilmədi: {upsert_failed[0]['error'] if upsert_failed else 'naməlum xəta'}")

    valid_emails = [r["email"] for r in unique if r["email"] not in {f["email"] for f in upsert_failed}]

    # 2) Create static segment for these emails
    segment_name = f"Marsol-bulk-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}-{str(uuid.uuid4())[:6]}"
    seg_res = await mailchimp_service.create_static_segment(audience_id, segment_name, valid_emails)
    if not seg_res.get("ok"):
        raise HTTPException(status_code=400, detail=f"Statik seqment yaradıla bilmədi: {seg_res.get('error')}")
    segment_id = seg_res["data"].get("id")
    segment_member_count = seg_res["data"].get("member_count", len(valid_emails))

    # 3) Create campaign for that segment
    title = f"{subject} ({segment_name})"
    cmp_res = await mailchimp_service.create_campaign_for_segment(audience_id, segment_id, subject, title, from_name, reply_to, preview_text)
    if not cmp_res.get("ok"):
        raise HTTPException(status_code=400, detail=f"Kampaniya yaradıla bilmədi: {cmp_res.get('error')}")
    campaign_id = cmp_res["data"].get("id")

    # 4) Set content
    content_res = await mailchimp_service.set_campaign_content(campaign_id, html)
    if not content_res.get("ok"):
        raise HTTPException(status_code=400, detail=f"Məzmun yüklənə bilmədi: {content_res.get('error')}")

    # 5) Send (if requested)
    sent = False
    send_error = None
    if send_now:
        send_res = await mailchimp_service.send_campaign(campaign_id)
        if send_res.get("ok"):
            sent = True
        else:
            send_error = send_res.get("error")

    # Persist a log row per recipient (status reflects upsert result; campaign-send is one-shot)
    for r in unique:
        in_failed = next((f for f in upsert_failed if f["email"] == r["email"]), None)
        await db.email_logs.insert_one({
            "id": str(uuid.uuid4()),
            "email": r["email"], "name": r.get("name", ""), "subject": subject,
            "category": "bulk", "provider": "mailchimp",
            "status": "failed" if in_failed else ("sent" if sent else "queued"),
            "error": (in_failed or {}).get("error") or send_error,
            "recipient_type": r.get("type", ""), "recipient_id": r.get("id", ""),
            "audience_id": audience_id, "segment_id": segment_id, "campaign_id": campaign_id,
            "sent_by": current_user.get("name", ""),
            "created_at": datetime.now(timezone.utc).isoformat(),
        })

    return {
        "ok": True,
        "total": len(unique),
        "synced_to_audience": len(valid_emails),
        "upsert_failed": len(upsert_failed),
        "segment_id": segment_id,
        "segment_member_count": segment_member_count,
        "campaign_id": campaign_id,
        "campaign_status": "sent" if sent else "draft",
        "send_error": send_error,
        "failures": upsert_failed[:30],
    }


@api_router.get("/marketing/email/logs")
async def marketing_email_logs(category: Optional[str] = None, limit: int = 200, current_user: dict = Depends(get_current_user)):
    _admin_only(current_user)
    q: Dict[str, Any] = {}
    if category:
        q["category"] = category
    items = await db.email_logs.find(q, {"_id": 0}).sort("created_at", -1).to_list(min(limit, 1000))
    total = await db.email_logs.count_documents(q)
    return {"items": items, "total": total}


# ==================== PARTNER EVALUATION (Reytinq) ====================
# 100-point system:
#   • Payment activity  — 40
#   • Event attendance  — 30
#   • Other projects    — 15
#   • Meetings          — 10
#   • Manual            —  5

async def _calc_partner_score(company: dict) -> Dict[str, Any]:
    cid = company.get("id")
    # Payment (40): use latest contract paid_amount / total_amount ratio if available
    contracts = await db.contracts.find({"company_id": cid}, {"_id": 0}).to_list(50) if hasattr(db, "contracts") else []
    payment_score = 0
    if contracts:
        active = sorted(contracts, key=lambda c: c.get("created_at", ""), reverse=True)[0]
        total = float(active.get("total_amount") or 0)
        paid = float(active.get("paid_amount") or 0)
        if total > 0:
            ratio = min(paid / total, 1.0)
            # Late payment penalty handled simply via ratio; bonus on full
            payment_score = int(round(ratio * 40))
    elif company.get("payment_status") == "Ödənilib":
        payment_score = 40

    # Events (30): how many invitations are 'Qatılır' / 'İştirak etdi' last 12 mo
    one_year_ago = (datetime.now(timezone.utc) - timedelta(days=365)).isoformat()
    invs = await db.invitations.find({"company_id": cid, "called_at": {"$gte": one_year_ago}}, {"_id": 0}).to_list(500)
    accepted = sum(1 for i in invs if (i.get("participation_status") or "") in ("Qatılır", "İştirak etdi"))
    event_total = len(invs) or 1
    event_score = int(round((accepted / event_total) * 30)) if invs else 0

    # Other projects (15): sales-leads with status Satıldı / Üzv oldu
    other = await db.sales_leads.count_documents({
        "company_name": company.get("brand_name", ""),
        "status": {"$in": ["Satıldı", "Üzv oldu"]},
        "created_at": {"$gte": one_year_ago},
    })
    other_score = min(other * 3, 15)  # 5 projects = full

    # Meetings (10): meetings with this company in last 12mo
    meetings = await db.meetings.count_documents({
        "company_id": cid,
        "date": {"$gte": one_year_ago[:10]},
    }) if cid else 0
    meeting_score = min(meetings, 10)

    # Manual (5): from db.partner_evaluations
    manual_doc = await db.partner_evaluations.find_one({"company_id": cid}, {"_id": 0})
    manual_score = int((manual_doc or {}).get("manual_bonus", 0) or 0)
    manual_score = max(0, min(manual_score, 5))

    total_score = payment_score + event_score + other_score + meeting_score + manual_score
    return {
        "company_id": cid,
        "brand_name": company.get("brand_name", ""),
        "scores": {
            "payment": payment_score,
            "event": event_score,
            "other_projects": other_score,
            "meetings": meeting_score,
            "manual": manual_score,
        },
        "total": total_score,
        "tier": "Platinum" if total_score >= 85 else "Qızıl" if total_score >= 65 else "Gümüş" if total_score >= 40 else "Standart",
    }


def _calc_partner_score_from_bulk(
    company: dict,
    inv_stats: Dict[str, int],
    other_count: int,
    meeting_count: int,
    manual_bonus: int,
) -> Dict[str, Any]:
    """Pure helper — combines pre-fetched aggregates into a partner score dict."""
    cid = company.get("id")
    # Payment score: simplified (no per-company contracts query — use company.payment_status)
    payment_score = 40 if company.get("payment_status") == "Ödənilib" else 0
    # Or use top-level paid/total ratio if available
    total = float(company.get("total_amount") or 0)
    paid = float(company.get("paid_amount") or 0)
    if total > 0:
        ratio = min(paid / total, 1.0)
        payment_score = int(round(ratio * 40))

    invited = int(inv_stats.get("invited", 0) or 0)
    accepted = int(inv_stats.get("accepted", 0) or 0)

    # Honor obligation_overrides written by the Excel import — when the user
    # migrates historical data they fill total_invited / total_attended, and
    # those should feed the event score too (otherwise reytinq always says 0).
    overrides_doc = company.get("obligation_overrides") or {}
    # Prefer the 'all' bucket; else merge any year-scoped buckets together
    # so we don't silently miss a 2026-only override when no 'all' exists.
    ov_all = overrides_doc.get("all") or {}
    if not ov_all and overrides_doc:
        merged_inv = 0
        merged_att = 0
        for v in overrides_doc.values():
            if isinstance(v, dict):
                if v.get("total_invited") is not None:
                    merged_inv += int(v.get("total_invited") or 0)
                if v.get("total_attended") is not None:
                    merged_att += int(v.get("total_attended") or 0)
        if merged_inv > 0 or merged_att > 0:
            ov_all = {"total_invited": merged_inv, "total_attended": merged_att}
    if ov_all.get("total_invited") is not None:
        invited = max(invited, int(ov_all.get("total_invited") or 0))
    if ov_all.get("total_attended") is not None:
        accepted = max(accepted, int(ov_all.get("total_attended") or 0))

    # Event score is based on ABSOLUTE attendance count so companies that
    # attend more events earn more points. Each attended event = 3 points,
    # capped at 30. Previous ratio-based formula (attended/invited * 30)
    # penalised companies that were invited many times — e.g. 1/1=30 vs
    # 5/10=15. With this scheme 5 attendances always beats 1.
    event_score = min(int(accepted) * 3, 30)

    other_score = min(int(other_count) * 3, 15)
    meeting_score = min(int(meeting_count), 10)
    manual_score = max(0, min(int(manual_bonus or 0), 5))

    total_score = payment_score + event_score + other_score + meeting_score + manual_score
    return {
        "company_id": cid,
        "display_id": company.get("display_id", ""),
        "brand_name": company.get("brand_name", ""),
        "scores": {
            "payment": payment_score,
            "event": event_score,
            "other_projects": other_score,
            "meetings": meeting_score,
            "manual": manual_score,
        },
        "total": total_score,
        "tier": "Platinum" if total_score >= 85 else "Qızıl" if total_score >= 65 else "Gümüş" if total_score >= 40 else "Standart",
    }


@api_router.get("/partner-evaluation")
async def partner_evaluation_list(current_user: dict = Depends(get_current_user)):
    """Compute reytinq for all active member companies. Bulk aggregation (4 queries total) instead of per-company loops."""
    companies = await db.companies.find({"status": "Aktiv"}, {"_id": 0}).to_list(2000)
    company_ids = [c.get("id", "") for c in companies]
    company_names = [c.get("brand_name", "") for c in companies if c.get("brand_name")]
    one_year_ago = (datetime.now(timezone.utc) - timedelta(days=365)).isoformat()
    one_year_ago_date = one_year_ago[:10]

    # Run 4 aggregations in parallel (1 per metric)
    inv_pipe = [
        {"$match": {"company_id": {"$in": company_ids}, "called_at": {"$gte": one_year_ago}}},
        {"$group": {
            "_id": "$company_id",
            "invited": {"$sum": 1},
            "accepted": {"$sum": {"$cond": [{"$in": ["$participation_status", ["Qatılır", "İştirak etdi"]]}, 1, 0]}},
        }},
    ]
    leads_pipe = [
        {"$match": {
            "company_name": {"$in": company_names},
            "status": {"$in": ["Satıldı", "Üzv oldu"]},
            "created_at": {"$gte": one_year_ago},
        }},
        {"$group": {"_id": "$company_name", "n": {"$sum": 1}}},
    ]
    meetings_pipe = [
        {"$match": {"company_id": {"$in": company_ids}, "date": {"$gte": one_year_ago_date}}},
        {"$group": {"_id": "$company_id", "n": {"$sum": 1}}},
    ]

    inv_rows, leads_rows, meeting_rows, manual_rows = await asyncio.gather(
        db.invitations.aggregate(inv_pipe).to_list(5000),
        db.sales_leads.aggregate(leads_pipe).to_list(5000),
        db.meetings.aggregate(meetings_pipe).to_list(5000),
        db.partner_evaluations.find({"company_id": {"$in": company_ids}}, {"_id": 0}).to_list(5000),
    )
    inv_map = {r["_id"]: r for r in inv_rows}
    leads_map = {r["_id"]: r["n"] for r in leads_rows}
    meeting_map = {r["_id"]: r["n"] for r in meeting_rows}
    manual_map = {r["company_id"]: int(r.get("manual_bonus", 0) or 0) for r in manual_rows}

    results = [
        _calc_partner_score_from_bulk(
            c,
            inv_map.get(c.get("id", ""), {}),
            leads_map.get(c.get("brand_name", ""), 0),
            meeting_map.get(c.get("id", ""), 0),
            manual_map.get(c.get("id", ""), 0),
        )
        for c in companies
    ]
    results.sort(key=lambda x: x["total"], reverse=True)
    return {"items": results, "total": len(results)}


@api_router.get("/partner-evaluation/{company_id}")
async def partner_evaluation_one(company_id: str, current_user: dict = Depends(get_current_user)):
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Şirkət tapılmadı")
    return await _calc_partner_score(company)


@api_router.put("/partner-evaluation/{company_id}/manual-bonus")
async def partner_evaluation_set_bonus(company_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    _admin_only(current_user)
    bonus = max(0, min(int(data.get("manual_bonus") or 0), 5))
    note = (data.get("note") or "").strip()
    await db.partner_evaluations.update_one(
        {"company_id": company_id},
        {"$set": {
            "company_id": company_id,
            "manual_bonus": bonus,
            "note": note,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user.get("name", ""),
        }},
        upsert=True,
    )
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    return await _calc_partner_score(company)


# ==================== MESSAGE GROUPS ====================

@api_router.get("/message-groups")
async def list_message_groups(current_user: dict = Depends(get_current_user)):
    user_id = current_user.get("id")
    # User sees groups they're a member of OR created
    groups = await db.message_groups.find({
        "$or": [{"members": user_id}, {"created_by_id": user_id}]
    }, {"_id": 0}).sort("updated_at", -1).to_list(200)
    return groups


@api_router.post("/message-groups")
async def create_message_group(data: dict, current_user: dict = Depends(get_current_user)):
    name = (data.get("name") or "").strip()
    members = data.get("members") or []
    if not name or len(members) < 2:
        raise HTTPException(status_code=400, detail="Ad və ən azı 2 üzv tələb olunur")
    user_id = current_user.get("id")
    if user_id and user_id not in members:
        members.append(user_id)
    group = {
        "id": str(uuid.uuid4()),
        "name": name,
        "description": (data.get("description") or "").strip(),
        "color": data.get("color") or "#9ACD32",
        "members": list(set(members)),
        "created_by": current_user.get("name", ""),
        "created_by_id": user_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.message_groups.insert_one(group)
    group.pop("_id", None)
    return group


@api_router.put("/message-groups/{group_id}")
async def update_message_group(group_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    g = await db.message_groups.find_one({"id": group_id}, {"_id": 0})
    if not g:
        raise HTTPException(status_code=404, detail="Qrup tapılmadı")
    if current_user.get("id") not in g.get("members", []) and current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="İcazə yoxdur")
    updates = {"updated_at": datetime.now(timezone.utc).isoformat()}
    for k in ("name", "description", "color", "members"):
        if k in data:
            updates[k] = data[k]
    if "members" in updates and current_user.get("id") not in updates["members"]:
        updates["members"].append(current_user.get("id"))
    await db.message_groups.update_one({"id": group_id}, {"$set": updates})
    g2 = await db.message_groups.find_one({"id": group_id}, {"_id": 0})
    return g2


@api_router.delete("/message-groups/{group_id}")
async def delete_message_group(group_id: str, current_user: dict = Depends(get_current_user)):
    g = await db.message_groups.find_one({"id": group_id}, {"_id": 0})
    if not g:
        raise HTTPException(status_code=404, detail="Qrup tapılmadı")
    if g.get("created_by_id") != current_user.get("id") and current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Yalnız yaradıcı və ya admin silə bilər")
    await db.message_groups.delete_one({"id": group_id})
    await db.group_messages.delete_many({"group_id": group_id})
    return {"deleted": True}


@api_router.get("/message-groups/{group_id}/messages")
async def get_group_messages(group_id: str, current_user: dict = Depends(get_current_user)):
    g = await db.message_groups.find_one({"id": group_id}, {"_id": 0})
    if not g or current_user.get("id") not in g.get("members", []):
        raise HTTPException(status_code=403, detail="İcazə yoxdur")
    msgs = await db.group_messages.find({"group_id": group_id}, {"_id": 0}).sort("created_at", 1).to_list(500)
    return msgs


@api_router.post("/message-groups/{group_id}/messages")
async def send_group_message(group_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    g = await db.message_groups.find_one({"id": group_id}, {"_id": 0})
    if not g or current_user.get("id") not in g.get("members", []):
        raise HTTPException(status_code=403, detail="İcazə yoxdur")
    body = (data.get("body") or "").strip()
    if not body:
        raise HTTPException(status_code=400, detail="Mətn boş ola bilməz")
    msg = {
        "id": str(uuid.uuid4()),
        "group_id": group_id,
        "body": body,
        "sender_id": current_user.get("id"),
        "sender_name": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.group_messages.insert_one(msg)
    await db.message_groups.update_one({"id": group_id}, {"$set": {"updated_at": msg["created_at"]}})
    msg.pop("_id", None)
    return msg


# Root
@api_router.get("/")
async def root():
    return {"message": "Marsol Group Management System API"}

# ==================== TƏŞKİLATÇILIQ / FƏALİYYƏTLƏR — VENDOR MODULLARI ====================

ORG_COLLECTIONS = {
    "venues": "org_venues",
    "catering": "org_catering",
    "decor": "org_decor",
    "musicians": "org_musicians",
    "photovideo": "org_photovideo",
    "transport": "org_transport",
    "materials": "org_materials",
}

ORG_LABELS = {
    "venues": "Məkanlar",
    "catering": "Catering",
    "decor": "Dekor və texniki təchizat",
    "musicians": "Musiqiçilər və şou komandaları",
    "photovideo": "Foto / Video",
    "transport": "Nəqliyyat və logistika",
    "materials": "Tədbir materialları",
}

# ===== RATINGS (DECLARE BEFORE /organization/{module} TO AVOID ROUTE CONFLICT) =====

@api_router.get("/organization/ratings/list")
async def list_ratings(
    vendor_type: Optional[str] = None,
    vendor_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if vendor_type and vendor_type != "all":
        query["vendor_type"] = vendor_type
    if vendor_id:
        query["vendor_id"] = vendor_id
    items = await db.org_ratings.find(query, {"_id": 0}).sort("event_date", -1).to_list(5000)
    return items

@api_router.post("/organization/ratings")
async def create_rating(data: dict, current_user: dict = Depends(check_permission("organization", "write"))):
    vendor_type = data.get("vendor_type", "")
    if vendor_type not in ORG_COLLECTIONS:
        raise HTTPException(status_code=400, detail="Yanlış təchizatçı növü")
    vendor = await db[ORG_COLLECTIONS[vendor_type]].find_one({"id": data.get("vendor_id", "")}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Təchizatçı tapılmadı")
    def f(x, default=0):
        try: return float(x or default)
        except (ValueError, TypeError): return default
    doc = {
        "id": str(uuid.uuid4()),
        "vendor_type": vendor_type,
        "vendor_id": data.get("vendor_id", ""),
        "vendor_name": vendor.get("name") or vendor.get("vendor_name") or vendor.get("venue_name") or "",
        "event_name": data.get("event_name", ""),
        "event_date": data.get("event_date", ""),
        "price_score": f(data.get("price_score"), 5),
        "quality_score": f(data.get("quality_score"), 5),
        "operativity_score": f(data.get("operativity_score"), 5),
        "behavior_score": f(data.get("behavior_score"), 5),
        "flexibility_score": f(data.get("flexibility_score"), 5),
        "event_fit_score": f(data.get("event_fit_score"), 5),
        "rehire_willingness": data.get("rehire_willingness", "Bəli"),
        "comment": data.get("comment", ""),
        "rated_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.org_ratings.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/organization/ratings/{rating_id}")
async def delete_rating(rating_id: str, current_user: dict = Depends(check_permission("organization", "write"))):
    await db.org_ratings.delete_one({"id": rating_id})
    return {"message": "Reytinq silindi"}

@api_router.get("/organization/ratings/summary")
async def ratings_summary(current_user: dict = Depends(get_current_user)):
    pipeline = [
        {"$group": {
            "_id": {"vendor_type": "$vendor_type", "vendor_id": "$vendor_id"},
            "vendor_name": {"$last": "$vendor_name"},
            "count": {"$sum": 1},
            "avg_price": {"$avg": "$price_score"},
            "avg_quality": {"$avg": "$quality_score"},
            "avg_operativity": {"$avg": "$operativity_score"},
            "avg_behavior": {"$avg": "$behavior_score"},
            "avg_flexibility": {"$avg": "$flexibility_score"},
            "avg_fit": {"$avg": "$event_fit_score"},
            "rehire_yes": {"$sum": {"$cond": [{"$eq": ["$rehire_willingness", "Bəli"]}, 1, 0]}},
            "last_event_date": {"$max": "$event_date"},
        }},
        {"$project": {
            "_id": 0,
            "vendor_type": "$_id.vendor_type",
            "vendor_id": "$_id.vendor_id",
            "vendor_name": 1, "count": 1,
            "avg_price": {"$round": ["$avg_price", 2]},
            "avg_quality": {"$round": ["$avg_quality", 2]},
            "avg_operativity": {"$round": ["$avg_operativity", 2]},
            "avg_behavior": {"$round": ["$avg_behavior", 2]},
            "avg_flexibility": {"$round": ["$avg_flexibility", 2]},
            "avg_fit": {"$round": ["$avg_fit", 2]},
            "overall": {"$round": [{"$divide": [{"$add": ["$avg_price", "$avg_quality", "$avg_operativity", "$avg_behavior"]}, 4]}, 2]},
            "rehire_rate": {"$round": [{"$multiply": [{"$divide": ["$rehire_yes", "$count"]}, 100]}, 1]},
            "last_event_date": 1,
        }},
        {"$sort": {"overall": -1}}
    ]
    results = await db.org_ratings.aggregate(pipeline).to_list(5000)
    for r in results:
        ov = r.get("overall") or 0
        rr = r.get("rehire_rate") or 0
        if ov >= 4.2 and rr >= 75:
            r["recommendation"] = "Tövsiyə edilir"
        elif ov >= 3.0:
            r["recommendation"] = "Şərtlə tövsiyə"
        else:
            r["recommendation"] = "Tövsiyə edilmir"
    return results

# ===== DASHBOARD (DECLARE BEFORE /organization/{module}) =====

@api_router.get("/organization/dashboard/stats")
async def org_dashboard(current_user: dict = Depends(get_current_user)):
    counts = {}
    for mod, col in ORG_COLLECTIONS.items():
        counts[mod] = await db[col].count_documents({})
    total_ratings = await db.org_ratings.count_documents({})
    pipeline = [
        {"$group": {
            "_id": {"vendor_type": "$vendor_type", "vendor_id": "$vendor_id"},
            "vendor_name": {"$last": "$vendor_name"},
            "avg_price": {"$avg": "$price_score"},
            "avg_quality": {"$avg": "$quality_score"},
            "avg_operativity": {"$avg": "$operativity_score"},
            "avg_behavior": {"$avg": "$behavior_score"},
            "count": {"$sum": 1},
        }},
        {"$project": {
            "_id": 0,
            "vendor_type": "$_id.vendor_type",
            "vendor_id": "$_id.vendor_id",
            "vendor_name": 1,
            "count": 1,
            "overall": {"$round": [{"$divide": [{"$add": ["$avg_price", "$avg_quality", "$avg_operativity", "$avg_behavior"]}, 4]}, 2]},
        }},
        {"$sort": {"overall": -1}},
        {"$limit": 5}
    ]
    top_rated = await db.org_ratings.aggregate(pipeline).to_list(5)
    recent = []
    for mod, col in ORG_COLLECTIONS.items():
        items = await db[col].find({}, {"_id": 0, "id": 1, "created_at": 1, "name": 1, "vendor_name": 1, "venue_name": 1}).sort("created_at", -1).limit(3).to_list(3)
        for i in items:
            recent.append({
                "module": mod,
                "module_label": ORG_LABELS[mod],
                "id": i.get("id"),
                "name": i.get("name") or i.get("vendor_name") or i.get("venue_name") or "—",
                "created_at": i.get("created_at")
            })
    recent.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    recent = recent[:8]
    return {"counts": counts, "total_ratings": total_ratings, "top_rated": top_rated, "recent_additions": recent}

# ===== VENDOR CRUD (parameterized — declare LAST) =====

@api_router.get("/organization/{module}")
async def get_org_vendors(module: str, current_user: dict = Depends(get_current_user)):
    if module not in ORG_COLLECTIONS:
        raise HTTPException(status_code=404, detail="Modul tapılmadı")
    items = await db[ORG_COLLECTIONS[module]].find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    # Attach rating aggregation
    vendor_ids = [i["id"] for i in items]
    ratings_pipeline = [
        {"$match": {"vendor_type": module, "vendor_id": {"$in": vendor_ids}}},
        {"$group": {
            "_id": "$vendor_id",
            "count": {"$sum": 1},
            "avg_price": {"$avg": "$price_score"},
            "avg_quality": {"$avg": "$quality_score"},
            "avg_operativity": {"$avg": "$operativity_score"},
            "avg_behavior": {"$avg": "$behavior_score"},
            "last_used": {"$max": "$event_date"},
        }}
    ]
    ratings = {r["_id"]: r async for r in db.org_ratings.aggregate(ratings_pipeline)}
    for i in items:
        r = ratings.get(i["id"])
        if r:
            overall = round((r["avg_price"] + r["avg_quality"] + r["avg_operativity"] + r["avg_behavior"]) / 4, 2)
            i["rating_count"] = r["count"]
            i["rating_avg"] = overall
            i["rating_last_used"] = r["last_used"]
        else:
            i["rating_count"] = 0
            i["rating_avg"] = None
            i["rating_last_used"] = None
    return items

@api_router.post("/organization/{module}")
async def create_org_vendor(module: str, data: dict, current_user: dict = Depends(check_permission("organization", "write"))):
    if module not in ORG_COLLECTIONS:
        raise HTTPException(status_code=404, detail="Modul tapılmadı")
    doc = {**data}
    doc["id"] = str(uuid.uuid4())
    doc["module"] = module
    doc["created_by"] = current_user.get("name", "")
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db[ORG_COLLECTIONS[module]].insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/organization/{module}/{item_id}")
async def update_org_vendor(module: str, item_id: str, data: dict, current_user: dict = Depends(check_permission("organization", "write"))):
    if module not in ORG_COLLECTIONS:
        raise HTTPException(status_code=404, detail="Modul tapılmadı")
    update = {k: v for k, v in data.items() if k not in ("id", "created_at", "created_by", "module")}
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db[ORG_COLLECTIONS[module]].update_one({"id": item_id}, {"$set": update})
    doc = await db[ORG_COLLECTIONS[module]].find_one({"id": item_id}, {"_id": 0})
    return doc

@api_router.delete("/organization/{module}/{item_id}")
async def delete_org_vendor(module: str, item_id: str, current_user: dict = Depends(check_permission("organization", "write"))):
    if module not in ORG_COLLECTIONS:
        raise HTTPException(status_code=404, detail="Modul tapılmadı")
    await db[ORG_COLLECTIONS[module]].delete_one({"id": item_id})
    # Also cascade delete ratings for this vendor
    await db.org_ratings.delete_many({"vendor_type": module, "vendor_id": item_id})
    return {"message": "Silindi"}

# ==================== AI DATA ANALYST (GPT-5.2 via Emergent LLM Key) ====================

AI_SCHEMA = """
MongoDB database with these collections (all documents use `id` field, `_id` is excluded in responses):

1. companies — Marsol üzvü olan şirkətlər (active members)
   top-level fields: id, brand_name, legal_name, voen, sector, sub_sector, company_size,
   registration_date, address, region, status (Aktiv/Passiv/Ləğv edilib),
   company_phone, company_website, logo_url, social_links (array),
   owner_name (LEGACY, tez-tez boş), owner_first_name, owner_last_name, owner_phone, owner_email,
   representative_name, representative_phone, representative_email,
   contact_first_name, contact_last_name, contact_phone, contact_email, contact_position,
   marsol_representative, reference_source, reference_person_name, reference_person_surname,
   reference_person_position, reference_company_name, joined_project, join_date,
   package (Premium/Business/Business+/Sponsor),
   contract_start_date, contract_end_date, contract_file, contracts (array),
   total_amount, paid_amount, debt_amount, last_payment_date, payment_due_date,
   employee_count,
   children_count (LEGACY – tez-tez 0 və ya boş), children_info (LEGACY – demək olar ki, həmişə boş array),

   **owners** (ƏSAS MƏLUMAT BURDADIR) — array of owner objects, hər elementdə:
      first_name, last_name, father_name, position, phone, email, birth_date (YYYY-MM-DD),
      citizenship, education, specialty, university, social_links (array), desired_activities (array),
      **children** — array of child objects: { name, surname, birth_date (YYYY-MM-DD), gender ("Oğlan"/"Qız") }

   co_founders (tez-tez boş)

   **VACİB**: Sahibkar(lar)ın adı və uşaqları `owners` array-dadır, köhnə `owner_name`/`children_info` sahələrinə GÜVƏNMƏ.
   Sahibkar üçün adı göstərmək üçün: $concat: ["$owners.first_name", " ", "$owners.last_name"] ya da unwind sonrası. owners boş olsa fallback owner_name istifadə et.
   Doğum günü və yaş sorğuları üçün `owners.birth_date` istifadə et.
   Uşaqlar üçün sorğular `owners.children.birth_date` üzərindən gedir.

2. employees — Marsol daxili əməkdaşlar
   fields: id, full_name, first_name, last_name, birth_date, gender, department, position,
   company_phone, personal_phone, email, contract_start_date, contract_end_date, probation_end_date,
   gross_salary, net_salary, children_count, children_birth_dates (array of YYYY-MM-DD), status, marital_status, education_level

3. meetings — görüşlər
   fields: id, date, time, employee, meeting_setter, company, contact_person, project, meeting_type,
   location, result, next_meeting, department, meeting_mode, notes

4. tasks — tapşırıqlar
   fields: id, task_name, department, assignee, responsible_person, priority, start_date, end_date,
   related_object, phase, status, task_code (T-XXX)

5. project_events — Marsol-un təşkil etdiyi tədbirlər (Sərgi, Forum, İftar, Təlim, Tur, Networking, Konfrans)
   fields: id, name, type, date, end_date, location, status (Planlaşdırılır/Aktiv/Tamamlandı)

6. event_invitations — tədbirlərə dəvət olunan qonaqlar (non-members)
   fields: id, event_id, event_name, guest_name, guest_company, guest_position, guest_phone,
   guest_email, status (Dəvət edilib/Gələcəm/Gəlməyəcəm/İştirak etdi), invited_by, converted_to_lead

7. sales_leads — satış pipeline
   fields: id, lead_code, company_name, contact_name, contact_person, phone, email, position, source,
   sale_type, stage, status, assigned_to, curator, expected_amount, package, notes

8. contact_lists — sahibkar/CEO siyahıları
   fields: id, title, description, created_by

9. contacts — contact_lists-dəki kontaktlar
   fields: id, list_id, name, surname, company, position, phone, email

10. barters — barter əməliyyatları
    fields: id, barter_code (B-XXX), partner_name, partner_contact, partner_phone,
    our_service, their_service, our_value, their_value, status (Təklif/Müzakirədə/Aktiv/Tamamlandı/Ləğv edilib),
    start_date, end_date, responsible

11. incomes — gəlirlər (şirkət ödənişləri)
    fields: id, company_id, company_name, owner_name, project, package, amount, paid_amount,
    debt_amount, currency, contract_start_date, contract_end_date

12. expenses — xərclər
    fields: id, expense_name, category, sub_category, amount, currency, date, project, department,
    responsible_person, payment_type, status

13. attendance — əməkdaş davamiyyəti (günlük)
    fields: id, employee_id, employee_name, date (YYYY-MM-DD), status (İşdə/Gəlməyib/Məzuniyyət/Xəstəlik/İcazəli/Uzaq),
    check_in, check_out

14. leave_requests — məzuniyyət sorğuları
    fields: id, employee_id, employee_name, type (Məzuniyyət/Xəstəlik/İcazə/Digər),
    start_date, end_date, reason, status (Gözləyir/Təsdiqlənib/Rədd edilib)

15. users — sistem istifadəçiləri (auth)
    fields: id, name, email, role, department

16. assemblies — iclaslar
    fields: id, title, date, time, location, agenda (array), participants, status

17. invitations — KÖHNƏ sistem: üzv şirkətlərin fəaliyyətlərə dəvətləri (Obligations modulu üçün)
    fields: id, event_id, event_name, event_type, event_date, company_id, company_name,
    call_status, participation_status, obligation_deducted

18. events — Təşkilatçılıq modulu fəaliyyətləri
    fields: id, name, date, event_type, location

Dates are stored as ISO strings (YYYY-MM-DD) or ISO datetimes. Birth dates follow YYYY-MM-DD.
Current reference date: {today}
Age calculation: if someone should be older than N years, their birth_date should be BEFORE (currentYear - N)-01-01.
For "daha böyük/yuxarı yaş" use $lt (before threshold). For "kiçik yaş" use $gt (after threshold).
"""

AI_SYSTEM_PROMPT = """You are an AI data analyst for Marsol Group B2B networking ERP system.
You receive user questions in Azerbaijani about the data in their MongoDB database.

Your job: produce a single JSON response containing a MongoDB aggregation pipeline that answers the question.

Use ONLY these MongoDB operators in $match, $project, $group, $sort, $limit, $unwind, $addFields, $lookup, $count.
NEVER use: $out, $merge, $function, $where, $accumulator, $expr with $function, $redact.

Response format (strict JSON):
{
  "title": "Qısa Azərbaycanca başlıq (cədvəl başlığı)",
  "collection": "companies|employees|meetings|...",
  "pipeline": [ ...mongodb aggregation stages... ],
  "list_mapping": {
    "name": "<column header that contains person/contact name>",
    "company": "<column header with company name, or null>",
    "phone": "<column header with phone, or null>",
    "email": "<column header with email, or null>",
    "position": "<column header with position/title, or null>",
    "notes": "<column header with extra info, or null>"
  }
}

CRITICAL RULES:
- Final $project stage MUST produce dictionary with HUMAN-READABLE Azerbaijani field names as column headers
- Always add {"_id": 0} in $project (exclude _id)
- Order $project fields logically (name/company first, then details)
- Add $sort when meaningful
- Keep results reasonable (add $limit only if user asks for top-N)
- list_mapping maps logical fields (name/company/phone/email) to YOUR OUTPUT COLUMN HEADERS so the result can be saved into contact lists
- If the result is a statistical breakdown (not contact-style), set list_mapping fields to null

SCHEMA:
""" + AI_SCHEMA + """

Examples:

Q: "3 yaşdan yuxarı uşağı olan sahibkarlar" (today is 2026-02-19, so child born before 2023-02-19)
A: {
  "title": "3 yaşdan yuxarı uşağı olan sahibkarlar",
  "collection": "companies",
  "pipeline": [
    {"$match": {"owners.children.birth_date": {"$exists": true, "$ne": "", "$lt": "2023-02-19"}}},
    {"$unwind": {"path": "$owners", "preserveNullAndEmptyArrays": false}},
    {"$match": {"owners.children": {"$elemMatch": {"birth_date": {"$lt": "2023-02-19", "$ne": ""}}}}},
    {"$project": {
      "_id": 0,
      "Sahibkar": {"$trim": {"input": {"$concat": [{"$ifNull": ["$owners.first_name", ""]}, " ", {"$ifNull": ["$owners.last_name", ""]}]}}},
      "Şirkət": "$brand_name",
      "Telefon": {"$ifNull": ["$owners.phone", "$owner_phone"]},
      "Email": {"$ifNull": ["$owners.email", "$owner_email"]},
      "Uşaqlar": {
        "$reduce": {
          "input": {"$filter": {"input": "$owners.children", "as": "c", "cond": {"$and": [{"$ne": ["$$c.birth_date", ""]}, {"$lt": ["$$c.birth_date", "2023-02-19"]}]}}},
          "initialValue": "",
          "in": {"$concat": ["$$value", {"$cond": [{"$eq": ["$$value", ""]}, "", ", "]}, {"$ifNull": ["$$this.name", ""]}, " (", {"$ifNull": ["$$this.birth_date", ""]}, ")"]}
        }
      },
      "Paket": "$package"
    }},
    {"$sort": {"Sahibkar": 1}}
  ],
  "list_mapping": {"name": "Sahibkar", "company": "Şirkət", "phone": "Telefon", "email": "Email", "position": null, "notes": "Uşaqlar"}
}

Q: "iyun ayında doğum günü olan sahibkarlar"
A: {
  "title": "İyun ayında doğum günləri (sahibkarlar)",
  "collection": "companies",
  "pipeline": [
    {"$unwind": {"path": "$owners", "preserveNullAndEmptyArrays": false}},
    {"$match": {"owners.birth_date": {"$regex": "^\\\\d{4}-06-"}}},
    {"$project": {
      "_id": 0,
      "Sahibkar": {"$trim": {"input": {"$concat": [{"$ifNull": ["$owners.first_name", ""]}, " ", {"$ifNull": ["$owners.last_name", ""]}]}}},
      "Şirkət": "$brand_name",
      "Doğum tarixi": "$owners.birth_date",
      "Telefon": "$owners.phone",
      "Email": "$owners.email"
    }},
    {"$sort": {"Doğum tarixi": 1}}
  ],
  "list_mapping": {"name": "Sahibkar", "company": "Şirkət", "phone": "Telefon", "email": "Email", "position": null, "notes": "Doğum tarixi"}
}

Q: "hansı sektordan neçə şirkət var"
A: {
  "title": "Sektorlar üzrə şirkət sayı",
  "collection": "companies",
  "pipeline": [
    {"$match": {"status": "Aktiv"}},
    {"$group": {"_id": "$sector", "count": {"$sum": 1}}},
    {"$project": {"_id": 0, "Sektor": "$_id", "Şirkət sayı": "$count"}},
    {"$sort": {"Şirkət sayı": -1}}
  ],
  "list_mapping": {"name": null, "company": null, "phone": null, "email": null, "position": null, "notes": null}
}

Q: "neçə görüş etmişik, nəticələri necə olub"
A: {
  "title": "Görüşlər nəticə bölgüsü",
  "collection": "meetings",
  "pipeline": [
    {"$group": {"_id": "$result", "count": {"$sum": 1}}},
    {"$project": {"_id": 0, "Nəticə": "$_id", "Say": "$count"}},
    {"$sort": {"Say": -1}}
  ],
  "list_mapping": {"name": null, "company": null, "phone": null, "email": null, "position": null, "notes": null}
}

Q: "aktiv üzv şirkətlər ad və telefon ilə"
A: {
  "title": "Aktiv üzv şirkətlər",
  "collection": "companies",
  "pipeline": [
    {"$match": {"status": "Aktiv"}},
    {"$project": {
      "_id": 0,
      "Şirkət": "$brand_name",
      "Sahibkar": {"$cond": [
        {"$gt": [{"$size": {"$ifNull": ["$owners", []]}}, 0]},
        {"$trim": {"input": {"$concat": [{"$ifNull": [{"$arrayElemAt": ["$owners.first_name", 0]}, ""]}, " ", {"$ifNull": [{"$arrayElemAt": ["$owners.last_name", 0]}, ""]}]}}},
        "$owner_name"
      ]},
      "Telefon": {"$ifNull": [{"$arrayElemAt": ["$owners.phone", 0]}, "$owner_phone"]},
      "Email": {"$ifNull": [{"$arrayElemAt": ["$owners.email", 0]}, "$owner_email"]},
      "Paket": "$package"
    }},
    {"$sort": {"Şirkət": 1}}
  ],
  "list_mapping": {"name": "Sahibkar", "company": "Şirkət", "phone": "Telefon", "email": "Email", "position": null, "notes": "Paket"}
}

Return ONLY the JSON object, no markdown, no commentary.
"""

AI_ALLOWED_COLLECTIONS = {
    "companies", "employees", "meetings", "tasks", "project_events", "event_invitations",
    "sales_leads", "contact_lists", "contacts", "barters", "incomes", "expenses",
    "attendance", "leave_requests", "users", "assemblies", "invitations", "events", "roles"
}

AI_FORBIDDEN_OPERATORS = {"$out", "$merge", "$function", "$where", "$accumulator", "$redact"}

def _scan_forbidden_operators(obj) -> Optional[str]:
    """Recursively scan for forbidden MongoDB operators. Returns operator name if found."""
    if isinstance(obj, dict):
        for k, v in obj.items():
            if isinstance(k, str) and k in AI_FORBIDDEN_OPERATORS:
                return k
            found = _scan_forbidden_operators(v)
            if found:
                return found
    elif isinstance(obj, list):
        for item in obj:
            found = _scan_forbidden_operators(item)
            if found:
                return found
    return None

# =====================================================
# CLOUDINARY UPLOADS — files, company logos, employee avatars, project media
# =====================================================
MAX_UPLOAD_BYTES = 25 * 1024 * 1024  # 25 MB
ALLOWED_FOLDER_PREFIXES = (
    "marsol/files",       # general files module
    "marsol/companies",   # company logos
    "marsol/employees",   # employee avatars
    "marsol/projects",    # project media
)


@api_router.post("/uploads")
async def upload_asset(
    file: UploadFile = File(...),
    folder: str = Form("marsol/files"),
    current_user: dict = Depends(get_current_user),
):
    """Stream a file to Cloudinary and return its CDN url + public_id.

    Folders are restricted to the four whitelisted prefixes; the caller is
    responsible for storing the returned url/public_id on whatever entity
    consumes it (company, employee, file record, etc.).
    """
    folder = (folder or "marsol/files").strip()
    if not any(folder.startswith(p) for p in ALLOWED_FOLDER_PREFIXES):
        raise HTTPException(status_code=400, detail=f"İcazə verilməyən qovluq: {folder}")
    raw = await file.read()
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail=f"Fayl çox böyükdür (max {MAX_UPLOAD_BYTES // (1024*1024)} MB)")
    try:
        result = _cl_upload(raw, file.filename or "file", folder)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Yükləmə alınmadı: {str(e)[:200]}")
    return result


@api_router.delete("/uploads")
async def delete_uploaded_asset(
    public_id: str,
    resource_type: str = "image",
    current_user: dict = Depends(get_current_user),
):
    """Remove an asset from Cloudinary by public_id."""
    ok = _cl_delete(public_id, resource_type=resource_type)
    if not ok:
        raise HTTPException(status_code=404, detail="Asset tapılmadı və ya silinə bilmədi")
    return {"ok": True}


# =====================================================
# FILES MODULE — CRUD wrapping uploaded asset metadata
# =====================================================
@api_router.get("/files")
async def list_files(
    folder: Optional[str] = None,
    current_user: dict = Depends(check_permission("files", "read")),
):
    query: Dict[str, Any] = {}
    if folder:
        query["folder"] = folder
    # Apply scope: own/department/all. Files store `uploaded_by_name` (mapped via
    # SCOPE_FIELDS where we treat it as the owner field).
    user_name = current_user.get("name", "")
    if (current_user.get("role") or "").lower() != "admin":
        scopes = await get_user_scopes(current_user)
        sc = scopes.get("files", "own")
        if sc == "own":
            query["uploaded_by_name"] = user_name
        elif sc == "department":
            my_dept = (current_user.get("department") or "").strip()
            if my_dept:
                dept_users = await db.users.find({"department": my_dept}, {"_id": 0, "name": 1}).to_list(500)
                names = [u.get("name") for u in dept_users if u.get("name")] or [user_name]
                if user_name not in names:
                    names.append(user_name)
                query["uploaded_by_name"] = {"$in": names}
            else:
                query["uploaded_by_name"] = user_name
    files = await db.files.find(query, {"_id": 0}).sort("uploaded_at", -1).to_list(2000)
    return files


@api_router.post("/files")
async def create_file(data: dict, current_user: dict = Depends(check_permission("files", "write"))):
    if not data.get("url") or not data.get("public_id"):
        raise HTTPException(status_code=400, detail="url və public_id məcburidir")
    # Reject URLs that aren't from our Cloudinary cloud — prevents metadata spoofing.
    cloud = os.environ.get("CLOUDINARY_CLOUD_NAME", "")
    expected_prefix = f"https://res.cloudinary.com/{cloud}/" if cloud else "https://res.cloudinary.com/"
    if not str(data["url"]).startswith(expected_prefix):
        raise HTTPException(status_code=400, detail="Yanlış Cloudinary URL-i")
    doc = {
        "id": str(uuid.uuid4()),
        "name": data.get("name") or data.get("original_filename") or "Untitled",
        "description": data.get("description", ""),
        "tags": data.get("tags", []),
        "folder": data.get("folder", "marsol/files"),
        "url": data["url"],
        "public_id": data["public_id"],
        "resource_type": data.get("resource_type", "raw"),
        "format": data.get("format"),
        "bytes": data.get("bytes"),
        "mime_type": data.get("mime_type"),
        "width": data.get("width"),
        "height": data.get("height"),
        "uploaded_by_id": current_user.get("id"),
        "uploaded_by_name": current_user.get("name"),
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.files.insert_one(doc.copy())
    return doc


@api_router.put("/files/{file_id}")
async def update_file_metadata(file_id: str, data: dict, current_user: dict = Depends(check_permission("files", "write"))):
    """Update description, name or tags for a file. Cloudinary asset itself is untouched."""
    update: Dict[str, Any] = {}
    for k in ("description", "name"):
        if k in data:
            update[k] = (str(data.get(k) or "")).strip()
    if "tags" in data and isinstance(data["tags"], list):
        update["tags"] = data["tags"]
    if not update:
        raise HTTPException(status_code=400, detail="Yenilənəcək məlumat yoxdur")
    result = await db.files.update_one({"id": file_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Fayl tapılmadı")
    return await db.files.find_one({"id": file_id}, {"_id": 0})


@api_router.delete("/files/{file_id}")
async def delete_file_record(file_id: str, current_user: dict = Depends(check_permission("files", "write"))):
    rec = await db.files.find_one({"id": file_id}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="Fayl tapılmadı")
    _cl_delete(rec["public_id"], resource_type=rec.get("resource_type", "image"))
    await db.files.delete_one({"id": file_id})
    return {"ok": True}



@api_router.post("/ai/analyze")
async def ai_analyze(data: dict, current_user: dict = Depends(get_current_user)):
    """AI Data Analyst — takes Azerbaijani prompt, generates aggregation pipeline, executes, returns table."""
    prompt = (data.get("prompt") or "").strip()
    if not prompt:
        raise HTTPException(status_code=400, detail="Prompt tələb olunur")
    if len(prompt) > 2000:
        raise HTTPException(status_code=400, detail="Prompt çox uzundur (max 2000 simvol)")

    openai_key = os.environ.get("OPENAI_API_KEY")
    emergent_key = os.environ.get("EMERGENT_LLM_KEY")
    if not openai_key and not emergent_key:
        raise HTTPException(status_code=500, detail="OPENAI_API_KEY və ya EMERGENT_LLM_KEY mövcud deyil")

    import json as _json
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    system_prompt = AI_SYSTEM_PROMPT.replace("{today}", today_str)
    ai_model = os.environ.get("AI_ANALYST_MODEL", "gpt-4o-mini")

    ai_text: Optional[str] = None
    last_err: Optional[str] = None

    # Path 1: Direct OpenAI SDK (works on Render — no private deps)
    if openai_key:
        try:
            from openai import AsyncOpenAI
            client = AsyncOpenAI(api_key=openai_key)
            resp = await client.chat.completions.create(
                model=ai_model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": prompt},
                ],
                response_format={"type": "json_object"},
                temperature=0.2,
                timeout=60,
            )
            ai_text = resp.choices[0].message.content
        except Exception as e:
            last_err = f"OpenAI: {str(e)[:200]}"
            logging.warning("OpenAI AI Analyst failed, trying fallback: %s", e)

    # Path 2: Fallback to emergentintegrations (local pod only)
    if ai_text is None and emergent_key:
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage  # type: ignore
            session_id = f"ai-analyst-{current_user.get('id', 'unknown')}-{uuid.uuid4().hex[:8]}"
            ai_provider = os.environ.get("AI_ANALYST_PROVIDER", "anthropic")
            fallback_model = os.environ.get("AI_ANALYST_FALLBACK_MODEL", "claude-sonnet-4-5-20250929")
            chat = LlmChat(api_key=emergent_key, session_id=session_id, system_message=system_prompt).with_model(ai_provider, fallback_model)
            ai_text = await chat.send_message(UserMessage(text=prompt))
        except Exception as e:
            last_err = (last_err + " | " if last_err else "") + f"Emergent: {str(e)[:200]}"

    if ai_text is None:
        raise HTTPException(status_code=502, detail=f"AI cavab vermədi: {last_err or 'naməlum xəta'}")

    # Parse AI response
    cleaned = (ai_text or "").strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.strip("`")
        if cleaned.lower().startswith("json"):
            cleaned = cleaned[4:].strip()
    # Find first { and last }
    first_brace = cleaned.find("{")
    last_brace = cleaned.rfind("}")
    if first_brace == -1 or last_brace == -1:
        raise HTTPException(status_code=422, detail="AI cavabını parse etmək mümkün olmadı")
    try:
        plan = _json.loads(cleaned[first_brace:last_brace + 1])
    except _json.JSONDecodeError as e:
        raise HTTPException(status_code=422, detail=f"AI JSON xətası: {str(e)[:100]}")

    collection = plan.get("collection", "")
    pipeline = plan.get("pipeline", [])
    title = plan.get("title", "Nəticə")
    list_mapping = plan.get("list_mapping") or {}

    if collection not in AI_ALLOWED_COLLECTIONS:
        raise HTTPException(status_code=403, detail=f"Icazəsiz kolleksiya: {collection}")
    if not isinstance(pipeline, list) or len(pipeline) == 0 or len(pipeline) > 20:
        raise HTTPException(status_code=422, detail="Yanlış pipeline strukturu")

    forbidden = _scan_forbidden_operators(pipeline)
    if forbidden:
        raise HTTPException(status_code=403, detail=f"Qadağan edilmiş operator: {forbidden}")

    # Validate $lookup only targets allowed collections
    for stage in pipeline:
        if isinstance(stage, dict) and "$lookup" in stage:
            lookup = stage["$lookup"]
            if isinstance(lookup, dict):
                from_col = lookup.get("from", "")
                if from_col not in AI_ALLOWED_COLLECTIONS:
                    raise HTTPException(status_code=403, detail=f"$lookup qadağan: {from_col}")

    # Execute pipeline
    try:
        docs = await db[collection].aggregate(pipeline).to_list(10000)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Pipeline icra xətası: {str(e)[:200]}")

    # Sanitize: remove any stray _id
    for d in docs:
        d.pop("_id", None)

    # Extract headers (preserve order from first doc if possible)
    headers = []
    if docs:
        headers = list(docs[0].keys())
    # Convert rows to arrays following header order
    rows = [[d.get(h, "") for h in headers] for d in docs]

    return {
        "title": title,
        "headers": headers,
        "rows": rows,
        "row_count": len(rows),
        "collection": collection,
        "list_mapping": list_mapping,
        "prompt": prompt
    }

@api_router.post("/ai/save-to-list")
async def ai_save_to_list(data: dict, current_user: dict = Depends(check_permission("sales", "write"))):
    """Save AI analysis result rows to a contact list."""
    title = (data.get("title") or "").strip()
    description = (data.get("description") or "").strip()
    headers = data.get("headers") or []
    rows = data.get("rows") or []
    mapping = data.get("mapping") or {}

    if not title:
        raise HTTPException(status_code=400, detail="Siyahı başlığı tələb olunur")
    if not headers or not rows:
        raise HTTPException(status_code=400, detail="Məlumat yoxdur")

    # Create list
    list_doc = {
        "id": str(uuid.uuid4()),
        "title": title,
        "description": description or f"AI tərəfindən yaradıldı",
        "created_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.contact_lists.insert_one(list_doc)
    list_doc.pop("_id", None)

    # Column indices by header name
    def col_idx(header_name):
        if not header_name:
            return -1
        try:
            return headers.index(header_name)
        except ValueError:
            return -1

    idx_name = col_idx(mapping.get("name"))
    idx_company = col_idx(mapping.get("company"))
    idx_phone = col_idx(mapping.get("phone"))
    idx_email = col_idx(mapping.get("email"))
    idx_position = col_idx(mapping.get("position"))
    idx_notes = col_idx(mapping.get("notes"))

    def safe_get(row, i):
        if i < 0 or i >= len(row):
            return ""
        v = row[i]
        return str(v) if v is not None else ""

    contact_docs = []
    for row in rows:
        full_name = safe_get(row, idx_name)
        parts = full_name.split(" ", 1) if full_name else ["", ""]
        first = parts[0] if parts else ""
        last = parts[1] if len(parts) > 1 else ""
        contact_docs.append({
            "id": str(uuid.uuid4()),
            "list_id": list_doc["id"],
            "name": first,
            "surname": last,
            "company": safe_get(row, idx_company),
            "position": safe_get(row, idx_position),
            "phone": safe_get(row, idx_phone),
            "email": safe_get(row, idx_email),
            "notes": safe_get(row, idx_notes),
            "created_at": datetime.now(timezone.utc).isoformat()
        })

    if contact_docs:
        await db.contacts.insert_many(contact_docs)

    return {"message": f"{len(contact_docs)} kontakt əlavə edildi", "list_id": list_doc["id"], "list_title": list_doc["title"]}

@api_router.get("/ai/examples")
async def ai_examples(current_user: dict = Depends(get_current_user)):
    """Return example prompts for the AI Data Analyst UI."""
    return {"examples": [
        "5 yaş üzəri uşağı olan sahibkarların siyahısı",
        "İyun ayında doğum günü olan sahibkarlar",
        "Hansı sektordan neçə şirkət var",
        "Aktiv üzvlər (paket və müqavilə bitmə tarixi ilə)",
        "Bu ay keçirilmiş görüşlərin say və nəticələri",
        "Bitməkdə olan müqavilələr (30 gün içində)",
        "Borcu olan şirkətlər (borc məbləğinə görə sıralanmış)",
        "Tamamlanmamış tapşırıqlar",
        "Bu ay davamiyyət statistikası",
        "Aktiv barter əməliyyatları və balans",
        "İyul ayında doğum günü olan əməkdaşlar",
        "Hər paket üzrə üzv şirkət sayı",
        "Bakıda olan premium üzvlər",
    ]}


# ==================== NOTES MODULE ====================

NOTE_COLORS = ["#FFFFFF", "#FEF3C7", "#FED7AA", "#FECACA", "#DBEAFE", "#D1FAE5", "#E9D5FF", "#FBCFE8"]


@api_router.get("/notes")
async def list_notes(
    company_id: Optional[str] = None,
    tag: Optional[str] = None,
    pinned: Optional[bool] = None,
    q: Optional[str] = None,
    current_user: dict = Depends(check_permission("notes", "read")),
):
    """Personal Google-Keep style notes. Each user sees their own notes plus
    notes shared (via `shared_with_users` containing their id) or visible to all
    (`shared_with_all=True`). Admin sees everything."""
    user_id = current_user.get("id") or current_user.get("email", "")
    is_admin = current_user.get("role") == "admin"
    base = {} if is_admin else {
        "$or": [
            {"created_by_id": user_id},
            {"shared_with_users": user_id},
            {"shared_with_all": True},
        ]
    }
    extra = {}
    if company_id:
        extra["related_company_id"] = company_id
    if tag:
        extra["tags"] = tag
    if pinned is not None:
        extra["pinned"] = bool(pinned)
    if q:
        extra["$or"] = [
            {"title": {"$regex": q, "$options": "i"}},
            {"content": {"$regex": q, "$options": "i"}},
        ]
        # If user already had an $or for visibility, combine via $and
        if base:
            query = {"$and": [base, {"$or": extra["$or"]}, {k: v for k, v in extra.items() if k != "$or"}]}
        else:
            query = {**extra}
    else:
        query = {**base, **extra} if base else extra
    rows = await db.notes.find(query, {"_id": 0}).sort([("pinned", -1), ("updated_at", -1)]).to_list(2000)
    return rows


@api_router.post("/notes")
async def create_note(data: dict, current_user: dict = Depends(check_permission("notes", "write"))):
    user_id = current_user.get("id") or current_user.get("email", "")
    color = data.get("color") or NOTE_COLORS[1]
    if color not in NOTE_COLORS:
        color = NOTE_COLORS[1]
    doc = {
        "id": str(uuid.uuid4()),
        "title": (data.get("title") or "").strip(),
        "content": (data.get("content") or "").strip(),
        "color": color,
        "pinned": bool(data.get("pinned", False)),
        "tags": [t.strip() for t in (data.get("tags") or []) if str(t).strip()],
        "related_company_id": data.get("related_company_id") or "",
        "related_module": data.get("related_module") or "",
        "shared_with_all": bool(data.get("shared_with_all", False)),
        "shared_with_users": list(data.get("shared_with_users") or []),
        "created_by": current_user.get("name", ""),
        "created_by_id": user_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    if not doc["title"] and not doc["content"]:
        raise HTTPException(status_code=400, detail="Başlıq və ya məzmun lazımdır")
    await db.notes.insert_one(doc)
    doc.pop("_id", None)

    # Notify the explicitly-shared users
    target_ids = [uid for uid in doc["shared_with_users"] if uid and uid != user_id]
    if target_ids:
        recipients = await db.users.find({"id": {"$in": target_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(200)
        preview = (doc["title"] or doc["content"])[:160]
        push_recipients = []
        for r in recipients:
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()),
                "type": "note_shared",
                "title": f"Yeni qeyd sizinlə paylaşıldı — {doc['created_by']}",
                "body": preview,
                "note_id": doc["id"],
                "recipient_name": r.get("name", ""),
                "is_read": False,
                "created_at": doc["created_at"],
            })
            if r.get("name"):
                push_recipients.append(r["name"])
        if push_recipients:
            _safe_push(
                push_recipients,
                f"Yeni qeyd — {doc['created_by']}",
                preview,
                link=f"/notes?id={doc['id']}",
                data={"type": "note_shared", "note_id": doc["id"]},
            )
    return doc


async def _ensure_note_owner(note_id: str, current_user: dict):
    note = await db.notes.find_one({"id": note_id}, {"_id": 0})
    if not note:
        raise HTTPException(status_code=404, detail="Qeyd tapılmadı")
    user_id = current_user.get("id") or current_user.get("email", "")
    if current_user.get("role") != "admin" and note.get("created_by_id") != user_id:
        raise HTTPException(status_code=403, detail="Yalnız müəllif redaktə edə bilər")
    return note


@api_router.put("/notes/{note_id}")
async def update_note(note_id: str, data: dict, current_user: dict = Depends(check_permission("notes", "write"))):
    await _ensure_note_owner(note_id, current_user)
    update = {}
    for k in ("title", "content", "related_company_id", "related_module"):
        if k in data and data[k] is not None:
            update[k] = str(data[k]).strip() if isinstance(data[k], str) else data[k]
    if "color" in data and data["color"] in NOTE_COLORS:
        update["color"] = data["color"]
    if "pinned" in data:
        update["pinned"] = bool(data["pinned"])
    if "tags" in data and isinstance(data["tags"], list):
        update["tags"] = [str(t).strip() for t in data["tags"] if str(t).strip()]
    if "shared_with_all" in data:
        update["shared_with_all"] = bool(data["shared_with_all"])
    if "shared_with_users" in data and isinstance(data["shared_with_users"], list):
        update["shared_with_users"] = list(data["shared_with_users"])
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.notes.update_one({"id": note_id}, {"$set": update})
    doc = await db.notes.find_one({"id": note_id}, {"_id": 0})
    return doc


@api_router.delete("/notes/{note_id}")
async def delete_note(note_id: str, current_user: dict = Depends(check_permission("notes", "write"))):
    await _ensure_note_owner(note_id, current_user)
    await db.notes.delete_one({"id": note_id})
    return {"message": "Qeyd silindi"}


@api_router.get("/notes/tags")
async def list_note_tags(current_user: dict = Depends(check_permission("notes", "read"))):
    """Return distinct tags + colors used by the current user (or all for admin)."""
    user_id = current_user.get("id") or current_user.get("email", "")
    is_admin = current_user.get("role") == "admin"
    match = {} if is_admin else {
        "$or": [
            {"created_by_id": user_id},
            {"shared_with_users": user_id},
            {"shared_with_all": True},
        ]
    }
    pipeline = [
        {"$match": match},
        {"$unwind": "$tags"},
        {"$group": {"_id": "$tags", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    rows = await db.notes.aggregate(pipeline).to_list(200)
    return [{"tag": r["_id"], "count": r["count"]} for r in rows if r.get("_id")]


# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


async def _backfill_marsol_company_tenant():
    """One-time backfill so multi-tenant filtering doesn't accidentally hide
    existing data:

    1. Ensure at least one entry exists in `marsol_companies` (default "Marsol Group").
    2. Users with empty/None `marsol_company` → set to the default tenant name.
    3. Records in tenant-aware collections missing `marsol_company` → backfill
       with the default tenant name.
    """
    # 1. Default tenant
    first = await db.marsol_companies.find_one({}, {"_id": 0, "name": 1})
    if not first:
        default_doc = {
            "id": str(uuid.uuid4()),
            "name": "Marsol Group",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.marsol_companies.insert_one(default_doc)
        default_name = "Marsol Group"
    else:
        default_name = first.get("name") or "Marsol Group"

    # 2. Backfill users
    u_res = await db.users.update_many(
        {"$or": [{"marsol_company": {"$exists": False}}, {"marsol_company": None}, {"marsol_company": ""}]},
        {"$set": {"marsol_company": default_name}},
    )
    if u_res.modified_count:
        logging.info("Tenant backfill: %s users → '%s'", u_res.modified_count, default_name)

    # 3. Backfill records across tenant-aware collections
    collections = ["companies", "tasks", "meetings", "sales_leads", "assemblies", "project_events", "notes", "files"]
    for coll_name in collections:
        coll = db[coll_name]
        res = await coll.update_many(
            {"$or": [{"marsol_company": {"$exists": False}}, {"marsol_company": None}, {"marsol_company": ""}]},
            {"$set": {"marsol_company": default_name}},
        )
        if res.modified_count:
            logging.info("Tenant backfill: %s %s → '%s'", res.modified_count, coll_name, default_name)

    # 4. Smart re-backfill — repair legacy records whose tenant got mis-assigned
    # to the default. We look up the creator's actual tenant and override.
    # Idempotent: skips records flagged `tenant_backfill_v2: True`.
    await _smart_retenant_records(default_name)


async def _smart_retenant_records(default_name: str):
    """For each tenant-aware collection, find records whose `marsol_company` is
    still the global default but whose creator belongs to a DIFFERENT müəssisə.
    Update the record's tenant to match its creator. Idempotent via the flag
    `tenant_backfill_v2`.
    """
    # Build user name → tenant map (only users on non-default tenants — those
    # are the candidates whose legacy data may currently be hidden)
    user_tenant_map = {}
    cursor = db.users.find({"marsol_company": {"$nin": [None, "", default_name]}}, {"_id": 0, "name": 1, "marsol_company": 1})
    async for u in cursor:
        nm = (u.get("name") or "").strip()
        if nm:
            user_tenant_map[nm] = u.get("marsol_company")

    if not user_tenant_map:
        return

    creator_field_by_collection = {
        "tasks": ["created_by", "assignee", "responsible_person"],
        "meetings": ["created_by", "employee", "meeting_setter"],
        "companies": ["created_by", "curator"],
        "sales_leads": ["created_by", "curator"],
        "assemblies": ["created_by", "curator"],
        "project_events": ["created_by"],
        "notes": ["created_by"],
        "files": ["uploaded_by", "owner"],
    }

    total_repaired = 0
    for coll_name, fields in creator_field_by_collection.items():
        coll = db[coll_name]
        # Only candidates: currently default-tenant AND not yet repaired
        cursor = coll.find(
            {"marsol_company": default_name, "tenant_backfill_v2": {"$ne": True}},
            {"_id": 0, "id": 1, **{f: 1 for f in fields}},
        )
        async for rec in cursor:
            new_tenant = None
            for f in fields:
                v = rec.get(f)
                # assignee/responsible_person may be a list (multi-assignee)
                candidates = v if isinstance(v, list) else [v]
                for cand in candidates:
                    cand_name = (str(cand).strip() if cand else "")
                    if cand_name and cand_name in user_tenant_map:
                        new_tenant = user_tenant_map[cand_name]
                        break
                if new_tenant:
                    break
            update = {"tenant_backfill_v2": True}
            if new_tenant and new_tenant != default_name:
                update["marsol_company"] = new_tenant
            await coll.update_one({"id": rec.get("id")}, {"$set": update})
            if new_tenant and new_tenant != default_name:
                total_repaired += 1
        # Mark records that already had a non-default tenant as v2 too, to
        # avoid re-scanning on every restart.
        await coll.update_many(
            {"marsol_company": {"$ne": default_name}, "tenant_backfill_v2": {"$ne": True}},
            {"$set": {"tenant_backfill_v2": True}},
        )
    if total_repaired:
        logging.info("Tenant smart re-backfill: %s records re-assigned by creator tenant", total_repaired)


@app.on_event("startup")
async def startup_backfills():
    try:
        await _backfill_company_display_ids()
    except Exception as e:
        logging.warning("Company display_id backfill failed: %s", e)
    try:
        await _backfill_marsol_company_tenant()
    except Exception as e:
        logging.warning("Marsol company tenant backfill failed: %s", e)
    try:
        res = await _auto_archive_completed_tasks()
        if res.get("archived"):
            logging.info("Auto-archive at startup: %s completed tasks moved to archive", res.get("archived"))
    except Exception as e:
        logging.warning("Auto-archive at startup failed: %s", e)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
